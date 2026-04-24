import io
import json
import os
import traceback
import zipfile
from datetime import datetime

import openpyxl
from django.core.management.base import BaseCommand
from django.db import connections

from ownerclan.ownerclan_product_service import (
    EXCEL_COL_MAP, DB,
    _parse_excel_row, _safe_str, _field_changed,
    record_field_changes,
)

PROGRESS_INTERVAL = 1000


class Command(BaseCommand):
    help = '오너클랜 상품대장 비동기 업로드 워커'

    def add_arguments(self, parser):
        parser.add_argument('task_id', type=int)

    def handle(self, *args, **options):
        task_id = options['task_id']

        with connections[DB].cursor() as cur:
            cur.execute("SELECT input_data FROM lohas_task WHERE id=%s", [task_id])
            row = cur.fetchone()
            if not row:
                return
            input_data = row[0]
            if isinstance(input_data, str):
                input_data = json.loads(input_data)

            cur.execute(
                "UPDATE lohas_task SET status='running', pid=%s, updated_at=%s WHERE id=%s",
                [os.getpid(), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), task_id],
            )

        file_path = input_data.get('file_path', '')

        try:
            result = _process_upload(file_path, task_id)
            _update_task(task_id, 'done', result)
        except Exception:
            _update_task(task_id, 'error', {'error': traceback.format_exc()})
        finally:
            try:
                os.unlink(file_path)
            except Exception:
                pass


def _update_task(task_id, status, result_data):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with connections[DB].cursor() as cur:
        cur.execute(
            "UPDATE lohas_task SET status=%s, result_data=%s, updated_at=%s WHERE id=%s",
            [status, json.dumps(result_data), now, task_id],
        )


def _update_progress(task_id, data):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with connections[DB].cursor() as cur:
        cur.execute(
            "UPDATE lohas_task SET result_data=%s, updated_at=%s WHERE id=%s",
            [json.dumps(data), now, task_id],
        )


def _load_workbooks(file_path):
    name = file_path.lower()
    if name.endswith('.zip'):
        with open(file_path, 'rb') as f:
            zf = zipfile.ZipFile(io.BytesIO(f.read()))
        xlsx_names = sorted(n for n in zf.namelist() if n.lower().endswith('.xlsx'))
        if not xlsx_names:
            raise ValueError('ZIP 안에 .xlsx 파일이 없습니다.')
        wbs = []
        for xn in xlsx_names:
            xlsx_bytes = zf.read(xn)
            wbs.append((xn, openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)))
        return wbs
    elif name.endswith('.xlsx'):
        return [(os.path.basename(file_path), openpyxl.load_workbook(file_path, read_only=True))]
    else:
        raise ValueError('xlsx 또는 zip 파일만 업로드 가능합니다.')


def _process_upload(file_path, task_id):
    workbooks = _load_workbooks(file_path)
    now = datetime.now()

    inserted = 0
    updated = 0
    skipped = 0

    existing = {}
    with connections[DB].cursor() as cur:
        cur.execute("SELECT id, product_code FROM ownerclan_product")
        for row in cur.fetchall():
            existing[row[1]] = row[0]

    rows_to_process = []
    for wb_name, wb in workbooks:
        ws = wb.active
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 3:
                continue
            product_code = _safe_str(row[2])
            if not product_code:
                continue
            data = _parse_excel_row(list(row))
            rows_to_process.append((product_code, data))
        wb.close()

    total_rows = len(rows_to_process)
    _update_progress(task_id, {
        'progress': 0, 'inserted': 0, 'updated': 0,
        'skipped': 0, 'total_rows': total_rows,
    })

    fields = list(EXCEL_COL_MAP.values())
    orig_fields = [f'orig_{f}' for f in fields]

    with connections[DB].cursor() as cur:
        for idx, (product_code, data) in enumerate(rows_to_process, 1):
            if product_code in existing:
                pid = existing[product_code]
                cur.execute(
                    f"SELECT {', '.join(fields)} FROM ownerclan_product WHERE id=%s",
                    [pid],
                )
                old_row = cur.fetchone()
                old_data = dict(zip(fields, old_row))

                any_current_changed = False
                for f in fields:
                    if _field_changed(old_data[f], data[f], f):
                        any_current_changed = True
                        break

                if not any_current_changed:
                    skipped += 1
                else:
                    # 변경사항 로그 기록 (UPDATE 전에 old_data와 비교)
                    record_field_changes(pid, product_code, old_data, data)

                    set_parts = [f"{f}=%s" for f in fields]
                    set_parts.append("uploaded_at=%s")
                    vals = [data[f] for f in fields] + [now]

                    cur.execute(
                        f"UPDATE ownerclan_product SET {', '.join(set_parts)} WHERE id=%s",
                        vals + [pid],
                    )

                    cur.execute(
                        f"SELECT {', '.join(orig_fields)} FROM ownerclan_product WHERE id=%s",
                        [pid],
                    )
                    orig_row = cur.fetchone()
                    orig_data = dict(zip(fields, orig_row))

                    is_synced = 1
                    for f in fields:
                        if _field_changed(orig_data[f], data[f], f):
                            is_synced = 0
                            break

                    cur.execute(
                        "UPDATE ownerclan_product SET is_synced=%s WHERE id=%s",
                        [is_synced, pid],
                    )
                    updated += 1
            else:
                all_fields = ['product_code'] + fields + orig_fields + [
                    'sale_status', 'is_synced', 'uploaded_at',
                ]
                placeholders = ', '.join(['%s'] * len(all_fields))
                vals = (
                    [product_code]
                    + [data[f] for f in fields]
                    + [data[f] for f in fields]
                    + [1, 1, now]
                )
                cur.execute(
                    f"INSERT INTO ownerclan_product ({', '.join(all_fields)}) "
                    f"VALUES ({placeholders})",
                    vals,
                )
                existing[product_code] = cur.lastrowid
                inserted += 1

            if idx % PROGRESS_INTERVAL == 0 or idx == total_rows:
                progress = int(idx * 100 / total_rows) if total_rows else 100
                _update_progress(task_id, {
                    'progress': progress,
                    'inserted': inserted,
                    'updated': updated,
                    'skipped': skipped,
                    'total_rows': total_rows,
                })

    # 이번 업로드에 포함된 상품 중 판매중지(sale_status!=1)인 것 파악
    uploaded_codes = [pc for pc, _ in rows_to_process]
    suspended_count = 0
    if uploaded_codes:
        batch_size = 500
        for i in range(0, len(uploaded_codes), batch_size):
            batch = uploaded_codes[i:i + batch_size]
            placeholders = ','.join(['%s'] * len(batch))
            with connections[DB].cursor() as cur:
                cur.execute(
                    f"SELECT COUNT(*) FROM ownerclan_product "
                    f"WHERE product_code IN ({placeholders}) AND sale_status != 1",
                    batch,
                )
                suspended_count += cur.fetchone()[0]

    return {
        'inserted': inserted,
        'updated': updated,
        'skipped': skipped,
        'total': inserted + updated + skipped,
        'suspended_count': suspended_count,
    }
