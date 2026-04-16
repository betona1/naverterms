import io

from django.db import connections
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response

from . import ownerclan_product_service


class OwnerClanProductUploadView(APIView):
    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'file required'}, status=400)

        running = ownerclan_product_service.check_running_task()
        if running:
            return Response({
                'error': '이미 업로드 처리 중입니다.',
                'task_id': running,
            }, status=409)

        try:
            result = ownerclan_product_service.upload_excel_async(f)
            return Response(result, status=202)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    def get(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=400)
        task = ownerclan_product_service.get_upload_task(int(task_id))
        if not task:
            return Response({'error': 'not found'}, status=404)
        return Response({
            'task_id': task['id'],
            'status': task['status'],
            'result_data': task['result_data'],
        })


class OwnerClanProductCsvUploadView(APIView):
    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'file required'}, status=400)
        try:
            result = ownerclan_product_service.upload_csv_status(f)
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class OwnerClanSoldoutTxtUploadView(APIView):
    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'file required'}, status=400)
        try:
            result = ownerclan_product_service.upload_soldout_txt(f)
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class OwnerClanProductListView(APIView):
    def get(self, request):
        page = int(request.query_params.get('page', 1))
        per_page = int(request.query_params.get('per_page', 50))
        sale_status = request.query_params.get('sale_status')
        is_synced = request.query_params.get('is_synced')
        search = request.query_params.get('search') or None
        changed_field = request.query_params.get('changed_field') or None
        result = ownerclan_product_service.get_products(
            page, per_page,
            sale_status=int(sale_status) if sale_status else None,
            is_synced=int(is_synced) if is_synced is not None and is_synced != '' else None,
            search=search,
            changed_field=changed_field,
        )
        return Response(result)


class OwnerClanProductDetailView(APIView):
    def get(self, request, pk):
        result = ownerclan_product_service.get_product_detail(pk)
        if not result:
            return Response({'error': '상품을 찾을 수 없습니다.'}, status=404)
        return Response(result)


class OwnerClanProductActivateView(APIView):
    def post(self, request):
        task_id = request.data.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=400)
        result = ownerclan_product_service.activate_suspended_from_task(int(task_id))
        return Response(result)


class OwnerClanProductSyncView(APIView):
    def post(self, request):
        product_ids = request.data.get('product_ids')
        if product_ids and isinstance(product_ids, list):
            product_ids = [int(i) for i in product_ids]
        else:
            product_ids = None
        result = ownerclan_product_service.sync_products(product_ids)
        return Response(result)


class OwnerClanProductStatsView(APIView):
    def get(self, request):
        return Response(ownerclan_product_service.get_stats())


class OwnerClanProductChangedFieldsView(APIView):
    def get(self, request):
        return Response(ownerclan_product_service.get_changed_field_counts())


class OwnerClanProductExcelExportView(APIView):
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        sale_status = request.query_params.get('sale_status')
        is_synced = request.query_params.get('is_synced')
        search = request.query_params.get('search') or None
        changed_field = request.query_params.get('changed_field') or None

        rows = ownerclan_product_service.get_products_for_export(
            sale_status=int(sale_status) if sale_status else None,
            is_synced=int(is_synced) if is_synced is not None and is_synced != '' else None,
            search=search,
            changed_field=changed_field,
        )

        STATUS_LABELS = {1: '판매중', 2: '품절', 3: '단종'}

        wb = Workbook()
        ws = wb.active
        ws.title = '오너클랜 상품대장'

        headers = ['W코드', '상태', '동기화', '상품명', '원본상품명',
                    '마켓상품명', '원본마켓상품명', '오너클랜가', '원본오너클랜가',
                    '마켓가', '원본마켓가', '배송비', '원본배송비',
                    '반품비', '원본반품비', '카테고리', '제조사', '원산지']
        col_widths = [12, 8, 8, 35, 35, 35, 35, 12, 12, 12, 12, 8, 8, 8, 8, 20, 15, 10]

        header_font = Font(bold=True, size=10)
        header_fill = PatternFill('solid', fgColor='F0F0F0')
        changed_fill = PatternFill('solid', fgColor='FFF3E0')
        thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))
        money_fmt = '#,##0'

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        for i, r in enumerate(rows, 2):
            is_changed = r.get('is_synced') == 0
            fill = changed_fill if is_changed else None

            def _cell(col, val, fmt=None):
                c = ws.cell(row=i, column=col, value=val)
                c.border = thin_border
                if fmt:
                    c.number_format = fmt
                if fill:
                    c.fill = fill
                return c

            _cell(1, r.get('product_code'))
            _cell(2, STATUS_LABELS.get(r.get('sale_status'), '?'))
            _cell(3, '변경됨' if is_changed else '일치')
            _cell(4, r.get('product_name'))
            _cell(5, r.get('orig_product_name'))
            _cell(6, r.get('market_product_name'))
            _cell(7, r.get('orig_market_product_name'))
            _cell(8, r.get('ownerclan_price', 0), money_fmt)
            _cell(9, r.get('orig_ownerclan_price', 0), money_fmt)
            _cell(10, r.get('market_price', 0), money_fmt)
            _cell(11, r.get('orig_market_price', 0), money_fmt)
            _cell(12, r.get('shipping_fee', 0), money_fmt)
            _cell(13, r.get('orig_shipping_fee', 0), money_fmt)
            _cell(14, r.get('return_fee', 0), money_fmt)
            _cell(15, r.get('orig_return_fee', 0), money_fmt)
            _cell(16, r.get('category_name'))
            _cell(17, r.get('manufacturer'))
            _cell(18, r.get('origin'))

        ws.auto_filter.ref = ws.dimensions

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="ownerclan_products.xlsx"'
        return response


class OwnerClanProductSampleExcelView(APIView):
    """DB에서 5건을 가져와 업로드 형식(이셀러스/오너클랜 상품대장) 샘플 엑셀을 생성"""
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        COL_MAP = ownerclan_product_service.EXCEL_COL_MAP
        fields = ['product_code'] + list(COL_MAP.values())

        with connections['ads'].cursor() as cur:
            cur.execute(
                f"SELECT {', '.join(fields)} FROM ownerclan_product "
                "WHERE sale_status=1 ORDER BY id LIMIT 5"
            )
            rows = cur.fetchall()

        HEADER_LABELS = {
            'seller_code1': '셀러코드1', 'seller_code2': '셀러코드2',
            'product_code': 'W코드',
            'category_code': '카테고리코드', 'category_name': '카테고리명',
            'market_category': '마켓카테고리',
            'product_name': '상품명', 'market_product_name': '마켓상품명',
            'ownerclan_price': '오너클랜가', 'consumer_price': '소비자가',
            'market_price': '마켓가', 'shipping_fee': '배송비',
            'shipping_type': '배송타입', 'min_qty': '최소수량', 'max_qty': '최대수량',
            'company_notice': '업체공지', 'special_notice': '특이사항',
            'option1_name': '옵션1명', 'option1_values': '옵션1값',
            'option2_name': '옵션2명', 'option2_values': '옵션2값',
            'combined_option': '조합옵션', 'product_attribute': '상품속성',
            'product_grade': '상품등급', 'tax_type': '과세유형',
            'compliance': '인증정보', 'age_restriction': '연령제한',
            'return_possible': '반품가능여부',
            'image_large': '대표이미지(대)', 'image_medium': '대표이미지(중)',
            'image_small': '대표이미지(소)',
            'manufacturer': '제조사', 'brand': '브랜드', 'model_name': '모델명',
            'origin': '원산지', 'keywords': '검색키워드',
            'registered_at': '등록일', 'modified_at': '수정일',
            'header_text': '상단문구', 'detail_html': '상세설명HTML',
            'notice_code': '고시코드', 'notice_category': '고시카테고리',
            'notice_info': '고시정보', 'notice_html': '고시HTML',
            'market_gmarket': '지마켓', 'market_auction': '옥션',
            'market_11st': '11번가', 'market_coupang': '쿠팡',
            'market_smartstore': '스마트스토어', 'market_promo': '프로모션',
            'market_gift': '사은품',
            'certification_type': '인증유형', 'certification_info': '인증상세',
            'return_fee': '반품비',
            'independent_option': '독립옵션', 'combined_option_detail': '조합옵션상세',
        }

        # 엑셀 컬럼 순서: seller_code1, seller_code2, product_code, category_code, ...
        ordered_fields = ['seller_code1', 'seller_code2', 'product_code'] + \
                         [COL_MAP[i] for i in sorted(COL_MAP.keys()) if i >= 3]

        wb = Workbook()
        ws = wb.active
        ws.title = '오너클랜 상품대장 샘플'

        header_font = Font(bold=True, size=10)
        header_fill = PatternFill('solid', fgColor='E8F5E9')
        guide_font = Font(size=9, color='888888', italic=True)
        thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))

        # Row 1: 안내문
        ws.cell(row=1, column=1, value='[업로드 형식 샘플] 3행부터 데이터. 이 파일을 참고하여 이셀러스/오너클랜에서 다운로드한 상품대장을 업로드하세요.')
        ws.cell(row=1, column=1).font = Font(bold=True, size=11, color='D32F2F')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

        # Row 2: 컬럼 헤더
        for col_idx, f in enumerate(ordered_fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=HEADER_LABELS.get(f, f))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # 컬럼 너비
        for col_idx in range(1, len(ordered_fields) + 1):
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = 18

        # Row 3~: 데이터
        field_to_idx = {f: i for i, f in enumerate(fields)}
        for row_num, db_row in enumerate(rows, 3):
            for col_idx, f in enumerate(ordered_fields, 1):
                val = db_row[field_to_idx[f]] if f in field_to_idx else ''
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = thin_border

        ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(ordered_fields)).column_letter}{2 + len(rows)}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="ownerclan_upload_sample.xlsx"'
        return response


class OwnerClanProductWCodesView(APIView):
    def get(self, request):
        sale_status = request.query_params.get('sale_status')
        is_synced = request.query_params.get('is_synced')
        search = request.query_params.get('search') or None
        changed_field = request.query_params.get('changed_field') or None

        codes = ownerclan_product_service.get_w_codes(
            sale_status=int(sale_status) if sale_status else None,
            is_synced=int(is_synced) if is_synced is not None and is_synced != '' else None,
            search=search,
            changed_field=changed_field,
        )
        return Response({'codes': codes, 'count': len(codes)})
