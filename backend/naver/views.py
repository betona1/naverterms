from collections import defaultdict
from datetime import timedelta
from django.utils import timezone
from django.db.models import Max, Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import (
    NaverKeyword, NaverSearchSnapshot, NaverTermAnalysis,
    NaverRankTarget, NaverRankHistory, NaverTrackingSchedule,
    NaverReport, NaverCrawlLog,
    NaverPurchaseTarget, NaverPurchaseHistory,
    NaverSynonym,
)
from .serializers import (
    NaverKeywordSerializer, NaverTermAnalysisSerializer,
    NaverRankTargetSerializer, NaverRankHistorySerializer,
    NaverTrackingScheduleSerializer,
    NaverPurchaseTargetSerializer, NaverPurchaseHistorySerializer,
    NaverSynonymSerializer,
)
from . import services
from . import uc_crawler


# ──── 키워드 ────

class KeywordListView(APIView):
    def get(self, request):
        # 순위추적 전용 키워드만 제외 (스냅샷/분석 없이 순위대상만 있는 키워드)
        keywords = NaverKeyword.objects.exclude(
            snapshots__isnull=True, analyses__isnull=True, rank_targets__isnull=False
        ).distinct().order_by('-created_at')
        return Response(NaverKeywordSerializer(keywords, many=True).data)

    def post(self, request):
        keyword_text = request.data.get('keyword', '').strip()
        if not keyword_text:
            return Response({'error': '키워드를 입력하세요'}, status=400)
        kw, created = NaverKeyword.objects.get_or_create(keyword=keyword_text)
        return Response(NaverKeywordSerializer(kw).data, status=201 if created else 200)


class KeywordDetailView(APIView):
    def delete(self, request, pk):
        try:
            NaverKeyword.objects.get(id=pk).delete()
            return Response(status=204)
        except NaverKeyword.DoesNotExist:
            return Response({'error': 'not found'}, status=404)


class KeywordFavoriteView(APIView):
    """즐겨찾기 토글 (PATCH). body 없이 호출하면 flip, {"is_favorite": true/false} 로 명시 가능"""
    def patch(self, request, pk):
        try:
            kw = NaverKeyword.objects.get(id=pk)
        except NaverKeyword.DoesNotExist:
            return Response({'error': 'not found'}, status=404)
        want = request.data.get('is_favorite')
        kw.is_favorite = (not kw.is_favorite) if want is None else bool(want)
        kw.save(update_fields=['is_favorite'])
        return Response({'id': kw.id, 'is_favorite': kw.is_favorite})


class CrawlLogView(APIView):
    """수집 로그 조회/추가/전체삭제"""
    def get(self, request):
        limit = int(request.query_params.get('limit', 200))
        since_id = request.query_params.get('since_id')
        qs = NaverCrawlLog.objects.all()
        if since_id:
            qs = qs.filter(id__gt=int(since_id)).order_by('id')
        else:
            qs = qs.order_by('-id')[:limit]
        return Response({'logs': [
            {
                'id': l.id,
                'timestamp': l.timestamp.isoformat(),
                'type': l.type,
                'message': l.message,
                'keyword': l.keyword,
                'session_id': l.session_id,
            } for l in qs
        ]})

    def post(self, request):
        body = request.data
        items = body if isinstance(body, list) else [body]
        created_ids = []
        for item in items:
            log = NaverCrawlLog.objects.create(
                type=(item.get('type') or 'info')[:20],
                message=str(item.get('message', '')),
                keyword=str(item.get('keyword', ''))[:200],
                session_id=str(item.get('session_id', ''))[:64],
            )
            created_ids.append(log.id)
        return Response({'count': len(items), 'ids': created_ids}, status=201)

    def delete(self, request):
        n, _ = NaverCrawlLog.objects.all().delete()
        return Response({'deleted': n})


class ResultsKeywordListView(APIView):
    """결과보기 전용 — 모든 키워드 + 그룹별 분류 + 탭별 수집 상태 + 최근 업데이트 순

    응답 구조:
    {
      "keywords": [
        {
          "id", "keyword", "is_favorite", "last_searched_at",
          "total_count", "term_count",
          "collected": { "total": {count, total, collected_at}, "model": {...}, "checkout": {...} },
          "has_data": bool
        }, ...
      ]
    }
    """
    def get(self, request):
        qs = NaverKeyword.objects.exclude(
            snapshots__isnull=True, analyses__isnull=True, rank_targets__isnull=False
        ).distinct()
        # 최근 업데이트 순 — last_searched_at 우선, 없으면 created_at
        from django.db.models.functions import Coalesce
        qs = qs.annotate(
            sort_key=Coalesce('last_searched_at', 'created_at')
        ).order_by('-sort_key')

        kw_ids = list(qs.values_list('id', flat=True))
        # 각 키워드의 탭별 최신 스냅샷 한번에 조회
        snap_map = defaultdict(dict)  # {kw_id: {tab_type: snapshot}}
        for snap in (NaverSearchSnapshot.objects
                     .filter(keyword_id__in=kw_ids)
                     .order_by('keyword_id', 'tab_type', '-collected_at')):
            if snap.tab_type not in snap_map[snap.keyword_id]:
                snap_map[snap.keyword_id][snap.tab_type] = snap

        results = []
        for kw in qs:
            collected = {}
            tab_snaps = snap_map.get(kw.id, {})
            for tab_type in ('total', 'model', 'checkout'):
                s = tab_snaps.get(tab_type)
                if s:
                    collected[tab_type] = {
                        'count': len(s.products) if s.products else 0,
                        'total': s.total,
                        'collected_at': s.collected_at,
                    }
            has_data = any(v.get('count', 0) > 0 for v in collected.values())
            results.append({
                'id': kw.id,
                'keyword': kw.keyword,
                'is_favorite': kw.is_favorite,
                'last_searched_at': kw.last_searched_at,
                'created_at': kw.created_at,
                'total_count': kw.total_count,
                'term_count': kw.term_count,
                'terms': kw.terms or [],
                'collected': collected,
                'has_data': has_data,
            })
        return Response({'keywords': results})


class SnapshotListView(APIView):
    """키워드의 시간순 스냅샷 메타 (시계열 슬라이더용)
    GET /api/naver/snapshots/<kw_id>/?tab=total&limit=20
    """
    def get(self, request, keyword_id):
        tab = request.query_params.get('tab', 'total')
        limit = min(int(request.query_params.get('limit', 30)), 100)
        qs = (NaverSearchSnapshot.objects
              .filter(keyword_id=keyword_id, tab_type=tab)
              .order_by('-collected_at')[:limit])
        return Response({'snapshots': [
            {
                'id': s.id,
                'collected_at': s.collected_at.isoformat(),
                'tab_type': s.tab_type,
                'total': s.total,
                'product_count': len(s.products) if s.products else 0,
            } for s in qs
        ]})


class ProductHistoryView(APIView):
    """특정 상품(id/nvMid) 의 시계열 변동
    GET /api/naver/product-history/<kw_id>/?product_id=12345&tab=total&limit=20
    응답: 시간순(오래된→최근)으로 변동 행 + 직전 대비 변화 플래그
    """
    def get(self, request, keyword_id):
        tab = request.query_params.get('tab', 'total')
        product_id = (request.query_params.get('product_id') or '').strip()
        limit = min(int(request.query_params.get('limit', 30)), 60)
        if not product_id:
            return Response({'error': 'product_id 필요'}, status=400)

        snaps = (NaverSearchSnapshot.objects
                 .filter(keyword_id=keyword_id, tab_type=tab)
                 .order_by('-collected_at')[:limit])
        # 시간순 (오래된 → 최근)
        snaps = list(snaps)[::-1]

        history = []
        for s in snaps:
            for i, p in enumerate(s.products or []):
                pid = str(p.get('nvMid') or p.get('id') or '')
                if pid == str(product_id):
                    history.append({
                        'snapshot_id': s.id,
                        'collected_at': s.collected_at.isoformat(),
                        'rank': i + 1,
                        'productName': p.get('productName') or p.get('productTitle') or '',
                        'mallName': p.get('mallName') or '',
                        'lowPrice': p.get('lowPrice') or p.get('price') or 0,
                        'reviewCount': p.get('reviewCount') or 0,
                        'imageUrl': p.get('imageUrl') or '',
                    })
                    break  # 같은 스냅샷 안에 같은 상품은 1번만

        # 직전 대비 변동 계산
        for idx, h in enumerate(history):
            if idx == 0:
                h['delta'] = {'rank': 0, 'price': 0, 'name_changed': False, 'image_changed': False}
            else:
                prev = history[idx - 1]
                p_now = int(h['lowPrice']) if h['lowPrice'] else 0
                p_prev = int(prev['lowPrice']) if prev['lowPrice'] else 0
                h['delta'] = {
                    'rank': prev['rank'] - h['rank'],  # 양수면 순위 상승
                    'price': p_now - p_prev,
                    'name_changed': prev['productName'] != h['productName'],
                    'image_changed': prev['imageUrl'] != h['imageUrl'],
                }

        return Response({
            'product_id': product_id,
            'tab': tab,
            'history': history,
            'count': len(history),
        })


# ──── 확장프로그램 → 검색결과 저장 ────

class ExtSearchResultView(APIView):
    def post(self, request):
        keyword_text = request.data.get('keyword', '').strip()
        tab_type = request.data.get('tab_type', 'total')
        products = request.data.get('products', [])
        total = request.data.get('total', 0)
        terms = request.data.get('terms', [])
        term_count = request.data.get('term_count', 0)

        kw, _ = NaverKeyword.objects.get_or_create(keyword=keyword_text)
        kw.last_searched_at = timezone.now()

        if terms:
            kw.terms = terms
            kw.term_count = term_count or len(terms)

        if tab_type == 'total':
            kw.total_count = total
        elif tab_type == 'checkout':
            kw.naverpay_count = total
        elif tab_type == 'model':
            kw.price_compare_count = total
        kw.save()

        NaverSearchSnapshot.objects.create(
            keyword=kw,
            tab_type=tab_type,
            products=products,
            total=total,
        )

        return Response({'status': 'ok', 'keyword_id': kw.id})


class ExtCaptchaStatusView(APIView):
    def post(self, request):
        captcha_type = request.data.get('type', 'unknown')
        resolved = request.data.get('resolved', False)
        return Response({'status': 'ok', 'type': captcha_type, 'resolved': resolved})


# ──── 가중치 분석 ────

class AnalysisView(APIView):
    def get(self, request, keyword_id):
        analyses = NaverTermAnalysis.objects.filter(keyword_id=keyword_id).order_by('-analyzed_at')
        return Response(NaverTermAnalysisSerializer(analyses, many=True).data)

    def post(self, request, keyword_id):
        analysis = services.run_full_analysis(keyword_id)
        if not analysis:
            return Response({'error': '분석할 데이터가 없습니다'}, status=400)
        return Response(NaverTermAnalysisSerializer(analysis).data)


# ──── 상품 데이터 조회 ────

class ProductsView(APIView):
    def get(self, request, keyword_id):
        tab = request.query_params.get('tab', 'total')
        snapshot = NaverSearchSnapshot.objects.filter(
            keyword_id=keyword_id, tab_type=tab
        ).order_by('-collected_at').first()
        if not snapshot:
            return Response({'products': [], 'total': 0})
        return Response({'products': snapshot.products, 'total': snapshot.total, 'collected_at': snapshot.collected_at})


# ──── 태그 통계 ────

class TagStatsView(APIView):
    def get(self, request, keyword_id):
        stats = services.get_tag_statistics(keyword_id)
        return Response(stats)


# ──── 순위추적 대상 ────

class RankTargetListView(APIView):
    def get(self, request):
        targets = NaverRankTarget.objects.filter(is_active=True).select_related('keyword')
        return Response(NaverRankTargetSerializer(targets, many=True).data)

    def post(self, request):
        keyword_text = request.data.get('keyword', '').strip()
        target_type = request.data.get('target_type', 'store')
        target_value = request.data.get('target_value', '').strip()
        display_name = request.data.get('display_name', '').strip()
        source_product_id = request.data.get('source_product_id')
        source_product_name = request.data.get('source_product_name', '').strip()

        if not keyword_text or not target_value:
            return Response({'error': '키워드와 대상을 입력하세요'}, status=400)

        kw, _ = NaverKeyword.objects.get_or_create(keyword=keyword_text)
        defaults = {'display_name': display_name or target_value}
        if source_product_id is not None:
            defaults['source_product_id'] = source_product_id
        if source_product_name:
            defaults['source_product_name'] = source_product_name

        target, created = NaverRankTarget.objects.get_or_create(
            keyword=kw, target_type=target_type, target_value=target_value,
            defaults=defaults,
        )
        return Response(NaverRankTargetSerializer(target).data, status=201 if created else 200)


class RankTargetDetailView(APIView):
    def delete(self, request, pk):
        try:
            NaverRankTarget.objects.get(id=pk).delete()
            return Response(status=204)
        except NaverRankTarget.DoesNotExist:
            return Response({'error': 'not found'}, status=404)


# ──── 확장프로그램 → 순위결과 저장 ────

class ExtRankResultView(APIView):
    def post(self, request):
        target_id = request.data.get('target_id')
        rank_position = request.data.get('rank_position')  # None = 미발견
        tab_type = request.data.get('tab_type', 'total')
        total_results = request.data.get('total_results', 0)
        found_product_name = request.data.get('found_product_name', '')
        found_product_price = request.data.get('found_product_price')
        found_review_count = request.data.get('found_review_count')

        try:
            target = NaverRankTarget.objects.get(id=target_id)
        except NaverRankTarget.DoesNotExist:
            return Response({'error': 'target not found'}, status=404)

        history = NaverRankHistory.objects.create(
            target=target,
            rank_position=rank_position,
            tab_type=tab_type,
            total_results=total_results,
            found_product_name=found_product_name,
            found_product_price=found_product_price,
            found_review_count=found_review_count,
        )
        return Response(NaverRankHistorySerializer(history).data, status=201)


# ──── 순위추적 실행 (네이버 API) ────

class RunRankTrackingView(APIView):
    def post(self, request):
        target_ids = request.data.get('target_ids')  # None = 전체
        try:
            result = services.run_rank_tracking(target_ids)
            return Response(result)
        except ValueError as e:
            return Response({'error': str(e)}, status=400)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ──── 순위 이력 ────

class RankHistoryView(APIView):
    def get(self, request):
        target_id = request.query_params.get('target_id')
        days = int(request.query_params.get('days', 30))
        since = timezone.now() - timedelta(days=days)

        qs = NaverRankHistory.objects.filter(tracked_at__gte=since)
        if target_id:
            qs = qs.filter(target_id=target_id)
        qs = qs.order_by('-tracked_at')
        return Response(NaverRankHistorySerializer(qs, many=True).data)


class RankSummaryView(APIView):
    def get(self, request):
        targets = NaverRankTarget.objects.filter(is_active=True).select_related('keyword')
        results = []
        for target in targets:
            records = list(target.history.order_by('-tracked_at')[:2])
            current = records[0].rank_position if records else None
            previous = records[1].rank_position if len(records) > 1 else None
            change = (previous - current) if (current and previous) else None
            results.append({
                'id': target.id,
                'keyword': target.keyword.keyword,
                'target_type': target.target_type,
                'target_value': target.target_value,
                'display_name': target.display_name,
                'current_rank': current,
                'previous_rank': previous,
                'change': change,
                'tracked_at': records[0].tracked_at.isoformat() if records else None,
            })
        return Response(results)


# ──── 스케줄 ────

class ScheduleListView(APIView):
    def get(self, request):
        schedules = NaverTrackingSchedule.objects.all().order_by('-created_at')
        result = []
        for s in schedules:
            data = NaverTrackingScheduleSerializer(s).data
            # 타겟 상세 정보 추가
            targets = NaverRankTarget.objects.filter(id__in=s.target_ids).select_related('keyword')
            data['targets'] = []
            for t in targets:
                latest = t.history.order_by('-tracked_at').first()
                data['targets'].append({
                    'id': t.id,
                    'keyword': t.keyword.keyword,
                    'target_value': t.target_value,
                    'display_name': t.display_name or '',
                    'matched_product_id': t.matched_product_id or '',
                    'rank': latest.rank_position if latest else None,
                    'tracked_at': latest.tracked_at.isoformat() if latest else None,
                })
            result.append(data)
        return Response(result)

    def post(self, request):
        serializer = NaverTrackingScheduleSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)


class ScheduleDetailView(APIView):
    def put(self, request, pk):
        try:
            schedule = NaverTrackingSchedule.objects.get(id=pk)
        except NaverTrackingSchedule.DoesNotExist:
            return Response({'error': 'not found'}, status=404)
        serializer = NaverTrackingScheduleSerializer(schedule, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def delete(self, request, pk):
        try:
            NaverTrackingSchedule.objects.get(id=pk).delete()
            return Response(status=204)
        except NaverTrackingSchedule.DoesNotExist:
            return Response({'error': 'not found'}, status=404)


# ──── 데이터 초기화 ────

class DataResetView(APIView):
    """키워드는 유지하고 스냅샷/분석/순위 데이터만 삭제"""
    def post(self, request):
        snap_cnt = NaverSearchSnapshot.objects.count()
        anal_cnt = NaverTermAnalysis.objects.count()
        rank_cnt = NaverRankHistory.objects.count()

        NaverSearchSnapshot.objects.all().delete()
        NaverTermAnalysis.objects.all().delete()
        NaverRankHistory.objects.all().delete()

        # 키워드 카운트 초기화
        NaverKeyword.objects.all().update(
            total_count=0, naverpay_count=0, price_compare_count=0,
            term_count=0, terms=[], last_searched_at=None,
        )

        return Response({
            'status': 'ok',
            'deleted': {
                'snapshots': snap_cnt,
                'analyses': anal_cnt,
                'rank_history': rank_cnt,
            }
        })


# ──── 상품별 그룹 순위 Summary ────

class RankGroupedSummaryView(APIView):
    """matched_product_id 기준으로 상품별 그룹핑된 순위 요약.
    같은 스토어라도 매칭된 상품이 다르면 별도 그룹."""
    def get(self, request):
        targets = NaverRankTarget.objects.filter(is_active=True).select_related('keyword')
        groups = defaultdict(list)
        for t in targets:
            if t.matched_product_id:
                key = t.matched_product_id
            else:
                key = f'_unmatched_{t.target_type}::{t.target_value}'
            groups[key].append(t)

        results = []
        for gkey, tlist in groups.items():
            first = tlist[0]
            keywords = []
            for t in tlist:
                records = list(t.history.order_by('-tracked_at')[:2])
                current = records[0].rank_position if records else None
                previous = records[1].rank_position if len(records) > 1 else None
                change = (previous - current) if (current and previous) else None
                keywords.append({
                    'target_id': t.id,
                    'keyword': t.keyword.keyword,
                    'keyword_id': t.keyword.id,
                    'current_rank': current,
                    'previous_rank': previous,
                    'change': change,
                    'tracked_at': records[0].tracked_at.isoformat() if records else None,
                })

            display = first.matched_product_name or first.display_name or first.target_value
            results.append({
                'group_key': gkey,
                'target_type': first.target_type,
                'target_value': first.target_value,
                'display_name': display,
                'matched_product_id': first.matched_product_id,
                'source_product_id': first.source_product_id,
                'source_product_name': first.source_product_name,
                'auto_track': first.auto_track,
                'auto_track_times': first.auto_track_times or [],
                'keyword_count': len(tlist),
                'keywords': keywords,
            })
        return Response(results)


class RankPivotView(APIView):
    """특정 상품 그룹의 키워드×날짜 피벗 데이터"""
    def get(self, request):
        group_key = request.query_params.get('group_key', '').strip()
        days = int(request.query_params.get('days', 30))

        if not group_key:
            return Response({'error': 'group_key 필요'}, status=400)

        since = timezone.now() - timedelta(days=days)

        if group_key.startswith('_unmatched_'):
            # 미매칭 그룹: target_type::target_value
            parts = group_key.replace('_unmatched_', '').split('::', 1)
            if len(parts) == 2:
                targets = NaverRankTarget.objects.filter(
                    target_type=parts[0], target_value=parts[1], is_active=True
                ).select_related('keyword')
            else:
                return Response({'keywords': [], 'dates': [], 'data': {}})
        else:
            targets = NaverRankTarget.objects.filter(
                matched_product_id=group_key, is_active=True
            ).select_related('keyword')

        if not targets.exists():
            return Response({'keywords': [], 'dates': [], 'data': {}})

        keyword_names = []
        target_map = {}
        for t in targets:
            kw = t.keyword.keyword
            keyword_names.append(kw)
            target_map[t.id] = kw

        histories = NaverRankHistory.objects.filter(
            target__in=targets, tracked_at__gte=since
        ).order_by('-tracked_at')

        # 같은 날짜(MM/DD)에 여러 기록 → 최적(가장 낮은) 순위만 표시
        date_map = {}  # day_key → max datetime (정렬용)
        data = defaultdict(dict)  # kw → {day_key: best_rank}
        for h in histories:
            kw = target_map[h.target_id]
            day_key = h.tracked_at.strftime('%m/%d')
            if day_key not in date_map or h.tracked_at > date_map[day_key]:
                date_map[day_key] = h.tracked_at
            rank = h.rank_position
            if rank is not None:
                prev = data[kw].get(day_key)
                if prev is None or rank < prev:
                    data[kw][day_key] = rank
            elif day_key not in data[kw]:
                data[kw][day_key] = None

        date_labels = sorted(date_map.keys(), key=lambda k: date_map[k], reverse=True)

        return Response({
            'keywords': keyword_names,
            'dates': date_labels,
            'data': dict(data),
        })


class RankTrackedProductsView(APIView):
    """순위추적 중인 source_product_id 목록 (스마트스토어 배지용)"""
    def get(self, request):
        product_ids = list(
            NaverRankTarget.objects.filter(
                is_active=True, source_product_id__isnull=False
            ).values_list('source_product_id', flat=True).distinct()
        )
        return Response(product_ids)


class RankToggleAutoView(APIView):
    """상품 그룹 단위 자동추적 토글 + 수집 시간 설정"""
    def post(self, request):
        group_key = request.data.get('group_key', '').strip()
        enabled = request.data.get('enabled', False)
        times = request.data.get('times')  # ["09:00","13:00"] 또는 None

        if not group_key:
            return Response({'error': 'group_key 필요'}, status=400)

        if group_key.startswith('_unmatched_'):
            parts = group_key.replace('_unmatched_', '').split('::', 1)
            if len(parts) == 2:
                qs = NaverRankTarget.objects.filter(
                    target_type=parts[0], target_value=parts[1]
                )
            else:
                return Response({'error': 'invalid group_key'}, status=400)
        else:
            qs = NaverRankTarget.objects.filter(matched_product_id=group_key)

        update_kwargs = {'auto_track': enabled}
        if times is not None:
            # 최대 4개, HH:MM 형식 검증
            valid_times = sorted(set(
                t.strip() for t in times[:4]
                if isinstance(t, str) and len(t.strip()) == 5
            ))
            update_kwargs['auto_track_times'] = valid_times

        updated = qs.update(**update_kwargs)
        return Response({
            'updated': updated,
            'auto_track': enabled,
            'auto_track_times': update_kwargs.get('auto_track_times', []),
        })


# ──── 엑셀 다운로드 ────

def _fmt_store(p):
    return (
        p.get('mallName')
        or ((p.get('lowMallList') or [{}])[0].get('name'))
        or ((p.get('lowMallList') or [{}])[0].get('mallName'))
        or ''
    )

def _fmt_category(p):
    return ' > '.join([
        str(p.get(k) or '') for k in ('category1Name', 'category2Name', 'category3Name', 'category4Name')
        if p.get(k)
    ])

def _fmt_pipe(v):
    if not v:
        return ''
    if isinstance(v, list):
        return '|'.join(str(x) for x in v)
    return str(v)

def _fmt_tag(v):
    if not v:
        return ''
    if isinstance(v, list):
        return ','.join(
            x if isinstance(x, str) else (x.get('name') or x.get('value') or x.get('text') or '')
            for x in v if x
        )
    return str(v)

def _fmt_date(d):
    if not d:
        return ''
    s = str(d)
    if len(s) >= 8 and s[:8].isdigit():
        return f'{s[:4]}-{s[4:6]}-{s[6:8]}'
    return s

def _product_row(p):
    return [
        p.get('productName') or p.get('productTitle') or '',
        _fmt_store(p),
        _fmt_category(p),
        _fmt_pipe(p.get('attributeValue')),
        _fmt_pipe(p.get('characterValue')),
        _fmt_tag(p.get('manuTag')),
        p.get('brand') or '',
        p.get('maker') or '',
        int(p.get('reviewCount') or 0),
        _fmt_date(p.get('openDate')),
        p.get('imageUrl') or '',
    ]


class ExportProductsView(APIView):
    """키워드별 상품 상세 — 3시트 xlsx (전체/가격비교/네이버페이)
    ?images=true → 각 행 끝에 이미지 임베드 (~90px 리사이즈)
    """
    COLS = ['상품명', '스토어명', '카테고리명', '속성항목', '속성값', '태그',
            '브랜드', '제조사', '리뷰수', '등록일', '이미지URL']
    COLS_WITH_IMG = COLS + ['이미지']
    TABS = [('total', '전체'), ('model', '가격비교'), ('checkout', '네이버페이')]

    def get(self, request, keyword_id):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        with_images = (request.query_params.get('images', 'false').lower() in ('1', 'true', 'yes'))

        try:
            kw = NaverKeyword.objects.get(id=keyword_id)
        except NaverKeyword.DoesNotExist:
            return Response({'error': 'not found'}, status=404)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        hdr_font = Font(bold=True, color='FFFFFFFF')
        hdr_fill = PatternFill('solid', fgColor='FF03C75A')
        hdr_align = Alignment(horizontal='center', vertical='center')
        widths = [42, 22, 30, 28, 28, 30, 14, 14, 10, 14, 46]
        cols = self.COLS_WITH_IMG if with_images else self.COLS
        if with_images:
            widths = widths + [16]  # 이미지 컬럼 너비

        for tab_key, tab_name in self.TABS:
            ws = wb.create_sheet(title=tab_name)
            ws.append(cols)
            for c, w in zip(ws[1], widths):
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            snapshot = NaverSearchSnapshot.objects.filter(
                keyword=kw, tab_type=tab_key
            ).order_by('-collected_at').first()
            if not snapshot or not snapshot.products:
                continue

            for idx, p in enumerate(snapshot.products):
                row = _product_row(p)
                if with_images:
                    row = row + ['']
                ws.append(row)

            # 이미지 임베드
            if with_images:
                _embed_product_images(ws, snapshot.products, image_col_idx=len(cols))

        from urllib.parse import quote
        suffix = '_전체+이미지' if with_images else '_전체'
        fname = quote(f'{kw.keyword}{suffix}.xlsx')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{fname}"
        wb.save(response)
        return response


def _embed_product_images(ws, products, image_col_idx, max_size=90, timeout=4):
    """openpyxl 워크시트에 상품 이미지를 fetch + 리사이즈해 셀에 임베드.
    동시 다운로드(ThreadPoolExecutor) 후 직렬로 임베드 (openpyxl 은 thread-safe 아님).
    """
    import requests as req_mod
    from io import BytesIO
    from PIL import Image as PILImage
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    from concurrent.futures import ThreadPoolExecutor

    def fetch_and_resize(url):
        try:
            r = req_mod.get(url, timeout=timeout)
            if not r.ok or len(r.content) < 200:
                return None
            pil = PILImage.open(BytesIO(r.content))
            pil.thumbnail((max_size, max_size))
            if pil.mode in ('RGBA', 'P'):
                pil = pil.convert('RGB')
            buf_out = BytesIO()
            pil.save(buf_out, format='JPEG', quality=78)
            buf_out.seek(0)
            return buf_out
        except Exception:
            return None

    urls = [p.get('imageUrl') or '' for p in products]
    bufs = [None] * len(urls)
    with ThreadPoolExecutor(max_workers=10) as ex:
        for i, buf in enumerate(ex.map(lambda u: fetch_and_resize(u) if u else None, urls)):
            bufs[i] = buf

    col_letter = get_column_letter(image_col_idx)
    row_height_pt = max_size * 0.75 + 4
    for i, buf in enumerate(bufs):
        if not buf:
            continue
        try:
            xl_img = XLImage(buf)
            ws.add_image(xl_img, f'{col_letter}{i + 2}')
            ws.row_dimensions[i + 2].height = row_height_pt
        except Exception:
            continue


class ExportTermsView(APIView):
    def get(self, request):
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Term 분석'
        headers = ['키워드', '1term', '2term', '3term', '4term',
                    '순서고정가중치', '위치가중치', '상품명가중치',
                    '파트가중치', '카테고리우선여부', '총검색수', '상품수']
        ws.append(headers)

        keywords = NaverKeyword.objects.all()
        for kw in keywords:
            analysis = NaverTermAnalysis.objects.filter(keyword=kw).order_by('-analyzed_at').first()
            snapshot = NaverSearchSnapshot.objects.filter(keyword=kw, tab_type='total').order_by('-collected_at').first()

            row = [kw.keyword]
            terms = kw.terms or []
            for i in range(4):
                row.append(terms[i] if i < len(terms) else '')

            if analysis:
                row.append(str(analysis.order_weight))
                row.append(str(analysis.position_weight))
                row.append(str(analysis.name_weight))
                row.append(str(analysis.part_weight))
                row.append(str(analysis.category_priority))
            else:
                row.extend([''] * 5)

            row.append(kw.total_count)
            row.append(len(snapshot.products) if snapshot else 0)
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="naver_terms.xlsx"'
        wb.save(response)
        return response


class ExportRankView(APIView):
    def get(self, request):
        import openpyxl
        from django.http import HttpResponse

        days = int(request.query_params.get('days', 30))
        since = timezone.now() - timedelta(days=days)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '순위 이력'
        headers = ['날짜시간', '키워드', '대상', '순위', '탭', '상품명', '가격', '리뷰수']
        ws.append(headers)

        records = NaverRankHistory.objects.filter(
            tracked_at__gte=since
        ).select_related('target__keyword').order_by('-tracked_at')

        for r in records:
            ws.append([
                r.tracked_at.strftime('%Y-%m-%d %H:%M'),
                r.target.keyword.keyword,
                r.target.display_name or r.target.target_value,
                r.rank_position,
                r.tab_type,
                r.found_product_name,
                r.found_product_price,
                r.found_review_count,
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="naver_rank.xlsx"'
        wb.save(response)
        return response


# ──── 스마트 수집 / 분석 ────

class SmartCollectView(APIView):
    """스마트 수집 — HTTP 크롤링 + 자동 API fallback"""
    def post(self, request):
        keywords = request.data.get('keywords', [])
        method = request.data.get('method', 'auto')  # 'auto', 'http', 'api'
        tabs = request.data.get('tabs')

        if not keywords:
            return Response({'error': '키워드를 입력하세요'}, status=400)

        if method == 'api':
            # 공식 API 직행
            results = []
            for kw_text in keywords:
                kw_obj, _ = NaverKeyword.objects.get_or_create(keyword=kw_text)
                try:
                    api_data = services.fetch_products_via_api(kw_text, max_items=40)
                    NaverSearchSnapshot.objects.create(
                        keyword=kw_obj, tab_type='total',
                        products=api_data['products'], total=api_data['total'],
                    )
                    kw_obj.total_count = api_data['total']
                    kw_obj.last_searched_at = timezone.now()
                    kw_obj.save(update_fields=['total_count', 'last_searched_at'])
                    results.append({
                        'keyword': kw_text,
                        'tab': 'total',
                        'count': len(api_data['products']),
                        'total': api_data['total'],
                        'keyword_id': kw_obj.id,
                        'has_terms': bool(kw_obj.terms),
                    })
                except Exception as e:
                    results.append({'keyword': kw_text, 'error': str(e)})
                import time; time.sleep(0.2)

            return Response({
                'method_used': 'api',
                'terms_source': 'cache',
                'products_source': 'api',
                'results': results,
                'blocked': False,
            })

        # HTTP 크롤링 (auto 또는 http)
        from . import http_crawler
        logs = []

        def on_progress(event, data):
            logs.append({'event': event, **data})

        http_result = http_crawler.crawl_keywords(keywords, tabs=tabs, on_progress=on_progress)

        if not http_result['blocked']:
            return Response({
                'method_used': 'http',
                'terms_source': 'http',
                'products_source': 'http',
                'results': http_result['results'],
                'blocked': False,
                'logs': logs,
            })

        # 차단 감지 → API fallback (auto만)
        if method == 'http':
            return Response({
                'method_used': 'http',
                'results': http_result['results'],
                'blocked': True,
                'logs': logs,
            })

        # auto fallback → 남은 키워드를 API로
        done_kws = {r['keyword'] for r in http_result['results']}
        remaining = [kw for kw in keywords if kw not in done_kws]
        api_results = []

        for kw_text in remaining:
            kw_obj, _ = NaverKeyword.objects.get_or_create(keyword=kw_text)
            try:
                api_data = services.fetch_products_via_api(kw_text, max_items=40)
                NaverSearchSnapshot.objects.create(
                    keyword=kw_obj, tab_type='total',
                    products=api_data['products'], total=api_data['total'],
                )
                kw_obj.total_count = api_data['total']
                kw_obj.last_searched_at = timezone.now()
                kw_obj.save(update_fields=['total_count', 'last_searched_at'])
                api_results.append({
                    'keyword': kw_text, 'tab': 'total',
                    'count': len(api_data['products']),
                    'total': api_data['total'],
                    'keyword_id': kw_obj.id,
                    'has_terms': bool(kw_obj.terms),
                    'source': 'api',
                })
            except Exception as e:
                api_results.append({'keyword': kw_text, 'error': str(e), 'source': 'api'})
            import time; time.sleep(0.2)

        return Response({
            'method_used': 'auto',
            'terms_source': 'mixed',
            'products_source': 'mixed',
            'results': http_result['results'] + api_results,
            'blocked': True,
            'fallback_count': len(api_results),
            'logs': logs,
        })


class SmartAnalysisView(APIView):
    """스마트 분석 — 캐시된 terms + API products"""
    def post(self, request, keyword_id):
        method = request.data.get('method', 'auto')
        try:
            result = services.run_smart_analysis(keyword_id, method=method)
            if result.get('analysis'):
                from .serializers import NaverTermAnalysisSerializer
                result['analysis'] = NaverTermAnalysisSerializer(result['analysis']).data
            return Response(result)
        except NaverKeyword.DoesNotExist:
            return Response({'error': '키워드를 찾을 수 없습니다'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ──── UC 크롤러 ────

# ──── 연관키워드 ────

class RelatedKeywordView(APIView):
    def get(self, request):
        keyword = request.query_params.get('keyword', '').strip()
        if not keyword:
            return Response({'error': '키워드를 입력하세요'}, status=400)
        try:
            keyword_list = services.search_related_keywords(keyword)
            return Response({'keywordList': keyword_list})
        except ValueError as e:
            return Response({'error': str(e)}, status=500)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ──── 카테고리키워드 (데이터랩) ────

class DatalabCategoryView(APIView):
    def get(self, request):
        parent_cid = request.query_params.get('cid', '0')
        try:
            categories = services.get_datalab_categories(parent_cid)
            return Response(categories)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class DatalabCategoryKeywordRankView(APIView):
    def get(self, request):
        cid = request.query_params.get('cid', '').strip()
        start_date = request.query_params.get('startDate', '').strip()
        end_date = request.query_params.get('endDate', '').strip()
        age = request.query_params.get('age', '')
        gender = request.query_params.get('gender', '')
        device = request.query_params.get('device', '')

        if not cid:
            return Response({'error': '카테고리를 선택하세요'}, status=400)
        if not start_date or not end_date:
            return Response({'error': '날짜 범위를 입력하세요'}, status=400)

        try:
            result = services.get_category_keyword_rank(
                cid, start_date, end_date, age, gender, device,
            )
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class KeywordEnrichView(APIView):
    """키워드 목록에 대해 검색량/상품수/카테고리 데이터 보강"""
    def post(self, request):
        keywords = request.data.get('keywords', [])
        if not keywords or not isinstance(keywords, list):
            return Response({'error': '키워드 목록이 필요합니다'}, status=400)
        keywords = keywords[:50]  # 1회 최대 50개
        try:
            data = services.enrich_keywords(keywords)
            return Response({'data': data})
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class KeywordAutoMatchView(APIView):
    """상품명 기반 키워드 자동매칭"""
    def post(self, request):
        product_name = request.data.get('product_name', '').strip()
        keywords = request.data.get('keywords', [])
        if not product_name or not keywords:
            return Response({'error': 'product_name, keywords 필요'}, status=400)
        matches = services.match_keywords_for_product(product_name, keywords[:500])
        return Response({'matches': matches})


class CategoryNameLookupView(APIView):
    """카테고리 CID → 이름 조회 (naver_category DB)"""
    def get(self, request):
        cids = request.query_params.get('cids', '').strip()
        if not cids:
            return Response({'error': 'cids 필요'}, status=400)
        cid_list = [c.strip() for c in cids.split(',') if c.strip()]
        from django.db import connections
        result = {}
        try:
            with connections['myproduct'].cursor() as c:
                placeholders = ','.join(['%s'] * len(cid_list))
                c.execute(f'SELECT category_id, name FROM naver_category WHERE category_id IN ({placeholders})', cid_list)
                for row in c.fetchall():
                    result[str(row[0])] = row[1]
        except Exception:
            pass
        return Response(result)


class UCStartView(APIView):
    def post(self, request):
        keywords = request.data.get('keywords', [])
        if not keywords:
            return Response({'error': '키워드를 입력하세요'}, status=400)
        headless = request.data.get('headless', False)
        tab_order = request.data.get('tab_order', None)
        ok, msg = uc_crawler.start(keywords, headless=headless, tab_order=tab_order)
        if not ok:
            return Response({'ok': False, 'message': msg}, status=409)
        return Response({'ok': True, 'message': msg, 'count': len(keywords)})


class UCStatusView(APIView):
    def get(self, request):
        since = int(request.query_params.get('logSince', 0))
        st = uc_crawler.get_status()
        if since > 0:
            st['logs'] = [l for i, l in enumerate(st['logs']) if i >= since]
        return Response(st)


class UCStopView(APIView):
    def post(self, request):
        uc_crawler.stop()
        return Response({'ok': True})


class ProductBuyKeywordProxyView(APIView):
    """order 시스템의 구매키워드 API를 프록시"""
    def get(self, request, product_code):
        import requests as http_requests
        try:
            resp = http_requests.get(
                f'http://localhost:8989/orders/api_product_buy_keyword/{product_code}/',
                timeout=10,
            )
            return Response(resp.json())
        except Exception as e:
            return Response({'success': False, 'error': str(e)}, status=502)


# ──── 보고서 ────

class ReportListView(APIView):
    def get(self, request):
        reports = NaverReport.objects.values('id', 'title', 'report_type', 'created_at')
        return Response(list(reports))

    def post(self, request):
        title = request.data.get('title', '').strip()
        content = request.data.get('content', '').strip()
        report_type = request.data.get('report_type', 'monthly')
        if not title or not content:
            return Response({'error': '제목과 내용을 입력하세요'}, status=400)
        report = NaverReport.objects.create(title=title, content=content, report_type=report_type)
        return Response({'id': report.id, 'title': report.title, 'report_type': report.report_type,
                         'created_at': report.created_at}, status=201)


class ReportDetailView(APIView):
    def get(self, request, pk):
        try:
            r = NaverReport.objects.get(pk=pk)
        except NaverReport.DoesNotExist:
            return Response({'error': '보고서를 찾을 수 없습니다'}, status=404)
        return Response({'id': r.id, 'title': r.title, 'content': r.content,
                         'report_type': r.report_type, 'created_at': r.created_at})

    def delete(self, request, pk):
        deleted, _ = NaverReport.objects.filter(pk=pk).delete()
        if not deleted:
            return Response({'error': '보고서를 찾을 수 없습니다'}, status=404)
        return Response(status=204)


class ReportDownloadView(APIView):
    def get(self, request, pk):
        from django.http import HttpResponse
        try:
            r = NaverReport.objects.get(pk=pk)
        except NaverReport.DoesNotExist:
            return Response({'error': '보고서를 찾을 수 없습니다'}, status=404)
        filename = f'{r.title}.md'
        resp = HttpResponse(r.content, content_type='text/markdown; charset=utf-8')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{__import__('urllib.parse', fromlist=['quote']).quote(filename)}"
        return resp


# ──── 순위컨닝 ────

class RankCunningListView(APIView):
    def get(self, request):
        import pymysql
        import os
        pw = os.environ.get('MYPRODUCT_DB_PASSWORD', '')
        conn = pymysql.connect(host='192.168.219.200', user='root', password=pw, database='naverdb', autocommit=True)
        cur = conn.cursor(pymysql.cursors.DictCursor)
        cur.execute('SELECT * FROM rank_cunning_product ORDER BY created_at DESC')
        rows = cur.fetchall()
        conn.close()
        for r in rows:
            if r.get('created_at'):
                r['created_at'] = r['created_at'].isoformat()
        return Response(rows)

    def post(self, request):
        import pymysql
        import os
        products = request.data.get('products', [])
        if not products:
            return Response({'error': 'products 필드 필요'}, status=400)
        pw = os.environ.get('MYPRODUCT_DB_PASSWORD', '')
        conn = pymysql.connect(host='192.168.219.200', user='root', password=pw, database='naverdb', autocommit=True)
        cur = conn.cursor()
        added = 0
        for p in products:
            try:
                cur.execute('''INSERT IGNORE INTO rank_cunning_product
                    (origin_product_no, store_name, store_id, product_name, sale_price, category_id, product_image_url, seller_management_code)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)''',
                    (p['origin_product_no'], p.get('store_name',''), p.get('store_id'),
                     p.get('product_name',''), p.get('sale_price',0), p.get('category_id',''),
                     p.get('product_image_url',''), p.get('seller_management_code','')))
                added += cur.rowcount
            except Exception:
                pass
        conn.close()
        return Response({'added': added, 'total': len(products)}, status=201)


class RankCunningDetailView(APIView):
    def delete(self, request, pk):
        import pymysql
        import os
        pw = os.environ.get('MYPRODUCT_DB_PASSWORD', '')
        conn = pymysql.connect(host='192.168.219.200', user='root', password=pw, database='naverdb', autocommit=True)
        cur = conn.cursor()
        cur.execute('DELETE FROM rank_cunning_product WHERE id=%s', (pk,))
        conn.close()
        return Response(status=204)


# ══════════════════════════════════════════
# 구매수 추적
# ══════════════════════════════════════════

class PurchaseTargetListView(APIView):
    def get(self, request):
        targets = NaverPurchaseTarget.objects.filter(is_active=True).order_by('-created_at')
        data = NaverPurchaseTargetSerializer(targets, many=True).data
        return Response(data)

    def post(self, request):
        nv_mid = request.data.get('nv_mid', '').strip()
        if not nv_mid:
            return Response({'error': 'nv_mid 필수'}, status=400)
        target, created = NaverPurchaseTarget.objects.get_or_create(
            nv_mid=nv_mid,
            defaults={
                'product_name': request.data.get('product_name', ''),
                'store_name': request.data.get('store_name', ''),
                'image_url': request.data.get('image_url', ''),
                'category': request.data.get('category', ''),
                'source_keyword': request.data.get('source_keyword', ''),
                'source_rank': request.data.get('source_rank'),
            },
        )
        if not created:
            target.is_active = True
            target.save(update_fields=['is_active'])
        return Response(NaverPurchaseTargetSerializer(target).data, status=201 if created else 200)


class PurchaseTargetDetailView(APIView):
    def put(self, request, pk):
        try:
            target = NaverPurchaseTarget.objects.get(pk=pk)
        except NaverPurchaseTarget.DoesNotExist:
            return Response(status=404)
        for f in ['product_name', 'store_name', 'is_active', 'auto_track', 'auto_track_time']:
            if f in request.data:
                setattr(target, f, request.data[f])
        target.save()
        return Response(NaverPurchaseTargetSerializer(target).data)

    def delete(self, request, pk):
        NaverPurchaseTarget.objects.filter(pk=pk).delete()
        return Response(status=204)


class RunPurchaseTrackingView(APIView):
    def post(self, request):
        from . import purchase_crawler

        target_ids = request.data.get('target_ids')
        headless = request.data.get('headless', True)

        if target_ids:
            targets = NaverPurchaseTarget.objects.filter(pk__in=target_ids, is_active=True)
        else:
            targets = NaverPurchaseTarget.objects.filter(is_active=True)

        if not targets.exists():
            return Response({'ok': False, 'message': '추적 대상 없음'})

        ok, msg = purchase_crawler.start(targets, headless=headless)
        return Response({'ok': ok, 'message': msg, 'count': targets.count()})


class PurchaseTrackStatusView(APIView):
    def get(self, request):
        from . import purchase_crawler
        log_since = int(request.query_params.get('logSince', 0))
        return Response(purchase_crawler.get_status(log_since))


class PurchaseTrackStopView(APIView):
    def post(self, request):
        from . import purchase_crawler
        purchase_crawler.stop()
        return Response({'ok': True})


class PurchaseHistoryView(APIView):
    def get(self, request):
        target_id = request.query_params.get('target_id')
        days = int(request.query_params.get('days', 30))
        since = timezone.now() - timedelta(days=days)

        qs = NaverPurchaseHistory.objects.filter(tracked_at__gte=since)
        if target_id:
            qs = qs.filter(target_id=target_id)
        qs = qs.select_related('target').order_by('-tracked_at')[:500]

        data = []
        for h in qs:
            data.append({
                'id': h.id,
                'target_id': h.target_id,
                'target_name': h.target.product_name,
                'target_nv_mid': h.target.nv_mid,
                'purchase_count': h.purchase_count,
                'review_count': h.review_count,
                'keep_count': h.keep_count,
                'price': h.price,
                'crawl_success': h.crawl_success,
                'error_message': h.error_message,
                'tracked_at': h.tracked_at.isoformat(),
            })
        return Response(data)


class PurchaseSummaryView(APIView):
    def get(self, request):
        targets = NaverPurchaseTarget.objects.filter(is_active=True).order_by('-created_at')
        result = []
        for t in targets:
            records = list(t.history.filter(crawl_success=True).order_by('-tracked_at')[:2])
            curr = records[0] if records else None
            prev = records[1] if len(records) > 1 else None
            delta = None
            if curr and prev and curr.purchase_count is not None and prev.purchase_count is not None:
                delta = curr.purchase_count - prev.purchase_count

            result.append({
                'id': t.id,
                'nv_mid': t.nv_mid,
                'product_name': t.product_name,
                'store_name': t.store_name,
                'image_url': t.image_url,
                'category': t.category,
                'source_keyword': t.source_keyword,
                'source_rank': t.source_rank,
                'current_purchase_count': curr.purchase_count if curr else None,
                'previous_purchase_count': prev.purchase_count if prev else None,
                'purchase_delta': delta,
                'current_review_count': curr.review_count if curr else None,
                'current_keep_count': curr.keep_count if curr else None,
                'current_price': curr.price if curr else None,
                'last_tracked_at': curr.tracked_at.isoformat() if curr else None,
                'auto_track': t.auto_track,
                'auto_track_time': t.auto_track_time,
            })
        return Response(result)


class PurchaseToggleAutoView(APIView):
    def post(self, request):
        target_id = request.data.get('target_id')
        enabled = request.data.get('enabled', False)
        track_time = request.data.get('time', '09:00')
        try:
            target = NaverPurchaseTarget.objects.get(pk=target_id)
        except NaverPurchaseTarget.DoesNotExist:
            return Response({'error': '대상 없음'}, status=404)
        target.auto_track = enabled
        target.auto_track_time = track_time
        target.save(update_fields=['auto_track', 'auto_track_time'])
        return Response({'ok': True})


# ══════════════════════════════════════════
# 동의어 (키워드별)
# ══════════════════════════════════════════

class SynonymListView(APIView):
    """키워드별 동의어 목록 조회 + 신규 추가 (수동/자동완성/사전 출처)"""
    def get(self, request, keyword_id):
        try:
            NaverKeyword.objects.get(id=keyword_id)
        except NaverKeyword.DoesNotExist:
            return Response({'error': 'keyword not found'}, status=404)
        qs = NaverSynonym.objects.filter(keyword_id=keyword_id).order_by(
            '-is_confirmed', '-verification_score', 'word'
        )
        return Response(NaverSynonymSerializer(qs, many=True).data)

    def post(self, request, keyword_id):
        try:
            kw = NaverKeyword.objects.get(id=keyword_id)
        except NaverKeyword.DoesNotExist:
            return Response({'error': 'keyword not found'}, status=404)
        word = (request.data.get('word') or '').strip()
        if not word:
            return Response({'error': 'word 필요'}, status=400)
        if word == kw.keyword:
            return Response({'error': '같은 단어는 추가할 수 없습니다'}, status=400)
        source = request.data.get('source') or 'manual'
        is_confirmed = request.data.get('is_confirmed')

        defaults = {'source': source}
        if is_confirmed is not None:
            defaults['is_confirmed'] = bool(is_confirmed)
        syn, created = NaverSynonym.objects.get_or_create(
            keyword=kw, word=word, defaults=defaults,
        )
        if not created and is_confirmed is not None:
            syn.is_confirmed = bool(is_confirmed)
            syn.save(update_fields=['is_confirmed', 'updated_at'])
        return Response(NaverSynonymSerializer(syn).data, status=201 if created else 200)


class SynonymDetailView(APIView):
    """동의어 확정여부 변경/삭제"""
    def patch(self, request, pk):
        try:
            syn = NaverSynonym.objects.get(id=pk)
        except NaverSynonym.DoesNotExist:
            return Response({'error': 'not found'}, status=404)
        if 'is_confirmed' in request.data:
            v = request.data['is_confirmed']
            syn.is_confirmed = None if v is None else bool(v)
        syn.save()
        return Response(NaverSynonymSerializer(syn).data)

    def delete(self, request, pk):
        try:
            NaverSynonym.objects.get(id=pk).delete()
            return Response(status=204)
        except NaverSynonym.DoesNotExist:
            return Response({'error': 'not found'}, status=404)


class SynonymLookupView(APIView):
    """네이버 사전 + 자동완성에서 동의어 후보를 가져와 미확정으로 등록.
    POST /api/naver/synonyms/<kw>/lookup/
    body: {"include_autocomplete": true|false}
    """
    def post(self, request, keyword_id):
        try:
            kw = NaverKeyword.objects.get(id=keyword_id)
        except NaverKeyword.DoesNotExist:
            return Response({'error': 'keyword not found'}, status=404)

        include_ac = request.data.get('include_autocomplete', True)

        candidates = {}  # word → source

        # 1. 네이버 사전
        try:
            for w in services.fetch_naver_dict_synonyms(kw.keyword):
                if w != kw.keyword:
                    candidates.setdefault(w, 'naver_dict')
        except Exception:
            pass

        # 2. 자동완성 (옵션)
        ac_words = []
        if include_ac:
            try:
                ac_words = services.fetch_autocomplete_naver(kw.keyword)
            except Exception:
                ac_words = []
            for w in ac_words:
                # "강아지 사료" 처럼 원 키워드로 시작하는 복합어는 동의어 후보 아님 → 제외
                if kw.keyword in w and w != kw.keyword:
                    continue
                if w in (kw.keyword,):
                    continue
                candidates.setdefault(w, 'autocomplete')

        added = []
        for word, source in candidates.items():
            syn, created = NaverSynonym.objects.get_or_create(
                keyword=kw, word=word, defaults={'source': source},
            )
            if created:
                added.append(syn)

        all_qs = NaverSynonym.objects.filter(keyword=kw).order_by(
            '-is_confirmed', '-verification_score', 'word'
        )
        return Response({
            'added': len(added),
            'candidates_count': len(candidates),
            'autocomplete_count': len(ac_words),
            'synonyms': NaverSynonymSerializer(all_qs, many=True).data,
        })


class SynonymVerifyView(APIView):
    """네이버쇼핑 검색으로 동의어 검증.
    POST /api/naver/synonyms/<kw>/verify/  body: {"word": "..."}  (이미 등록된 후보)
    또는 body: {"synonym_id": 123}
    검증결과를 verification_score/verification_data 에 저장.
    """
    def post(self, request, keyword_id):
        try:
            kw = NaverKeyword.objects.get(id=keyword_id)
        except NaverKeyword.DoesNotExist:
            return Response({'error': 'keyword not found'}, status=404)

        synonym_id = request.data.get('synonym_id')
        word = (request.data.get('word') or '').strip()

        syn = None
        if synonym_id:
            try:
                syn = NaverSynonym.objects.get(id=synonym_id, keyword=kw)
                word = syn.word
            except NaverSynonym.DoesNotExist:
                return Response({'error': 'synonym not found'}, status=404)
        elif word:
            syn = NaverSynonym.objects.filter(keyword=kw, word=word).first()

        if not word:
            return Response({'error': 'word 또는 synonym_id 필요'}, status=400)

        try:
            result = services.verify_synonym_in_shopping(kw.keyword, word)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

        if 'error' in result:
            return Response({'error': result['error']}, status=502)

        # 저장 (없으면 생성)
        if syn is None:
            syn = NaverSynonym.objects.create(keyword=kw, word=word, source='manual')
        syn.verification_score = result.get('score')
        syn.verification_data = result
        syn.save(update_fields=['verification_score', 'verification_data', 'updated_at'])

        data = NaverSynonymSerializer(syn).data
        data['verification'] = result
        return Response(data)


class AutocompleteView(APIView):
    """마켓별 자동완성 조회 (현재 1단계: 네이버, 쿠팡)
    POST /api/naver/autocomplete/
    body: {"query": "강아지", "markets": ["naver","coupang"]}
    응답: {"query":..., "results": {"naver": {"keywords":[...], "error":null}, ...}}
    """
    def post(self, request):
        query = (request.data.get('query') or '').strip()
        markets = request.data.get('markets') or ['naver', 'coupang']
        if not query:
            return Response({'error': 'query 필요'}, status=400)
        if not isinstance(markets, list):
            return Response({'error': 'markets는 배열이어야 합니다'}, status=400)

        results = services.fetch_autocomplete_multi(query, markets)
        return Response({'query': query, 'results': results})


# ── 확장프로그램 → 구매수 결과 수신 ──
class ExtPurchaseResultView(APIView):
    def post(self, request):
        target_id = request.data.get('target_id')
        if not target_id:
            return Response({'error': 'target_id 필수'}, status=400)
        try:
            target = NaverPurchaseTarget.objects.get(pk=target_id)
        except NaverPurchaseTarget.DoesNotExist:
            return Response({'error': 'target not found'}, status=404)

        pc = request.data.get('purchase_count')
        NaverPurchaseHistory.objects.create(
            target=target,
            purchase_count=pc,
            review_count=request.data.get('review_count'),
            keep_count=request.data.get('keep_count'),
            price=request.data.get('price'),
            crawl_success=pc is not None,
            error_message=request.data.get('error', ''),
        )
        return Response({'ok': True, 'target_id': target_id}, status=201)
