"""2025-04-01 이후 스마트스토어로 판매된 상품 리스트 추출."""
import os
import django
from datetime import datetime

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from django.db import connections
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

VALID_ORDER_STATUSES = (
    '고객주문', '신규주문', '배송준비', '입금확인', '배송준비중',
    '배송중', '배송완료', '거래완료', '오더완료',
)
SMARTSTORE_SITE = '04.스마트스토어'
START_DATE = '2025-04-01'

status_ph = ','.join(['%s'] * len(VALID_ORDER_STATUSES))
sql = f"""
    SELECT
        product_seller_code,
        MAX(product_name)         AS product_name,
        COUNT(*)                  AS order_count,
        SUM(quantity)             AS total_qty,
        SUM(payment_price)        AS total_amount,
        SUM(settlement_price)     AS total_settle,
        MIN(order_date)           AS first_date,
        MAX(order_date)           AS last_date
    FROM orders_order
    WHERE site_name = %s
      AND order_date >= %s
      AND order_status IN ({status_ph})
      AND product_seller_code IS NOT NULL
      AND product_seller_code <> ''
    GROUP BY product_seller_code
    ORDER BY total_amount DESC
"""

with connections['joacham'].cursor() as cur:
    cur.execute(sql, [SMARTSTORE_SITE, START_DATE, *VALID_ORDER_STATUSES])
    cols = [c[0] for c in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]

print(f'[1/3] orders_order 집계: {len(rows):,}개 SKU')

# smartstore_product 매핑 (관리코드 → 상품정보 + 상점명)
codes = [r['product_seller_code'] for r in rows]
product_map = {}
if codes:
    chunk = 1000
    with connections['myproduct'].cursor() as cur:
        for i in range(0, len(codes), chunk):
            sub = codes[i:i + chunk]
            ph = ','.join(['%s'] * len(sub))
            cur.execute(
                f"""
                SELECT p.seller_management_code, p.name, p.origin_product_no,
                       p.channel_product_no, p.sale_price, p.status_type,
                       p.stock_quantity, s.store_name
                FROM smartstore_product p
                JOIN smartstoreIdList s ON s.id = p.store_id
                WHERE p.seller_management_code IN ({ph})
                """,
                sub,
            )
            for r in cur.fetchall():
                product_map[r[0]] = {
                    'name': r[1] or '',
                    'origin_product_no': r[2] or '',
                    'channel_product_no': r[3] or '',
                    'sale_price': r[4] or 0,
                    'status_type': r[5] or '',
                    'stock_quantity': r[6] or 0,
                    'store_name': r[7] or '',
                }

print(f'[2/3] smartstore_product 매칭: {len(product_map):,}/{len(codes):,}개')

STATUS_LABELS = {
    'SALE': '판매중', 'SUSPENSION': '판매중지',
    'CLOSE': '판매종료', 'PROHIBITION': '판매금지',
    'WAIT': '대기',
}

wb = Workbook()
ws = wb.active
ws.title = '2025-04 이후 판매상품'

headers = [
    '순위', '상점명', '관리코드', '상품번호(origin)', '채널상품번호', '상품명',
    '현재상태', '판매가', '재고',
    '주문건수', '총수량', '총결제금액', '총정산금액',
    '첫주문일', '마지막주문일',
]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='03C75A')
    c.alignment = Alignment(horizontal='center', vertical='center')

thin = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD'),
)

for idx, r in enumerate(rows, 1):
    code = r['product_seller_code']
    info = product_map.get(code, {})
    name = info.get('name') or r['product_name'] or ''
    status_kr = STATUS_LABELS.get(info.get('status_type', ''), info.get('status_type', '') or '미동기화')
    fd = r['first_date']
    ld = r['last_date']
    values = [
        idx,
        info.get('store_name', ''),
        code,
        info.get('origin_product_no', ''),
        info.get('channel_product_no', ''),
        name,
        status_kr,
        info.get('sale_price', 0),
        info.get('stock_quantity', 0),
        int(r['order_count'] or 0),
        int(r['total_qty'] or 0),
        int(r['total_amount'] or 0),
        int(r['total_settle'] or 0),
        fd.strftime('%Y-%m-%d') if hasattr(fd, 'strftime') else str(fd),
        ld.strftime('%Y-%m-%d') if hasattr(ld, 'strftime') else str(ld),
    ]
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=idx + 1, column=col, value=v)
        cell.border = thin
        if col in (8, 9, 10, 11, 12, 13):
            cell.number_format = '#,##0'

widths = [6, 14, 18, 16, 16, 60, 12, 12, 8, 10, 10, 14, 14, 12, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
ws.freeze_panes = 'A2'

out_dir = os.path.join(os.path.dirname(__file__), 'exports')
os.makedirs(out_dir, exist_ok=True)
ts = datetime.now().strftime('%Y%m%d_%H%M%S')
out_path = os.path.join(out_dir, f'스스_2025_04이후_판매상품_{ts}.xlsx')
wb.save(out_path)

matched = sum(1 for r in rows if r['product_seller_code'] in product_map)
total_amt = sum(int(r['total_amount'] or 0) for r in rows)
print(f'[3/3] 저장 완료: {out_path}')
print(f'  - 총 SKU: {len(rows):,}개  (스스 카탈로그 매칭: {matched:,}개)')
print(f'  - 합계 결제금액: {total_amt:,}원')
