import os
from io import BytesIO
from urllib.parse import quote

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response

from . import smartstore_service
from . import smartstore_product_service
from . import smartstore_order_service
from . import smartstore_analytics_service
from . import product_audit_service
from . import attr_analytics_service
from . import missing_attrs_service
from . import worker_dashboard_service
from . import naver_my_product_service
from . import register_set_service
from . import bulk_register_service
from . import register_folder_service
from . import live_name_service
from . import bad_name_service
from . import dup_suspend_service
from . import diagnosis_service
from . import naver_name_generator
from . import naver_vision_analyzer


# ── 스마트스토어 상점 관리 ──

class SmartStoreStoreListView(APIView):
    def get(self, request):
        include_inactive = request.query_params.get('all') == '1'
        return Response(smartstore_service.get_stores(include_inactive))

    def post(self, request):
        d = request.data
        if not d.get('store_id') or not d.get('store_pw') or not d.get('store_name'):
            return Response({'error': 'store_id, store_pw, store_name required'}, status=400)
        try:
            store = smartstore_service.create_store(d)
            return Response(store, status=201)
        except Exception as e:
            msg = str(e)
            if 'Duplicate' in msg:
                return Response({'error': '이미 등록된 store_id입니다.'}, status=400)
            return Response({'error': msg}, status=400)


class SmartStoreStoreCountsView(APIView):
    """좌측 사이드바용: 활성 스토어 + status 별 상품 카운트."""
    def get(self, request):
        return Response({'items': smartstore_service.get_stores_with_counts()})


class SmartStoreStoreSampleExcelView(APIView):
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = '상점목록'

        headers = ['store_id', 'store_pw', 'store_name', 'store_url', 'application_id', 'application_secret', 'memo']
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill('solid', fgColor='D9D9D9')

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        stores = smartstore_service.get_stores(include_inactive=True)
        if stores:
            for row_idx, s in enumerate(stores, 2):
                ws.cell(row=row_idx, column=1, value=s.get('store_id', ''))
                ws.cell(row=row_idx, column=2, value=s.get('store_pw', ''))
                ws.cell(row=row_idx, column=3, value=s.get('store_name', ''))
                ws.cell(row=row_idx, column=4, value=s.get('store_url', '') or '')
                ws.cell(row=row_idx, column=5, value=s.get('commerce_api_key', '') or '')
                ws.cell(row=row_idx, column=6, value=s.get('commerce_secret_key', '') or '')
                ws.cell(row=row_idx, column=7, value=s.get('memo', '') or '')
        else:
            example = ['mystore', 'password123', '내상점', 'mystore-url', '', '', '테스트']
            for col, v in enumerate(example, 1):
                ws.cell(row=2, column=col, value=v)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 20

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = '스마트스토어_상점목록.xlsx'
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return resp


class SmartStoreStoreBulkUploadView(APIView):
    HEADER_MAP = {
        'store_id': 'store_id', '스토어id': 'store_id', '스토어 id': 'store_id',
        'store_pw': 'store_pw', '비밀번호': 'store_pw', 'password': 'store_pw',
        'store_name': 'store_name', '상점명': 'store_name', '스토어명': 'store_name',
        'store_url': 'store_url', '스토어주소': 'store_url', '스토어url': 'store_url', '상점주소': 'store_url',
        'commerce_api_key': 'commerce_api_key', '커머스api키': 'commerce_api_key', 'api키': 'commerce_api_key', '애플리케이션id': 'commerce_api_key', '앱id': 'commerce_api_key', 'application_id': 'commerce_api_key',
        'commerce_secret_key': 'commerce_secret_key', '커머스시크릿키': 'commerce_secret_key', '시크릿키': 'commerce_secret_key', 'application_secret': 'commerce_secret_key', '애플리케이션시크릿': 'commerce_secret_key', '앱시크릿': 'commerce_secret_key',
        'memo': 'memo', '메모': 'memo',
    }

    def post(self, request):
        from openpyxl import load_workbook

        xls_file = request.FILES.get('file')
        if not xls_file:
            return Response({'error': '파일을 선택하세요.'}, status=400)

        try:
            wb = load_workbook(xls_file, read_only=True)
        except Exception:
            return Response({'error': 'xlsx 파일만 지원합니다.'}, status=400)

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            return Response({'error': '빈 파일입니다.'}, status=400)

        col_map = {}
        for idx, val in enumerate(header_row):
            if val is None:
                continue
            key = str(val).strip().lower().replace(' ', '')
            mapped = self.HEADER_MAP.get(key)
            if mapped:
                col_map[idx] = mapped

        if 'store_id' not in col_map.values():
            return Response({'error': 'store_id 컬럼을 찾을 수 없습니다.'}, status=400)

        data_rows = []
        for row in rows_iter:
            if all(v is None or str(v).strip() == '' for v in row):
                continue
            item = {}
            for idx, field in col_map.items():
                val = row[idx] if idx < len(row) else None
                item[field] = str(val).strip() if val is not None else ''
            data_rows.append(item)

        wb.close()

        if not data_rows:
            return Response({'error': '데이터가 없습니다.'}, status=400)

        result = smartstore_service.bulk_create_stores(data_rows)
        return Response(result)


class SmartStoreStoreDetailView(APIView):
    def put(self, request, pk):
        store = smartstore_service.get_store(pk)
        if not store:
            return Response({'error': 'not found'}, status=404)
        updated = smartstore_service.update_store(pk, request.data)
        return Response(updated)

    def delete(self, request, pk):
        store = smartstore_service.get_store(pk)
        if not store:
            return Response({'error': 'not found'}, status=404)
        if store.get('is_active'):
            smartstore_service.deactivate_store(pk)
        else:
            smartstore_service.delete_store(pk)
        return Response(status=204)


# ── 스마트스토어 상품 관리 ──

class SmartStoreProductListView(APIView):
    def _parse_params(self, src):
        store_id = src.get('store_id')
        if store_id is None:
            return None, Response({'error': 'store_id required'}, status=400)
        page = int(src.get('page', 1))
        per_page = int(src.get('per_page', 50))
        status = src.get('status') or None
        search = src.get('search') or None
        ownerclan_soldout = src.get('ownerclan_soldout')
        is_focus = src.get('is_focus')
        has_orders = src.get('has_orders')
        sort_by = src.get('sort_by') or None
        sort_dir = src.get('sort_dir') or None
        min_ss_amount = src.get('min_ss_amount')
        has_changes = src.get('has_changes')
        reverse_margin = src.get('reverse_margin')
        restock_unchecked = src.get('restock_unchecked')
        no_master = src.get('no_master')
        result = smartstore_product_service.get_products(
            int(store_id), page, per_page, status, search,
            ownerclan_soldout=int(ownerclan_soldout) if ownerclan_soldout is not None else None,
            is_focus=int(is_focus) if is_focus is not None else None,
            has_orders=int(has_orders) if has_orders is not None else None,
            sort_by=sort_by,
            sort_dir=sort_dir,
            min_ss_amount=int(min_ss_amount) if min_ss_amount is not None else None,
            has_changes=int(has_changes) if has_changes is not None else None,
            reverse_margin=int(reverse_margin) if reverse_margin is not None else None,
            restock_unchecked=int(restock_unchecked) if restock_unchecked is not None else None,
            no_master=int(no_master) if no_master is not None else None,
        )
        return result, None

    def get(self, request):
        result, err = self._parse_params(request.query_params)
        return err or Response(result)


class SmartStoreProductSearchView(APIView):
    """POST 검색 — 다수 코드 검색 시 URL 길이 제한 회피"""
    def post(self, request):
        result, err = SmartStoreProductListView()._parse_params(request.data)
        return err or Response(result)


class SmartStoreProductSyncView(APIView):
    def post(self, request):
        store_id = request.data.get('store_id')
        if not store_id:
            return Response({'error': 'store_id required'}, status=400)
        result = smartstore_product_service.sync_products(int(store_id))
        if 'error' in result:
            return Response(result, status=400)
        return Response(result)


class SmartStoreRefreshTrackingView(APIView):
    def post(self, request):
        from . import smartstore_order_service
        store_id = int(request.data.get('store_id', 0))
        count = smartstore_product_service.refresh_master_tracking(store_id)
        order_count = smartstore_order_service.update_product_sales_summary()
        return Response({'refreshed': count, 'orders_updated': order_count})


class SmartStoreSalesSnapshotView(APIView):
    def post(self, request):
        from . import smartstore_order_service
        saved = smartstore_order_service.save_daily_snapshot()
        return Response({'saved': saved})

    def get(self, request):
        from . import smartstore_order_service
        store_id = int(request.query_params.get('store_id', 0))
        delta = smartstore_order_service.get_sales_delta(store_id)
        return Response(delta or {})


class SmartStoreSyncLogView(APIView):
    def get(self, request):
        store_id = int(request.query_params.get('store_id', 0))
        limit = int(request.query_params.get('limit', 50))
        return Response(smartstore_product_service.get_sync_logs(store_id, limit))


class SmartStoreSyncLogDetailView(APIView):
    def get(self, request, pk):
        limit = int(request.query_params.get('limit', 500))
        return Response(smartstore_product_service.get_sync_log_changes(pk, limit))


class SmartStoreProductStatsView(APIView):
    def get(self, request):
        store_id = request.query_params.get('store_id')
        if store_id:
            result = smartstore_product_service.get_product_stats(int(store_id))
        else:
            result = smartstore_product_service.get_all_stores_stats()
        return Response(result)


class SmartStoreProductExcelView(APIView):
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        store_ids = request.query_params.getlist('store_ids')
        statuses = request.query_params.getlist('statuses')
        w_only = request.query_params.get('w_only') == '1'
        search = request.query_params.get('search') or None
        has_orders = request.query_params.get('has_orders') == '1'
        is_focus = request.query_params.get('is_focus') == '1'
        sort_by = request.query_params.get('sort_by') or None
        sort_dir = request.query_params.get('sort_dir') or None

        store_ids = [int(s) for s in store_ids] if store_ids else None
        statuses = statuses if statuses else None

        rows = smartstore_product_service.get_products_for_export(
            store_ids, statuses, w_only,
            search=search, has_orders=has_orders, is_focus=is_focus,
            sort_by=sort_by, sort_dir=sort_dir,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = '상품목록'

        headers = ['상점명', '상품번호', '채널상품번호', '상품명', '판매가',
                    '재고', '상태', '노출상태', '관리코드', '카테고리ID',
                    '주문건수', '스마트주문건수', '판매금액', '스마트판매금액',
                    '판매수량', '스마트판매수량', '동기화일시']
        col_widths = [15, 14, 14, 40, 12, 8, 10, 10, 18, 14, 10, 12, 14, 14, 10, 12, 20]

        header_font = Font(bold=True, size=10)
        header_fill = PatternFill('solid', fgColor='F0F0F0')
        thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))
        money_fmt = '#,##0'

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        STATUS_LABELS = {
            'SALE': '판매중', 'SUSPENSION': '판매중지',
            'CLOSE': '판매종료', 'PROHIBITION': '판매금지',
            'WAIT': '대기',
        }

        for i, r in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=r['store_name']).border = thin_border
            ws.cell(row=i, column=2, value=r['origin_product_no']).border = thin_border
            ws.cell(row=i, column=3, value=r['channel_product_no']).border = thin_border
            ws.cell(row=i, column=4, value=r['name']).border = thin_border
            c = ws.cell(row=i, column=5, value=r['sale_price'])
            c.number_format = money_fmt
            c.border = thin_border
            ws.cell(row=i, column=6, value=r['stock_quantity']).border = thin_border
            ws.cell(row=i, column=7, value=STATUS_LABELS.get(r['status_type'] or '', r['status_type'] or '')).border = thin_border
            ws.cell(row=i, column=8, value=r['channel_product_display_status_type'] or '').border = thin_border
            ws.cell(row=i, column=9, value=r['seller_management_code'] or '').border = thin_border
            ws.cell(row=i, column=10, value=r['category_id'] or '').border = thin_border
            c = ws.cell(row=i, column=11, value=r.get('all_order_count', 0))
            c.number_format = '#,##0'
            c.border = thin_border
            c = ws.cell(row=i, column=12, value=r.get('total_order_count', 0))
            c.number_format = '#,##0'
            c.border = thin_border
            c = ws.cell(row=i, column=13, value=r.get('all_order_amount', 0))
            c.number_format = money_fmt
            c.border = thin_border
            c = ws.cell(row=i, column=14, value=r.get('total_order_amount', 0))
            c.number_format = money_fmt
            c.border = thin_border
            c = ws.cell(row=i, column=15, value=r.get('all_order_qty', 0))
            c.number_format = '#,##0'
            c.border = thin_border
            c = ws.cell(row=i, column=16, value=r.get('total_order_qty', 0))
            c.number_format = '#,##0'
            c.border = thin_border
            synced = r['synced_at']
            if hasattr(synced, 'strftime'):
                synced = synced.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=i, column=17, value=synced or '').border = thin_border

        for j, w in enumerate(col_widths):
            ws.column_dimensions[chr(65 + j)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = '스마트스토어_상품목록.xlsx'
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return resp


class SmartStoreProductCountView(APIView):
    def get(self, request):
        store_ids = request.query_params.getlist('store_ids')
        statuses = request.query_params.getlist('statuses')
        w_only = request.query_params.get('w_only') == '1'

        store_ids = [int(s) for s in store_ids] if store_ids else None
        statuses = statuses if statuses else None

        from django.db import connections
        where = ['1=1']
        params = []
        if store_ids:
            ph = ','.join(['%s'] * len(store_ids))
            where.append(f'p.store_id IN ({ph})')
            params.extend(store_ids)
        if statuses:
            ph = ','.join(['%s'] * len(statuses))
            where.append(f'p.status_type IN ({ph})')
            params.extend(statuses)
        if w_only:
            where.append("p.seller_management_code LIKE 'W%%'")

        with connections['myproduct'].cursor() as cur:
            cur.execute(
                f"SELECT COUNT(*) FROM smartstore_product p "
                f"JOIN smartstoreIdList s ON s.id = p.store_id "
                f"WHERE {' AND '.join(where)}", params)
            count = cur.fetchone()[0]
        return Response({'count': count})


class SmartStoreProductWCodesView(APIView):
    def get(self, request):
        store_ids = request.query_params.getlist('store_ids')
        statuses = request.query_params.getlist('statuses')

        store_ids = [int(s) for s in store_ids] if store_ids else None
        statuses = statuses if statuses else None

        codes = smartstore_product_service.get_w_codes(store_ids, statuses)
        return Response({'codes': codes})


class SmartStoreProductSuspendPreviewView(APIView):
    def post(self, request):
        product_ids = request.data.get('product_ids', [])
        select_all = request.data.get('select_all', False)
        filters = request.data.get('filters', {})
        result = smartstore_product_service.preview_suspend(product_ids, select_all, filters)
        return Response(result)


class SmartStoreProductSuspendView(APIView):
    def post(self, request):
        product_ids = request.data.get('product_ids', [])
        select_all = request.data.get('select_all', False)
        filters = request.data.get('filters', {})
        result = smartstore_product_service.suspend_products(product_ids, select_all, filters)
        return Response(result)


class SmartStoreProductFocusView(APIView):
    def post(self, request):
        product_ids = request.data.get('product_ids', [])
        is_focus = request.data.get('is_focus', 1)
        if not product_ids:
            return Response({'error': 'product_ids required'}, status=400)
        result = smartstore_product_service.toggle_focus(product_ids, is_focus)
        return Response(result)


class SmartStoreProductRestockCheckView(APIView):
    def post(self, request):
        product_ids = request.data.get('product_ids', [])
        checked = request.data.get('checked', 1)
        if not product_ids:
            return Response({'error': 'product_ids required'}, status=400)
        result = smartstore_product_service.toggle_restock_checked(product_ids, checked)
        return Response(result)


class SmartStoreProductOrphanWCodesView(APIView):
    def get(self, request):
        store_ids = request.query_params.getlist('store_ids')
        store_ids = [int(s) for s in store_ids] if store_ids else None
        codes = smartstore_product_service.get_orphan_w_codes(store_ids)
        return Response({'codes': codes, 'count': len(codes)})


class SmartStoreProductOrphanSoldoutView(APIView):
    """오너클랜 카탈로그 이탈 W코드인데 SS는 SALE 로 박제된 상품(품절 박제)."""
    def get(self, request):
        store_id = int(request.query_params.get('store_id', 0) or 0)
        return Response(smartstore_product_service.get_orphan_soldout(store_id))


class SmartStoreReconcileView(APIView):
    """전체동기화(리콘실) — 네이버 라이브와 DB 카운트 일치. 백그라운드 실행."""
    def post(self, request):
        from . import smartstore_reconcile_service as _r
        body = request.data or {}
        return Response(_r.start_reconcile(
            apply=bool(body.get('apply', True)),
            max_delete_ratio=float(body.get('max_delete_ratio', 0.5)),
            force=bool(body.get('force', False)),
        ))

    def get(self, request):
        from . import smartstore_reconcile_service as _r
        return Response(_r.get_status())


class SmartStoreProductOrdersView(APIView):
    def get(self, request):
        code = request.query_params.get('code', '')
        product_name = request.query_params.get('product_name', '')
        start_date = request.query_params.get('start_date', '')
        end_date = request.query_params.get('end_date', '')
        if not code and not product_name:
            return Response({'error': 'code or product_name required'}, status=400)
        result = smartstore_order_service.get_product_orders(
            seller_code=code or None,
            product_name=product_name or None,
            start_date=start_date or None,
            end_date=end_date or None,
        )
        return Response(result)


# ── 스마트스토어 분석 ──

class SmartStoreAnalyticsOverviewView(APIView):
    def get(self, request):
        start_date = request.query_params.get('start_date') or None
        end_date = request.query_params.get('end_date') or None
        return Response(smartstore_analytics_service.get_overview(start_date, end_date))


class SmartStoreAnalyticsStoreDetailView(APIView):
    def get(self, request, store_id):
        start_date = request.query_params.get('start_date') or None
        end_date = request.query_params.get('end_date') or None
        period = request.query_params.get('period', 'monthly')
        return Response(smartstore_analytics_service.get_store_detail(
            store_id, start_date, end_date, period,
        ))


class SmartStoreAnalyticsBusinessDetailView(APIView):
    def get(self, request, code):
        start_date = request.query_params.get('start_date') or None
        end_date = request.query_params.get('end_date') or None
        period = request.query_params.get('period', 'monthly')
        return Response(smartstore_analytics_service.get_business_detail(
            code, start_date, end_date, period,
        ))


class SmartStoreAnalyticsSyncCategoriesView(APIView):
    def post(self, request):
        result = smartstore_analytics_service.sync_category_names()
        status_code = 400 if 'error' in result else 200
        return Response(result, status=status_code)


class SmartStoreRegistrationLimitView(APIView):
    def get(self, request):
        return Response(smartstore_analytics_service.get_registration_limits())


class SmartStorePolicyCollectView(APIView):
    """POST /api/smartstore/policy/collect/  — 상품등록한도 API 수집 시작.
    body: {login_ids?: [int], concurrency?: int}  미지정 시 전체 로그인·순차(동시 1).
    주의: 동시>1 은 같은 디스플레이 크롬 충돌 + 네이버 동시로그인 차단으로 수집 실패율↑ → 순차 권장."""
    def post(self, request):
        from . import naver_store_policy_service as policy
        login_ids = request.data.get('login_ids') or None
        try:
            concurrency = int(request.data.get('concurrency') or 1)
        except (TypeError, ValueError):
            concurrency = 1
        concurrency = max(1, min(concurrency, 8))
        return Response(policy.start_collect(login_ids=login_ids, concurrency=concurrency))


class SmartStorePolicyStatusView(APIView):
    """GET /api/smartstore/policy/status/  — 수집 진행 상태/로그."""
    def get(self, request):
        from . import naver_store_policy_service as policy
        return Response(policy.get_status())


class SmartStorePolicyTrendView(APIView):
    """GET /api/smartstore/policy/trend/?store_id=<id>  — 정책 스냅샷 시계열."""
    def get(self, request):
        sid = request.query_params.get('store_id')
        sid = int(sid) if sid and sid.isdigit() else None
        return Response(smartstore_analytics_service.get_policy_trend(sid))


# ── 스토어 상품수집 (브라우저 자동화) ──

from . import store_collector


class StoreCollectStartView(APIView):
    def post(self, request):
        store_ids = request.data.get('store_ids')
        if store_ids is not None and not isinstance(store_ids, list):
            store_ids = [int(store_ids)]
        ok, msg = store_collector.start(store_ids)
        if not ok:
            return Response({'ok': False, 'message': msg}, status=409)
        return Response({'ok': True, 'message': msg})


class StoreCollectStatusView(APIView):
    def get(self, request):
        since = int(request.query_params.get('logSince', 0))
        st = store_collector.get_status()
        if since > 0:
            st['logs'] = [l for i, l in enumerate(st['logs']) if i >= since]
        return Response(st)


class StoreCollectStopView(APIView):
    def post(self, request):
        store_collector.stop()
        return Response({'ok': True})


class StoreCollectCsvView(APIView):
    def get(self, request):
        # store_name 또는 log_id로 CSV 다운로드
        store_name = request.query_params.get('store_name', '')
        log_id = request.query_params.get('log_id')

        file_path = None
        if log_id:
            # DB에서 파일 경로 조회
            logs = store_collector.get_collect_logs(limit=100)
            for log in logs:
                if str(log['id']) == str(log_id):
                    file_path = log.get('csv_file_path')
                    break
        elif store_name:
            file_path = store_collector.get_csv_file(store_name)

        if not file_path or not os.path.exists(file_path):
            return Response({'error': 'CSV 파일이 없습니다.'}, status=404)
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='text/csv; charset=utf-8-sig')
            fname = quote(os.path.basename(file_path))
            response['Content-Disposition'] = f"attachment; filename*=UTF-8''{fname}"
            return response


class StoreCollectLogsView(APIView):
    def get(self, request):
        limit = int(request.query_params.get('limit', 20))
        logs = store_collector.get_collect_logs(limit=limit)
        # datetime 직렬화
        for log in logs:
            if log.get('completed_at'):
                log['completed_at'] = str(log['completed_at'])
        return Response(logs)

    def delete(self, request):
        """실패 로그 삭제"""
        from django.db import connections
        with connections['myproduct'].cursor() as cur:
            cur.execute("DELETE FROM store_collect_log WHERE status = 'error'")
            deleted = cur.rowcount
        return Response({'deleted': deleted})


# ── 전상품 API 검증 ──

class ProductAuditStartView(APIView):
    def post(self, request):
        source = request.data.get('source', 'api')
        result = product_audit_service.start_audit(source=source)
        status = 200 if result['ok'] else 409
        return Response(result, status=status)


class ProductAuditStatusView(APIView):
    def get(self, request):
        return Response(product_audit_service.get_audit_status())


class ProductAuditStopView(APIView):
    def post(self, request):
        product_audit_service.stop_audit()
        return Response({'ok': True})


class ProductAuditLogsView(APIView):
    def get(self, request):
        limit = int(request.query_params.get('limit', 20))
        return Response(product_audit_service.get_audit_logs(limit=limit))


class ProductAuditLogDetailView(APIView):
    def get(self, request, pk):
        return Response(product_audit_service.get_audit_log_detail(pk))


# ── 상세페이지 크롤링 ──

class DetailCrawlStartView(APIView):
    def post(self, request):
        from smartstore.smartstore_product_service import start_detail_crawl
        batch_size = int(request.data.get('batch_size', 0))
        result = start_detail_crawl(batch_size=batch_size)
        status = 200 if result['ok'] else 409
        return Response(result, status=status)


class DetailCrawlStatusView(APIView):
    def get(self, request):
        from smartstore.smartstore_product_service import get_detail_crawl_status
        return Response(get_detail_crawl_status())


class DetailCrawlStopView(APIView):
    def post(self, request):
        from smartstore.smartstore_product_service import stop_detail_crawl
        stop_detail_crawl()
        return Response({'ok': True})


# ── 상품 편집 ──

class ProductFullDetailView(APIView):
    """상품 편집용 전체 상세 조회 (네이버 API v2 실시간)"""
    def get(self, request, opno):
        store_id = request.query_params.get('store_id')
        if not store_id:
            return Response({'error': 'store_id required'}, status=400)
        try:
            result = smartstore_product_service.get_product_full_detail(int(opno), int(store_id))
        except Exception as e:
            return Response({'error': str(e)}, status=500)
        if 'error' in result:
            return Response(result, status=400)
        return Response(result)


class ProductUpdateView(APIView):
    """상품 필드 수정 (GET→modify→PUT)"""
    def put(self, request, opno):
        store_id = request.data.get('store_id')
        if not store_id:
            return Response({'error': 'store_id required'}, status=400)
        updates = request.data.get('updates', {})
        if not updates:
            return Response({'error': 'updates required'}, status=400)
        try:
            result = smartstore_product_service.update_product_fields(int(opno), int(store_id), updates)
        except Exception as e:
            return Response({'error': str(e)}, status=500)
        if 'error' in result:
            return Response(result, status=400)
        return Response(result)


class ProductImageUploadView(APIView):
    """상품 이미지 네이버 CDN 업로드"""
    def post(self, request):
        store_id = request.data.get('store_id')
        if not store_id:
            return Response({'error': 'store_id required'}, status=400)
        image_file = request.FILES.get('image')
        if not image_file:
            return Response({'error': 'image file required'}, status=400)
        try:
            result = smartstore_product_service.upload_product_image(int(store_id), image_file)
        except Exception as e:
            return Response({'error': str(e)}, status=500)
        if 'error' in result:
            return Response(result, status=400)
        return Response(result)


class ZeroMarginPreviewView(APIView):
    """역마진 상품 0마진 가격 미리보기"""
    def get(self, request):
        store_id = request.query_params.get('store_id', 0)
        result = smartstore_product_service.zero_margin_preview(int(store_id))
        return Response(result)


class ZeroMarginUpdateView(APIView):
    """역마진 상품 가격을 0마진으로 일괄 수정"""
    def post(self, request):
        store_id = request.data.get('store_id', 0)
        try:
            result = smartstore_product_service.zero_margin_update(int(store_id))
        except Exception as e:
            return Response({'error': str(e)}, status=500)
        return Response(result)


class ZeroMarginLogsView(APIView):
    """0마진 처리 이력"""
    def get(self, request):
        limit = int(request.query_params.get('limit', 20))
        return Response(smartstore_product_service.zero_margin_logs(limit))


class ZeroMarginLogDetailView(APIView):
    """0마진 처리 상세"""
    def get(self, request, pk):
        items = smartstore_product_service.zero_margin_log_detail(pk)
        return Response({'items': items})


# ── 상품속성 분석 (크롤 결과 조회) ──

class AttrStatsView(APIView):
    def get(self, request):
        return Response(attr_analytics_service.get_stats())


class AttrProductsListView(APIView):
    def get(self, request):
        q = request.query_params
        page = int(q.get('page', 1))
        per_page = min(int(q.get('per_page', 50)), 200)
        store_id = q.get('store_id')
        needs_review = q.get('needs_review')
        has_quality = q.get('has_quality')
        return Response(attr_analytics_service.list_products(
            page=page, per_page=per_page, search=q.get('search', ''),
            store_id=int(store_id) if store_id else None,
            needs_review=int(needs_review) if needs_review else None,
            has_quality=int(has_quality) if has_quality else None,
        ))


class AttrProductDetailView(APIView):
    def get(self, request, seller_code):
        store_id = request.query_params.get('store_id')
        data = attr_analytics_service.get_product_detail(
            seller_code, int(store_id) if store_id else None
        )
        if not data:
            return Response({'error': 'not found'}, status=404)
        return Response(data)


class AttrTopTagsView(APIView):
    def get(self, request):
        q = request.query_params
        return Response({'items': attr_analytics_service.top_tags(
            limit=min(int(q.get('limit', 30)), 200),
            by=q.get('by', 'count'),
            category_id=q.get('category_id') or None,
        )})


class AttrQualityIssuesView(APIView):
    def get(self, request):
        limit = min(int(request.query_params.get('limit', 200)), 1000)
        return Response({'items': attr_analytics_service.quality_issues(limit)})


class AttrCategorySummaryView(APIView):
    def get(self, request):
        limit = min(int(request.query_params.get('limit', 50)), 500)
        return Response({'items': attr_analytics_service.category_summary(limit)})


# ── 빈 속성 검토 + 등록 ──

class MissingAttrsSummaryView(APIView):
    def get(self, request):
        return Response(missing_attrs_service.get_summary())


class MissingAttrsListView(APIView):
    def get(self, request):
        q = request.query_params
        return Response(missing_attrs_service.list_attributes(
            page=int(q.get('page', 1)),
            per_page=min(int(q.get('per_page', 50)), 200),
            search=q.get('search', ''),
            kind=q.get('kind', 'all'),
            status=q.get('status', 'pending'),
            sort=q.get('sort', 'count'),
        ))


class MissingAttrsSkusView(APIView):
    def get(self, request, attribute_seq):
        q = request.query_params
        return Response(missing_attrs_service.list_skus_for_attribute(
            attribute_seq=int(attribute_seq),
            page=int(q.get('page', 1)),
            per_page=min(int(q.get('per_page', 100)), 500),
            store_id=q.get('store_id'),
            category_id=q.get('category_id') or None,
            status=q.get('status', 'pending'),
            search=q.get('search', ''),
        ))


class MissingAttrsRegisterFilteredView(APIView):
    """필터 매칭 모든 SKU 에 일괄 등록 — body: {attribute_seq, value_seq, value_text, store_id, category_id, search, dry_run, max_skus}"""
    def post(self, request):
        d = request.data or {}
        try:
            r = missing_attrs_service.register_filtered(
                attribute_seq=int(d['attribute_seq']),
                value_seq=int(d.get('value_seq', 0)),
                value_text=d.get('value_text'),
                store_id=d.get('store_id'),
                category_id=d.get('category_id') or None,
                search=d.get('search', ''),
                max_skus=int(d.get('max_skus', 2000)),
                dry_run=bool(d.get('dry_run', False)),
            )
            return Response(r)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class MissingAttrsRefreshSummaryView(APIView):
    """attr_missing_summary + sku_missing_summary 재계산."""
    def post(self, request):
        from django.db import connections
        attr_count = missing_attrs_service.refresh_attr_summary()
        with connections['myproduct'].cursor() as c:
            c.execute("DELETE FROM smartstore_sku_missing_summary")
            c.execute("""
                INSERT INTO smartstore_sku_missing_summary
                  (seller_management_code, store_id, category_id,
                   missing_count, auto_count, free_count, registered_count, pending_count)
                SELECT
                  m.seller_management_code, m.store_id, MAX(m.category_id),
                  COUNT(*),
                  SUM(CASE WHEN m.candidate_count=1 AND m.status='pending' THEN 1 ELSE 0 END),
                  SUM(CASE WHEN m.candidate_count=0 THEN 1 ELSE 0 END),
                  SUM(CASE WHEN m.status='registered' THEN 1 ELSE 0 END),
                  SUM(CASE WHEN m.status='pending' THEN 1 ELSE 0 END)
                FROM smartstore_product_missing_attrs m
                GROUP BY m.seller_management_code, m.store_id
            """)
            sku_count = c.rowcount
        return Response({'attr_summary_rows': attr_count, 'sku_summary_rows': sku_count})


class MissingAttrsDetailView(APIView):
    def get(self, request, attribute_seq):
        return Response(missing_attrs_service.get_attribute_detail(int(attribute_seq)))


class MissingAttrsRegisterView(APIView):
    """일괄 등록 — body: {attribute_seq, value_seq, value_text, skus: [{...}], dry_run}"""
    def post(self, request):
        d = request.data or {}
        skus = d.get('skus', [])
        if not skus:
            return Response({'error': 'skus required'}, status=400)
        try:
            result = missing_attrs_service.register_bulk(
                skus=skus,
                attribute_seq=int(d['attribute_seq']),
                value_seq=int(d['value_seq']),
                value_text=d.get('value_text'),
                dry_run=bool(d.get('dry_run', False)),
            )
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class MissingAttrsSkuListView(APIView):
    """상품별 — 빈속성 개수 정렬"""
    def get(self, request):
        q = request.query_params
        store_id = q.get('store_id')
        return Response(missing_attrs_service.list_skus(
            page=int(q.get('page', 1)),
            per_page=min(int(q.get('per_page', 50)), 200),
            search=q.get('search', ''),
            store_id=int(store_id) if store_id else None,
            status=q.get('status', 'pending'),
            category_id=q.get('category_id') or None,
        ))


class MissingAttrsForSkuView(APIView):
    """단일 SKU 의 빈속성 (모달용)"""
    def get(self, request, seller_code):
        store_id = int(request.query_params.get('store_id', 0))
        return Response(missing_attrs_service.get_missing_for_sku(seller_code, store_id))


class MissingAttrsRegisterSkuView(APIView):
    """단일 SKU 다중 속성 일괄 등록 — body: {store_id, selections, dry_run}"""
    def post(self, request, seller_code):
        d = request.data or {}
        try:
            r = missing_attrs_service.register_for_sku(
                seller_management_code=seller_code,
                store_id=int(d['store_id']),
                selections=d.get('selections', []),
                dry_run=bool(d.get('dry_run', False)),
            )
            return Response(r)
        except Exception as e:
            return Response({'ok': False, 'error': str(e)}, status=500)


class MissingAttrsMarkView(APIView):
    """skipped/reviewed 마킹 — body: {attribute_seq, status, skus: [{seller, store_id}]}"""
    def post(self, request):
        d = request.data or {}
        skus = [(s['seller_management_code'], s['store_id']) for s in d.get('skus', [])]
        if not skus:
            return Response({'error': 'skus required'}, status=400)
        result = missing_attrs_service.mark_status(
            seller_codes_with_store=skus,
            attribute_seq=int(d['attribute_seq']),
            status=d.get('status', 'skipped'),
            value_seq=d.get('value_seq'),
            value_text=d.get('value_text'),
        )
        return Response(result)


# ── 워커 대시보드 ──

class WorkerDashboardView(APIView):
    def get(self, request):
        return Response({
            'workers': worker_dashboard_service.get_workers_status(),
            'aggregate': worker_dashboard_service.get_aggregate_status(),
        })


class AttrTopAttributesView(APIView):
    def get(self, request):
        q = request.query_params
        return Response({'items': attr_analytics_service.top_attributes(
            limit=min(int(q.get('limit', 200)), 1000),
            section=q.get('section') or None,
            category_id=q.get('category_id') or None,
        )})


class AttrValuesView(APIView):
    def get(self, request):
        q = request.query_params
        attr_label = q.get('attr_label')
        if not attr_label:
            return Response({'error': 'attr_label required'}, status=400)
        return Response({'items': attr_analytics_service.attribute_values(
            attr_label=attr_label,
            section=q.get('section') or None,
            category_id=q.get('category_id') or None,
            limit=min(int(q.get('limit', 100)), 500),
        )})


# ── 네이버 나의상품 (11번가 my_product 미러) ──────────────────────

class NaverMyProductFolderListView(APIView):
    def get(self, request):
        return Response({'items': naver_my_product_service.list_folders()})


class NaverMyProductFolderSyncView(APIView):
    """smartstoreIdList 스토어 → 폴더 자동 생성."""
    def post(self, request):
        return Response(naver_my_product_service.ensure_folders_from_stores())


class NaverMyProductImportFrom11stView(APIView):
    """[11번가 나의상품 가져오기] 버튼.
    백그라운드로 ads.my_product 전체를 naverdb.naver_my_product 로 UPSERT.
    """
    def post(self, request):
        batch_size = int(request.data.get('batch_size', 1000)) if request.data else 1000
        return Response(naver_my_product_service.start_import_from_11st(batch_size=batch_size))


class NaverMyProductImportStatusView(APIView):
    def get(self, request):
        return Response(naver_my_product_service.get_import_status())


class NaverMyProductListView(APIView):
    def get(self, request):
        q = request.query_params
        folder_id = q.get('folder_id')
        reg = q.get('registered')
        return Response(naver_my_product_service.get_products(
            page=int(q.get('page', 1)),
            per_page=int(q.get('per_page', 50)),
            folder_id=int(folder_id) if folder_id not in (None, '', 'all') else None,
            search=q.get('search') or None,
            sort=q.get('sort') or 'id_desc',
            include_sales=q.get('include_sales') == '1',
            category_code=q.get('category_code') or None,
            registered=int(reg) if reg in ('0', '1') else None,
            register_stage=q.get('register_stage') or None,
        ))


class NaverMyProductGenerateNameView(APIView):
    """단건 네이버 상품명 생성 (이미지 분석 캐싱 + Ollama 텍스트, 동기).

    use_vision 기본값:
      - body 에 명시 안 하면 False (벤치마크 모드, 텍스트만)
      - 단건 [🤖 재생성] 모달 호출은 명시적으로 True 보냄
    ollama_url:
      - 11st 워커가 자기 endpoint 를 body 로 전달 → 분산 호출
      - 없으면 backend default (localhost:11438)
    """
    def post(self, request, pk):
        body = request.data or {}
        use_vision = bool(body.get('use_vision', False))
        ollama_url = body.get('ollama_url') or None
        if ollama_url and not ollama_url.startswith('http'):
            ollama_url = f'http://{ollama_url}'
        result = naver_name_generator.generate_naver_name(
            int(pk), use_vision=use_vision, url=ollama_url)
        status = 200 if result.get('ok') else 502
        return Response(result, status=status)


class NaverMyProductAnalyzeImageView(APIView):
    """단건 이미지 분석만 (비전 모델). 캐시 우선, ?force=1 이면 재분석."""
    def post(self, request, pk):
        force = (request.data or {}).get('force', False) if request.data else False
        result = naver_vision_analyzer.analyze_product_image(int(pk), force=bool(force))
        status = 200 if result.get('ok') else 502
        return Response(result, status=status)


class NaverMyProductEnqueueView(APIView):
    """일괄 enqueue (ads.ai_keyword_task, platform='naver'). 워커 11개가 분산 처리.
    body: {ids:[...]} 또는 {folder_id:N} 또는 {top_sales:500} 또는 {} (=빠진 전체).
    """
    def post(self, request):
        d = request.data or {}
        ids = d.get('ids')
        folder_id = d.get('folder_id')
        top_sales = d.get('top_sales')
        only_missing = d.get('only_missing', True)
        result = naver_my_product_service.enqueue_products(
            ids=ids,
            folder_id=int(folder_id) if folder_id is not None else None,
            top_sales=int(top_sales) if top_sales is not None else None,
            only_missing=bool(only_missing),
        )
        return Response(result)


class NaverMyProductQueueStatusView(APIView):
    """큐 상태 + 워커별 처리 현황 (최근 1시간)."""
    def get(self, request):
        return Response(naver_my_product_service.get_queue_status())


class NaverDispatchTickView(APIView):
    """11st dispatch 데몬이 5초마다 호출 — 워커별 buffer 자동 보충.
    body: {endpoints: [...], target_buffer: 30}
    """
    authentication_classes = []
    permission_classes = []
    def post(self, request):
        d = request.data or {}
        endpoints = d.get('endpoints') or []
        target = int(d.get('target_buffer', 30))
        return Response(naver_my_product_service.dispatch_tick(endpoints, target))


class NaverFolderQueueView(APIView):
    """GET — 현재 큐에 들어간 폴더 list + 진척률
       POST {folder_ids: [...]} — 큐 순서 갱신"""
    def get(self, request):
        return Response({'items': naver_my_product_service.list_folder_queue()})

    def post(self, request):
        d = request.data or {}
        ids = d.get('folder_ids')
        force = bool(d.get('force_regenerate', False))
        if not isinstance(ids, list):
            return Response({'ok': False, 'error': 'folder_ids list required'}, status=400)
        ids = [int(x) for x in ids if str(x).isdigit()]
        return Response(naver_my_product_service.set_folder_queue_positions(ids, force_regenerate=force))


class NaverMyProductDetailView(APIView):
    """단건 전체 정보 조회 + 편집 (PATCH).
    GET   /naver-products/<pk>/        → 전체 컬럼 + image_analysis JSON
    PATCH /naver-products/<pk>/        → body 의 화이트리스트 필드 UPDATE
    DELETE /naver-products/<pk>/image-analysis/ 는 별도 뷰
    """
    def get(self, request, pk):
        d = naver_my_product_service.get_product_detail(int(pk))
        if not d:
            return Response({'ok': False, 'error': 'not_found'}, status=404)
        return Response(d)

    def patch(self, request, pk):
        r = naver_my_product_service.patch_product(int(pk), request.data or {})
        status = 200 if r.get('ok') else (404 if r.get('error') == 'not_found' else 400)
        return Response(r, status=status)


class NaverMyProductClearVisionView(APIView):
    """비전 분석 캐시 삭제."""
    def post(self, request, pk):
        r = naver_my_product_service.clear_image_analysis(int(pk))
        return Response(r, status=200 if r.get('ok') else 404)


class NaverMyProductKeywordPoolView(APIView):
    """편집 모달용 키워드 풀 (비전 + 11번가 ad_keyword + naver preset + 조회수 캐시)."""
    def get(self, request, pk):
        r = naver_my_product_service.get_keyword_pool(int(pk))
        return Response(r, status=200 if r.get('ok') else 404)


class NaverMyProductEditThumbView(APIView):
    """[Deprecated 호환] 단일 썸네일 저장 → 풀에 추가 후 활성화로 변환.
    POST   { image_b64 } → variant 풀에 추가 + 활성화
    DELETE                → 전체 비활성화 (원본 image_large 복귀)
    """
    def post(self, request, pk):
        from . import naver_thumbnail_variant_service as _v
        body = request.data or {}
        b64 = body.get('image_b64', '')
        if not b64:
            return Response({'ok': False, 'error': 'image_b64 required'}, status=400)
        r = _v.add_variant(int(pk), b64, source_type='manual', activate=True)
        status = 200 if r.get('ok') else (404 if r.get('error') == 'not_found' else 400)
        # 구 API 호환: edited_image_url 키 유지
        if r.get('ok'):
            r['edited_image_url'] = r.get('image_url')
        return Response(r, status=status)

    def delete(self, request, pk):
        from . import naver_thumbnail_variant_service as _v
        r = _v.deactivate_all(int(pk))
        return Response(r, status=200 if r.get('ok') else 404)


class NaverMyProductVariantListView(APIView):
    """제품 썸네일 변형 풀 (최대 20개).
    GET    /naver-products/<pk>/variants/                    → list
    POST   /naver-products/<pk>/variants/                    → add (body: image_b64, source_type, source_meta?, label?, activate?)
    DELETE /naver-products/<pk>/variants/                    → delete ALL (확인 후 사용)
    """
    def get(self, request, pk):
        from . import naver_thumbnail_variant_service as _v
        return Response(_v.list_variants(int(pk)))

    def post(self, request, pk):
        from . import naver_thumbnail_variant_service as _v
        body = request.data or {}
        b64 = body.get('image_b64', '')
        if not b64:
            return Response({'ok': False, 'error': 'image_b64 required'}, status=400)
        r = _v.add_variant(
            int(pk), b64,
            source_type=body.get('source_type', 'manual'),
            source_meta=body.get('source_meta'),
            label=body.get('label'),
            activate=bool(body.get('activate', True)),
        )
        return Response(r, status=200 if r.get('ok') else 400)

    def delete(self, request, pk):
        from . import naver_thumbnail_variant_service as _v
        return Response(_v.delete_all(int(pk)))


class NaverMyProductVariantDetailView(APIView):
    """단일 변형 — 삭제 / 라벨 수정.
    DELETE /naver-products/<pk>/variants/<vid>/              → 삭제
    PATCH  /naver-products/<pk>/variants/<vid>/              → body: { label }
    """
    def delete(self, request, pk, vid):
        from . import naver_thumbnail_variant_service as _v
        r = _v.delete_variant(int(pk), int(vid))
        return Response(r, status=200 if r.get('ok') else 404)

    def patch(self, request, pk, vid):
        from . import naver_thumbnail_variant_service as _v
        body = request.data or {}
        r = _v.update_label(int(pk), int(vid), body.get('label', ''))
        return Response(r, status=200 if r.get('ok') else 404)


class NaverUpscaleJobsView(APIView):
    def get(self, request):
        from . import naver_upscale_dispatcher as _d
        return Response(_d.list_jobs(limit=int(request.query_params.get('limit', 20))))

    def post(self, request):
        from . import naver_upscale_dispatcher as _d
        body = request.data or {}
        fid = body.get('folder_id')
        if fid is not None: fid = int(fid)
        r = _d.start_job(folder_id=fid, scale=float(body.get('scale', 1.5)),
                         model_family=body.get('model_family', 'sd15'),
                         workers=body.get('workers'))
        return Response(r, status=200 if r.get('ok') else 400)


class NaverUpscaleJobDetailView(APIView):
    def get(self, request, pk):
        from . import naver_upscale_dispatcher as _d
        r = _d.get_job(int(pk))
        return Response({'ok': bool(r), 'job': r} if r else {'ok': False, 'error': 'not_found'},
                        status=200 if r else 404)


class NaverUpscaleJobProgressView(APIView):
    def get(self, request, pk):
        from . import naver_upscale_dispatcher as _d
        return Response(_d.get_job_progress(int(pk)))


class NaverUpscaleJobControlView(APIView):
    def post(self, request, pk, action):
        from . import naver_upscale_dispatcher as _d
        if action == 'pause':  return Response(_d.pause_job(int(pk)))
        if action == 'resume': return Response(_d.resume_job(int(pk)))
        if action == 'cancel': return Response(_d.cancel_job(int(pk)))
        return Response({'ok': False, 'error': 'invalid_action'}, status=400)


class NaverUpscaleWorkersView(APIView):
    def get(self, request):
        from . import naver_upscale_dispatcher as _d
        return Response(_d.list_workers())


class NaverUpscaleStorageStatsView(APIView):
    """218 sshfs mount 파일 수 + 분당 증가 rate (60s sliding window).
    GET /api/smartstore/upscale/storage-stats/
    """
    _history: list = []  # list of (timestamp, count)

    def get(self, request):
        import os, time
        from . import naver_upscale_dispatcher as _d
        try:
            count = len(os.listdir(_d.STORAGE_LOCAL_DIR))
        except OSError:
            count = 0
        now = time.time()
        hist = self.__class__._history
        hist.append((now, count))
        # 60초 보다 오래된 거 제거
        while hist and hist[0][0] < now - 65:
            hist.pop(0)
        # 60초 전 데이터로 rate 계산
        rate_per_min = 0
        if len(hist) >= 2:
            old_t, old_c = hist[0]
            elapsed = now - old_t
            if elapsed > 0:
                rate_per_min = round((count - old_c) * 60.0 / elapsed)
        return Response({
            'ok': True,
            'count': count,
            'rate_per_min': rate_per_min,
            'window_s': round(now - hist[0][0], 1) if hist else 0,
            'host': '218 (sshfs mount)',
        })


class NaverMyProductVariantActivateView(APIView):
    """변형 활성화 — edited_image_url 미러링.
    POST /naver-products/<pk>/variants/<vid>/activate/
    """
    def post(self, request, pk, vid):
        from . import naver_thumbnail_variant_service as _v
        r = _v.activate_variant(int(pk), int(vid))
        return Response(r, status=200 if r.get('ok') else 404)


class NaverKeywordRelatedView(APIView):
    """GET /api/smartstore/naver-keyword-related/?seed=후드&limit=30
    또는 ?seeds=후드,자켓,여성&limit=1500  (multi-seed 대량 수집)
    네이버 검색광고 연관키워드 — 조회수 desc 정렬."""
    def get(self, request):
        seeds_str = (request.query_params.get('seeds') or '').strip()
        seed = (request.query_params.get('seed') or '').strip()
        from . import keyword_volume_service as kvs
        if seeds_str:
            seeds = [s.strip() for s in seeds_str.split(',') if s.strip()]
            limit = int(request.query_params.get('limit', 1500))
            items = kvs.get_related_keywords_multi(seeds, limit=limit)
            return Response({'ok': True, 'seeds': seeds, 'items': items})
        if not seed:
            return Response({'ok': False, 'error': 'seed or seeds required'}, status=400)
        limit = int(request.query_params.get('limit', 30))
        items = kvs.get_related_keywords(seed, limit=limit)
        return Response({'ok': True, 'seed': seed, 'items': items})


class NaverDupBaseView(APIView):
    """중복 검사 3-mode 공통 — query params: ?folder_id, ?page, ?per_page, ?refresh=1, ?strength=exact|token"""
    mode: str = ''
    def get(self, request):
        from . import naver_dup_service as ds
        q = request.query_params
        fid = q.get('folder_id')
        xfid = q.get('exclude_folder_id')
        kwargs = {
            'folder_id': int(fid) if fid not in (None, '', 'all') else None,
            'exclude_folder_id': int(xfid) if xfid not in (None, '', '0') else None,
            'page': int(q.get('page', 1)),
            'per_page': int(q.get('per_page', 50)),
            'refresh': q.get('refresh') == '1',
        }
        if self.mode == 'name':
            kwargs['strength_filter'] = q.get('strength') or None
            r = ds.by_name(**kwargs)
        elif self.mode == 'image':
            r = ds.by_image(**kwargs)
        elif self.mode == 'detail':
            r = ds.by_detail(**kwargs)
        else:
            return Response({'error': 'unknown mode'}, status=400)
        return Response(r)


class NaverDupByNameView(NaverDupBaseView):
    mode = 'name'


class NaverDupByImageView(NaverDupBaseView):
    mode = 'image'


class NaverDupByDetailView(NaverDupBaseView):
    mode = 'detail'


class NaverDupMarkView(APIView):
    """POST {ids:[...], action: 'mark'|'delete'|'dismiss'|'undismiss', mode?:'name|image|detail|all'}"""
    def post(self, request):
        from . import naver_dup_service as ds
        d = request.data or {}
        ids = d.get('ids') or []
        action = d.get('action', 'mark')
        if action == 'delete':
            return Response(ds.delete_products(ids))
        if action == 'dismiss':
            return Response(ds.dismiss_as_not_dup(ids, mode=d.get('mode', 'all')))
        if action == 'undismiss':
            return Response(ds.undismiss(ids))
        return Response(ds.mark_for_deletion(ids))


class NaverQueueTgReportView(APIView):
    """GET/POST 텔레그램 큐 보고 ON/OFF.
    GET → {enabled: bool}
    POST {enabled: bool} → 플래그 파일 생성/삭제 + 즉시 1회 보고(ON 일때)
    """
    FLAG = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.tg_reporter_enabled')

    def get(self, request):
        return Response({'enabled': os.path.exists(self.FLAG)})

    def post(self, request):
        d = request.data or {}
        enable = bool(d.get('enabled'))
        if enable:
            open(self.FLAG, 'w').close()
            # 즉시 1회 보고
            import subprocess
            try:
                subprocess.Popen(
                    ['python3', '-u', 'naver_queue_reporter.py'],
                    cwd=os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                    env={**os.environ, 'DJANGO_SETTINGS_MODULE': 'config.settings'},
                )
            except Exception as e:
                return Response({'ok': True, 'enabled': True, 'warn': f'즉시보고 실패: {e}'})
            return Response({'ok': True, 'enabled': True, 'sent_now': True})
        else:
            if os.path.exists(self.FLAG):
                os.remove(self.FLAG)
            return Response({'ok': True, 'enabled': False})


class NaverDupGroupView(APIView):
    """GET ?group_key=...&mode=name|image|detail&folder_id=... — 그룹 멤버 조회"""
    def get(self, request):
        from . import naver_dup_service as ds
        q = request.query_params
        gk = q.get('group_key')
        mode = q.get('mode', 'name')
        fid = q.get('folder_id')
        if not gk:
            return Response({'error': 'group_key required'}, status=400)
        return Response(ds.group_members(
            gk, mode,
            folder_id=int(fid) if fid not in (None, '', 'all') else None,
        ))


class NaverKeywordRelevanceView(APIView):
    """POST /api/smartstore/naver-products/<pk>/keyword-relevance/
    body: { keywords: [...], force?: bool }
    LLM 으로 키워드 적합도 검증 — 캐시 hit 시 즉시. naver_my_product 컨텍스트 활용.
    """
    def post(self, request, pk):
        d = request.data or {}
        kws = d.get('keywords') or []
        force = bool(d.get('force', False))
        if not kws or not isinstance(kws, list):
            return Response({'ok': False, 'error': 'keywords (list) required'}, status=400)
        from . import keyword_relevance_service as krs
        return Response(krs.get_relevance(int(pk), kws, force=force))


class NaverMyProductConfirmNameView(APIView):
    """POST /api/smartstore/naver-products/<pk>/confirm-name/
    body: { before_name, after_name, comment?, comment_type: wrong|missing|overall }
    AI 생성 상품명 컨펌 저장 — W코드 raw 누적 + 카테고리별 키워드 정책 학습.
    """
    def post(self, request, pk):
        from . import naver_name_confirmation_service as ncs
        d = request.data or {}
        r = ncs.save_confirmation(
            product_id=int(pk),
            before_name=d.get('before_name'),
            after_name=d.get('after_name'),
            comment=d.get('comment'),
            comment_type=d.get('comment_type', 'overall'),
        )
        status = 200 if r.get('ok') else (404 if r.get('error') == 'not_found' else 400)
        return Response(r, status=status)


class NaverMyProductConfirmationsView(APIView):
    """GET /api/smartstore/naver-products/<pk>/confirmations/
    해당 상품(W코드)의 컨펌 히스토리.
    """
    def get(self, request, pk):
        from . import naver_name_confirmation_service as ncs
        from django.db import connections
        with connections['naverdb'].cursor() as cur:
            cur.execute("SELECT product_code FROM naver_my_product WHERE id=%s", [int(pk)])
            row = cur.fetchone()
            if not row:
                return Response({'ok': False, 'error': 'not_found'}, status=404)
        r = ncs.list_confirmations(
            product_code=row[0],
            limit=int(request.query_params.get('limit', 20)),
            offset=int(request.query_params.get('offset', 0)),
        )
        return Response({'ok': True, **r})


class WcodeBlacklistView(APIView):
    """GET  /api/smartstore/wcode-blacklist/?search=&only_unprocessed=&page=&per_page=
    POST /api/smartstore/wcode-blacklist/ body: { codes: 'W001 W002 ...' | [...], reason?, source? }
    DELETE /api/smartstore/wcode-blacklist/?codes=W001,W002  (지정 해제, 실제 상품 삭제 아님)
    """
    def get(self, request):
        from . import wcode_blacklist_service as svc
        q = request.query_params
        return Response(svc.list_all(
            search=q.get('search') or None,
            only_unprocessed=q.get('only_unprocessed') == '1',
            page=int(q.get('page', 1)),
            per_page=int(q.get('per_page', 50)),
        ))

    def post(self, request):
        from . import wcode_blacklist_service as svc
        d = request.data or {}
        raw = d.get('codes')
        if isinstance(raw, list):
            codes = []
            seen: set = set()
            for x in raw:
                c = svc._normalize_code(str(x))
                if c and c not in seen:
                    seen.add(c)
                    codes.append(c)
        else:
            codes = svc._parse_codes(str(raw or ''))
        if not codes:
            return Response({'ok': False, 'error': 'no valid W-codes'}, status=400)
        r = svc.bulk_upsert(codes, reason=d.get('reason') or None,
                            source=d.get('source') or 'manual')
        return Response(r)

    def delete(self, request):
        from . import wcode_blacklist_service as svc
        raw = request.query_params.get('codes') or (request.data or {}).get('codes') or ''
        if isinstance(raw, list):
            codes = [svc._normalize_code(str(x)) for x in raw if x]
            codes = [c for c in codes if c]
        else:
            codes = svc._parse_codes(str(raw))
        if not codes:
            return Response({'ok': False, 'error': 'codes required'}, status=400)
        return Response(svc.delete(codes))


class WcodeBlacklistUploadExcelView(APIView):
    """POST /api/smartstore/wcode-blacklist/upload-excel/  multipart file 필드 'file'"""
    def post(self, request):
        from . import wcode_blacklist_service as svc
        f = request.FILES.get('file')
        if not f:
            return Response({'ok': False, 'error': '파일을 선택하세요'}, status=400)
        parsed = svc.parse_excel(f)
        if not parsed.get('ok'):
            return Response(parsed, status=400)
        codes = parsed['codes']
        if not codes:
            return Response({'ok': False, 'error': '엑셀에서 W코드를 찾지 못함'}, status=400)
        reason = (request.data.get('reason') if hasattr(request, 'data') else None) or None
        r = svc.bulk_upsert(codes, reason=reason, source='excel',
                            reasons=parsed.get('reasons'))
        return Response({**r, 'parsed': len(codes),
                         'with_reason': parsed.get('with_reason', 0)})


class WcodeBlacklistMatchView(APIView):
    """POST /api/smartstore/wcode-blacklist/match/
    3-DB 매칭 결과 (naver_my_product + ads.preliminary_product + ads.ownerclan_product).
    """
    def post(self, request):
        from . import wcode_blacklist_service as svc
        return Response(svc.run_match(refresh_flags=True))

    def get(self, request):
        from . import wcode_blacklist_service as svc
        return Response(svc.run_match(refresh_flags=False))


class WcodeBlacklistEnforceView(APIView):
    """POST /api/smartstore/wcode-blacklist/enforce/
    블랙리스트의 모든 W코드를 3개 타겟 테이블에서 일괄 DELETE.
    외부 sync(11st/ai100) 가 재등록한 행 정리용. is_processed 무관.
    """
    def post(self, request):
        from . import wcode_blacklist_service as svc
        return Response(svc.enforce_blacklist())


class WcodeBlacklistProcessView(APIView):
    """POST /api/smartstore/wcode-blacklist/process/
    body: { codes: [...], targets: ['my','pre','oc'] }
    체크된 W코드 → DB 행 삭제. 기본 'my' 만 안전. 마켓 API 호출은 하지 않음.
    """
    def post(self, request):
        from . import wcode_blacklist_service as svc
        d = request.data or {}
        raw_codes = d.get('codes') or []
        if isinstance(raw_codes, str):
            codes = svc._parse_codes(raw_codes)
        else:
            codes = []
            seen: set = set()
            for x in raw_codes:
                c = svc._normalize_code(str(x))
                if c and c not in seen:
                    seen.add(c)
                    codes.append(c)
        if not codes:
            return Response({'ok': False, 'error': 'codes required'}, status=400)
        targets = d.get('targets') or ['my']
        if not isinstance(targets, list):
            return Response({'ok': False, 'error': 'targets must be list'}, status=400)
        return Response(svc.process_delete(codes, targets))


class NaverMyProductNameVersionsView(APIView):
    """GET /api/smartstore/naver-products/<pk>/name-versions/
    버전 스냅샷 목록 + 현재 적용 버전.
    """
    def get(self, request, pk):
        from . import naver_name_confirmation_service as ncs
        r = ncs.list_version_snapshots(int(pk),
                                        limit=int(request.query_params.get('limit', 20)))
        # 현재 코드 버전(상수) + 패치 노트 같이 — 모달 헤더/안내 표시용
        from . import naver_name_generator as ng
        return Response({
            'ok': True,
            'code_version': ng.NAVER_NAME_VERSION,
            'patch_notes': ng.NAVER_NAME_PATCH_NOTES,
            **r,
        })


class NaverMyProductRollbackNameView(APIView):
    """POST /api/smartstore/naver-products/<pk>/rollback-name/
    body: { snapshot_id: int }
    """
    def post(self, request, pk):
        from . import naver_name_confirmation_service as ncs
        d = request.data or {}
        sid = d.get('snapshot_id')
        if sid is None:
            return Response({'ok': False, 'error': 'snapshot_id required'}, status=400)
        r = ncs.rollback_to_snapshot(int(pk), int(sid))
        return Response(r, status=200 if r.get('ok') else 400)


class NaverNamePolicyView(APIView):
    """GET  /api/smartstore/name-policy/?category_type=apparel&policy=white|black&limit=200
    카테고리별 누적 화/흑 키워드.
    """
    def get(self, request):
        from . import naver_name_confirmation_service as ncs
        q = request.query_params
        r = ncs.list_policy(
            category_type=q.get('category_type') or None,
            policy=q.get('policy') or None,
            limit=int(q.get('limit', 200)),
        )
        return Response({'ok': True, **r})


class NaverNamePolicyDeleteView(APIView):
    """DELETE /api/smartstore/name-policy/<category_type>/<keyword>/<policy>/"""
    def delete(self, request, category_type, keyword, policy):
        from . import naver_name_confirmation_service as ncs
        return Response(ncs.delete_policy(category_type, keyword, policy))


class BrandPolicyListView(APIView):
    """GET /api/smartstore/brand-policy/?policy=white|black&search=...&limit=200
        ?auto_discover=1 → 미등록 brand/manufacturer 빈도순
    POST body: {name, policy: white|black, note?}
    """
    def get(self, request):
        from . import brand_policy_service as bp
        q = request.query_params
        if q.get('auto_discover') == '1':
            items = bp.auto_discover(limit=int(q.get('limit', 50)))
            return Response({'items': items, 'mode': 'auto_discover'})
        r = bp.list_all(
            policy=q.get('policy') or None,
            search=q.get('search') or None,
            limit=int(q.get('limit', 200)),
            offset=int(q.get('offset', 0)),
        )
        return Response(r)

    def post(self, request):
        from . import brand_policy_service as bp
        d = request.data or {}
        r = bp.upsert(
            name=d.get('name'),
            policy=d.get('policy'),
            source=d.get('source', 'user'),
            note=d.get('note'),
        )
        return Response(r, status=200 if r.get('ok') else 400)


class BrandPolicyDeleteView(APIView):
    def delete(self, request, name):
        from . import brand_policy_service as bp
        return Response(bp.delete(name))


class NaverKeywordHotView(APIView):
    """GET /api/smartstore/naver-keyword-hot/?category=패션의류>여성의류
    카테고리 핫 키워드 (캐시 — 캐시 없으면 빈 배열)."""
    def get(self, request):
        cat = request.query_params.get('category') or ''
        from . import keyword_volume_service as kvs
        items = kvs.get_category_hot_keywords(cat, limit=int(request.query_params.get('limit', 20)))
        return Response({'ok': True, 'category': cat, 'items': items})


class NaverMyProductExcelView(APIView):
    """필터/정렬 그대로 엑셀 다운로드.
    Query params:
      folder_id, search, sort=sales|id_desc, top_sales=N (TOP N만), limit=N
    """
    def get(self, request):
        from io import BytesIO
        from urllib.parse import quote
        from datetime import datetime as _dt
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        q = request.query_params
        folder_id = q.get('folder_id')
        sort = q.get('sort') or 'id_desc'
        top_sales = q.get('top_sales')
        # 매출 모드: top_sales 있거나 sort=sales 면 자동 매출 정렬 + sales 조인
        if top_sales:
            sort = 'sales'
        include_sales = sort == 'sales' or q.get('include_sales') == '1'
        # limit: top_sales 우선, 없으면 q.get('limit') 또는 5000 (안전 상한)
        limit = int(top_sales) if top_sales else int(q.get('limit', 5000))
        limit = min(limit, 10000)

        data = naver_my_product_service.get_products(
            page=1, per_page=limit,
            folder_id=int(folder_id) if folder_id not in (None, '', 'all') else None,
            search=q.get('search') or None,
            sort=sort, include_sales=include_sales,
        )
        items = data['items']

        wb = Workbook()
        ws = wb.active
        ws.title = '네이버상품목록'

        # 헤더 정의 (매출 모드 여부에 따라)
        headers = ['순위', 'W코드']
        if include_sales:
            headers += ['총매출', '판매수량', '주문수']
        headers += [
            '원본상품명', '11번가 AI상품명', '편집명',
            '🌐 네이버상품명', '글자수', '바이트',
            '카테고리', '브랜드', '제조사', '원산지', '모델명',
            '판매가', '오너클랜가', '배송비', '반품비',
            '폴더', '이미지URL',
            '비전:형태', '비전:소재', '비전:색상', '비전:패키지',
            '비전:특징', '비전:글자',
            '비전분석일', '상품명생성일',
        ]
        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_fill = PatternFill('solid', fgColor='03c75a')
        center = Alignment(horizontal='center', vertical='center')

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center

        # 컬럼 너비 자동
        widths = {
            '순위': 6, 'W코드': 10,
            '총매출': 12, '판매수량': 9, '주문수': 8,
            '원본상품명': 35, '11번가 AI상품명': 35, '편집명': 30,
            '🌐 네이버상품명': 50, '글자수': 7, '바이트': 7,
            '카테고리': 30, '브랜드': 15, '제조사': 15, '원산지': 12, '모델명': 15,
            '판매가': 10, '오너클랜가': 10, '배송비': 8, '반품비': 8,
            '폴더': 12, '이미지URL': 40,
            '비전:형태': 12, '비전:소재': 12, '비전:색상': 20, '비전:패키지': 12,
            '비전:특징': 35, '비전:글자': 20,
            '비전분석일': 17, '상품명생성일': 17,
        }
        for col, h in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = widths.get(h, 12)

        import json as _json
        # naverdb 의 image_analysis JSON 가져오기 (get_products 응답엔 없음)
        from django.db import connections
        ids = [it['id'] for it in items]
        vision_map = {}
        if ids:
            ph = ','.join(['%s'] * len(ids))
            with connections['naverdb'].cursor() as cur:
                cur.execute(
                    f"SELECT id, image_analysis, image_analyzed_at FROM naver_my_product WHERE id IN ({ph})",
                    ids,
                )
                for row in cur.fetchall():
                    raw = row[1]
                    va = None
                    if raw:
                        try:
                            va = _json.loads(raw) if isinstance(raw, str) else raw
                        except Exception:
                            va = None
                    vision_map[row[0]] = (va, row[2])

        # 폴더 이름 매핑
        with connections['naverdb'].cursor() as cur:
            cur.execute("SELECT id, name FROM naver_my_product_folder")
            folder_name_map = {r[0]: r[1] for r in cur.fetchall()}

        def _calc_bytes(s):
            if not s:
                return 0
            return sum(2 if '가' <= c <= '힣' else 1 for c in s)

        def _color_str(v):
            c = v.get('color') if v else None
            if isinstance(c, list):
                return ', '.join(c)
            return c or ''

        def _list_str(v, key):
            x = v.get(key) if v else None
            if isinstance(x, list):
                return ', '.join(x)
            return x or ''

        for idx, it in enumerate(items, start=1):
            va, va_at = vision_map.get(it['id'], (None, None))
            name = it.get('naver_product_name') or ''
            sales = it.get('sales') or {}
            row = [idx, it.get('product_code', '')]
            if include_sales:
                row += [sales.get('total_amount', 0), sales.get('total_quantity', 0), sales.get('order_count', 0)]
            row += [
                it.get('product_name', ''),
                it.get('ai_recommended_name', '') or it.get('ai_product_name', ''),
                it.get('edited_product_name', '') or '',
                name,
                len(name) if name else 0,
                _calc_bytes(name),
                it.get('category_name', ''),
                it.get('brand', ''),
                it.get('manufacturer', ''),
                it.get('origin', ''),
                '',  # model_name (not in list response)
                it.get('market_price', 0),
                it.get('ownerclan_price', 0),
                it.get('shipping_fee', 0),
                it.get('return_fee', 0),
                folder_name_map.get(it.get('folder_id'), ''),
                it.get('image_large', '') or it.get('image_small', ''),
                (va or {}).get('form', '') or '',
                (va or {}).get('material', '') or '',
                _color_str(va),
                (va or {}).get('package_qty', '') or '',
                _list_str(va, 'key_features'),
                (va or {}).get('readable_text', '') or '',
                va_at.strftime('%Y-%m-%d %H:%M') if va_at else '',
                (it.get('synced_at') or '')[:16].replace('T', ' ') if it.get('synced_at') else '',
            ]
            for col, v in enumerate(row, 1):
                ws.cell(row=idx + 1, column=col, value=v)

        ws.freeze_panes = 'C2'

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        ts = _dt.now().strftime('%Y%m%d_%H%M%S')
        suffix = f'_TOP{top_sales}' if top_sales else (f'_매출정렬' if sort == 'sales' else '')
        filename = f'네이버상품목록{suffix}_{ts}.xlsx'

        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return resp


class NaverMyProductMoveView(APIView):
    """선택된 상품들을 지정 폴더로 이동.
    body: {"ids": [1,2,3], "folder_id": 5}
    """
    def post(self, request):
        d = request.data or {}
        ids = d.get('ids') or []
        folder_id = d.get('folder_id')
        if not ids or folder_id is None:
            return Response({'ok': False, 'error': 'ids와 folder_id 필수'}, status=400)
        result = naver_my_product_service.move_products(
            ids=[int(x) for x in ids], folder_id=int(folder_id))
        status = 200 if result.get('ok') else 400
        return Response(result, status=status)


# ─────────────────────────────────────────────────────────────────────
# 상품 일괄등록 (엑셀 일괄등록 양식)
# ─────────────────────────────────────────────────────────────────────

class RegisterSetListView(APIView):
    """등록 세트 목록 + 폴더(스토어) 목록 + 확정상품 수."""
    def get(self, request):
        folders = naver_my_product_service.list_folders()
        sets = {s['folder_id']: s for s in register_set_service.list_sets()}
        out = []
        for f in folders:
            if f.get('is_system'):
                continue
            out.append({
                'folder_id': f['id'],
                'store_id': f.get('store_id'),
                'name': f.get('name'),
                'product_count': f.get('product_count', 0),
                'has_set': f['id'] in sets,
                'set': sets.get(f['id']),
            })
        return Response({'folders': out})


class RegisterSetView(APIView):
    """폴더별 등록 세트 조회/저장."""
    def get(self, request, folder_id):
        store_name = request.query_params.get('store_name')
        s = register_set_service.get_or_create_set(int(folder_id), store_name)
        return Response({'set': s})

    def put(self, request, folder_id):
        s = register_set_service.upsert_set(int(folder_id), request.data or {})
        return Response({'ok': True, 'set': s})


class RegisterSetCalcView(APIView):
    """라이브 마진 계산기 — 세트 파라미터 + 원가/원본배송비 → 가격 산출.
    body: { set: {...}, cost?: 10000, orig_ship_fee?: 0 }
    """
    def post(self, request):
        body = request.data or {}
        set_row = body.get('set') or body
        cost = int(body.get('cost') or 10000)
        orig = int(body.get('orig_ship_fee') or 0)
        return Response(register_set_service.compute_price(set_row, cost, orig))


class BulkRegisterPreviewView(APIView):
    """일괄등록 미리보기 — 확정상품 상위 N건 판매가 계산 + 경고/파일수.
    body: { folder_id, set: {...}, n?: 10 }
    """
    def post(self, request):
        body = request.data or {}
        folder_id = body.get('folder_id')
        if not folder_id:
            return Response({'ok': False, 'error': 'folder_id 필수'}, status=400)
        set_row = body.get('set') or {}
        n = int(body.get('n') or 10)
        return Response(bulk_register_service.preview(int(folder_id), set_row, n=n))


class BulkRegisterGenerateView(APIView):
    """일괄등록 엑셀 생성 — 확정상품 전체 → 500개 단위 xlsx → ZIP 다운로드.
    body: { folder_id, store_name? }  (세트는 저장된 것 사용)
    """
    def post(self, request):
        body = request.data or {}
        folder_id = body.get('folder_id')
        if not folder_id:
            return Response({'ok': False, 'error': 'folder_id 필수'}, status=400)
        folder_id = int(folder_id)
        store_name = body.get('store_name') or ''
        set_row = register_set_service.get_set(folder_id)
        if not set_row:
            return Response({'ok': False, 'error': '등록 세트가 없습니다. 먼저 세트를 저장하세요.'}, status=400)
        zip_bytes, meta = bulk_register_service.generate_zip(folder_id, set_row, store_name)
        if not zip_bytes:
            return Response({'ok': False, 'error': '확정상품이 없습니다.'}, status=400)
        from datetime import datetime as _dt
        safe = quote(f"{store_name or ('folder'+str(folder_id))}_일괄등록_{_dt.now():%Y%m%d}.zip")
        resp = HttpResponse(zip_bytes, content_type='application/zip')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{safe}"
        resp['X-Meta-Files'] = str(meta.get('files', 0))
        resp['X-Meta-Total'] = str(meta.get('total', 0))
        return resp


# ─────────────────────────────────────────────────────────────────────
# 일괄등록 단계 폴더 (등록후보/작업대기) + 검증 + 작업대기 엑셀
# ─────────────────────────────────────────────────────────────────────

class RegisterFolderTreeView(APIView):
    """스토어별 단계 카운트 (전체/등록후보/작업대기/등록완료)."""
    def get(self, request):
        return Response({'stores': register_folder_service.list_store_stages()})


class RegisterStageView(APIView):
    """체크된 상품의 단계 설정 — stage: 'candidate'/'queue'/null(되돌리기)."""
    def post(self, request):
        body = request.data or {}
        ids = body.get('ids') or []
        stage = body.get('stage')   # 'candidate' | 'queue' | None
        if not ids:
            return Response({'ok': False, 'error': 'ids 필수'}, status=400)
        r = register_folder_service.set_stage([int(x) for x in ids], stage)
        return Response(r, status=200 if r.get('ok') else 400)


class RegisterSelectIdsView(APIView):
    """필터/AI추천에 맞는 상품 id 전체 반환 (전체선택 / AI추천 체크).
    params: folder_id, stage, registered, category_code, search, recommend(0/1), limit
    """
    def get(self, request):
        q = request.query_params
        fid = q.get('folder_id')
        if not fid:
            return Response({'ok': False, 'error': 'folder_id 필수'}, status=400)
        reg = q.get('registered')
        r = register_folder_service.select_ids(
            folder_id=int(fid),
            stage=q.get('stage') or None,
            registered=int(reg) if reg in ('0', '1') else None,
            category_code=q.get('category_code') or None,
            search=q.get('search') or None,
            recommend=q.get('recommend') == '1',
            limit=int(q.get('limit') or 2000),
        )
        return Response(r)


class RegisterMarkView(APIView):
    """상품등록완료 표기/해제."""
    def post(self, request):
        body = request.data or {}
        ids = body.get('ids') or []
        value = 1 if body.get('registered', True) else 0
        r = register_folder_service.mark_registered([int(x) for x in ids], value)
        return Response(r, status=200 if r.get('ok') else 400)


class RegisterVerifyView(APIView):
    """작업대기 상품의 실제 네이버 등록 여부 검증(동기화).
    body: { store_folder_id, sync?: true }
    """
    def post(self, request):
        body = request.data or {}
        sid = body.get('store_folder_id')
        if not sid:
            return Response({'ok': False, 'error': 'store_folder_id 필수'}, status=400)
        r = register_folder_service.verify_registration(int(sid), run_sync=bool(body.get('sync')))
        return Response(r, status=200 if r.get('ok') else 400)


class RegisterSyncInspectView(APIView):
    """전체 올린상품 동기화 + 특이사항 점검 + 로그.
    body: { store_folder_id, sync?: true }
    """
    def post(self, request):
        body = request.data or {}
        sid = body.get('store_folder_id')
        if not sid:
            return Response({'ok': False, 'error': 'store_folder_id 필수'}, status=400)
        r = register_folder_service.inspect_registration(int(sid), run_sync=bool(body.get('sync', True)))
        return Response(r, status=200 if r.get('ok') else 400)


class RegisterQueueExcelAllView(APIView):
    """전체 스토어 작업대기 → 스토어별 폴더로 한 ZIP. (세트 없으면 자동 복제)"""
    def post(self, request):
        stores = []
        for s in register_folder_service.list_store_stages():
            if s.get('store_id') and s.get('staged_pending', 0) > 0:
                set_row = register_set_service.ensure_set(s['id'], s['name'])
                stores.append({'folder_id': s['id'], 'name': s['name'], 'set_row': set_row})
        if not stores:
            return Response({'ok': False, 'error': '작업대기 상품이 있는 스토어가 없습니다.'}, status=400)
        zip_bytes, meta = bulk_register_service.generate_all_queue_zip(stores)
        if not zip_bytes:
            return Response({'ok': False, 'error': '생성된 상품이 없습니다.'}, status=400)
        from datetime import datetime as _dt
        safe = quote(f"전체스토어_작업대기_일괄등록_{_dt.now():%Y%m%d}.zip")
        resp = HttpResponse(zip_bytes, content_type='application/zip')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{safe}"
        resp['X-Meta-Stores'] = str(meta.get('stores', 0))
        resp['X-Meta-Files'] = str(meta.get('files', 0))
        resp['X-Meta-Total'] = str(meta.get('total', 0))
        return resp


class RegisterFailureAnalyzeView(APIView):
    """네이버 실패 엑셀 업로드 → 사유 분석 + 성공분 등록완료 표기.
    multipart: file=실패엑셀, store_folder_id, mark_success(기본 1)
    """
    def post(self, request):
        f = request.FILES.get('file')
        sid = request.data.get('store_folder_id')
        if not f or not sid:
            return Response({'ok': False, 'error': 'file, store_folder_id 필수'}, status=400)
        parsed = register_folder_service.parse_error_excel(f.read())
        mark = str(request.data.get('mark_success', '1')) in ('1', 'true', 'True')
        applied = register_folder_service.apply_failure_result(
            int(sid), parsed['failed_codes'], mark_success=mark)
        return Response({'ok': True, **parsed, 'applied': applied})


class RegisterQueueExcelView(APIView):
    """작업대기 폴더 전체 → 500개 단위 ZIP 다운로드.
    body: { store_folder_id }  (세트는 스토어 세트 사용)
    """
    def post(self, request):
        body = request.data or {}
        sid = body.get('store_folder_id')
        if not sid:
            return Response({'ok': False, 'error': 'store_folder_id 필수'}, status=400)
        sid = int(sid)
        # 스토어명
        from django.db import connections as _conns
        with _conns['naverdb'].cursor() as cur:
            cur.execute("SELECT name FROM naver_my_product_folder WHERE id=%s", [sid])
            row = cur.fetchone()
        store_name = (row[0] if row else '') or f'folder{sid}'
        # 세트 없으면 행원만물상 기준으로 자동 복제
        set_row = register_set_service.ensure_set(sid, store_name)
        if not set_row:
            return Response({'ok': False, 'error': '스토어 등록 세트 생성 실패.'}, status=400)
        zip_bytes, meta = bulk_register_service.generate_zip_from_stage(sid, 'queue', set_row, store_name)
        if not zip_bytes:
            return Response({'ok': False, 'error': '작업대기 상품이 없습니다.'}, status=400)
        from datetime import datetime as _dt
        safe = quote(f"{store_name}_작업대기_일괄등록_{_dt.now():%Y%m%d}.zip")
        resp = HttpResponse(zip_bytes, content_type='application/zip')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{safe}"
        resp['X-Meta-Files'] = str(meta.get('files', 0))
        resp['X-Meta-Total'] = str(meta.get('total', 0))
        return resp


# ─────────────────────────────────────────────────────────────────────
# 라이브 상품명 금지어 점검/수정
# ─────────────────────────────────────────────────────────────────────

class LiveNameScanView(APIView):
    """라이브 상품명 금지어(브랜드 black) 스캔. ?store_id= (선택)"""
    def get(self, request):
        sid = request.query_params.get('store_id')
        r = live_name_service.scan(store_id=int(sid) if sid else None,
                                   limit=int(request.query_params.get('limit', 1000)))
        return Response(r)


class LiveNameFixView(APIView):
    """라이브 상품명 수정 (네이버 API). body: { items:[{store_id, origin_product_no, new_name}] }"""
    def post(self, request):
        items = (request.data or {}).get('items') or []
        if not items:
            return Response({'ok': False, 'error': 'items 필수'}, status=400)
        return Response(live_name_service.fix_names(items))


# ─────────────────────────────────────────────────────────────────────
# 불량 상품명 탐지 (AI 안내문 누출 등)
# ─────────────────────────────────────────────────────────────────────

class BadNameScanView(APIView):
    """불량 상품명 탐지. ?scope=pool|live|both (기본 both), ?stages=queue,candidate"""
    def get(self, request):
        scope = request.query_params.get('scope', 'both')
        stages_q = request.query_params.get('stages', 'queue,candidate')
        stages = tuple(s for s in stages_q.split(',') if s) if stages_q and stages_q != 'all' else None
        out = {}
        if scope in ('pool', 'both'):
            out['pool'] = bad_name_service.scan_pool(stages=stages)
        if scope in ('live', 'both'):
            sid = request.query_params.get('store_id')
            out['live'] = bad_name_service.scan_live(store_id=int(sid) if sid else None)
        return Response(out)


class BadNamePurgeView(APIView):
    """불량명 상품을 작업대기/후보에서 제외(stage=NULL). body: { ids:[...] }"""
    def post(self, request):
        ids = (request.data or {}).get('ids') or []
        r = bad_name_service.purge_from_queue([int(x) for x in ids])
        return Response(r, status=200 if r.get('ok') else 400)


# ─────────────────────────────────────────────────────────────────────
# 중복 상품 초과분 품절처리
# ─────────────────────────────────────────────────────────────────────

class DupSuspendScanView(APIView):
    """중복 W코드 그룹 + 품절처리 대상(초과분) 스캔. ?store_id="""
    def get(self, request):
        sid = request.query_params.get('store_id')
        return Response(dup_suspend_service.scan_excess(store_id=int(sid) if sid else None))


class DupSuspendApplyView(APIView):
    """중복 초과분 품절처리(SUSPENSION). body: { store_id? , items?:[{store_id,origin_product_no}] }"""
    def post(self, request):
        body = request.data or {}
        sid = body.get('store_id')
        r = dup_suspend_service.suspend_excess(
            store_id=int(sid) if sid else None, items=body.get('items'))
        return Response(r, status=200 if r.get('ok') else 400)


# ─────────────────────────────────────────────────────────────────────
# 상품등록정보검토 (진단) 수집
# ─────────────────────────────────────────────────────────────────────

class DiagnosisSyncView(APIView):
    """진단 수집 — 아이디별 워커 병렬 실행. body: { login_ids?:[], concurrency?:5 }"""
    def post(self, request):
        body = request.data or {}
        r = diagnosis_service.dispatch_parallel(
            login_ids=body.get('login_ids'), concurrency=int(body.get('concurrency', 5)))
        return Response(r, status=200 if r.get('ok') else 400)


class DiagnosisStatusView(APIView):
    def get(self, request):
        return Response(diagnosis_service.sync_status())


class DiagnosisResultsView(APIView):
    """스토어별 진단 결과. ?store_id="""
    def get(self, request):
        sid = request.query_params.get('store_id')
        if not sid:
            return Response({'ok': False, 'error': 'store_id 필수'}, status=400)
        return Response(diagnosis_service.get_results(int(sid)))


class ProductTagAttrMetaView(APIView):
    """W코드 목록 → 태그/등록속성 메타. ?codes=W1,W2 (&store_id=)"""
    def get(self, request):
        codes = [c for c in (request.query_params.get('codes', '').split(',')) if c]
        if not codes:
            return Response({'meta': {}})
        sid = request.query_params.get('store_id')
        from django.db import connections as _c
        out = {c: {'tags': [], 'tag_count': 0, 'attr_count': 0} for c in codes}
        ph = ','.join(['%s'] * len(codes))
        # 태그
        import json as _j
        with _c['naverdb'].cursor() as cur:
            cur.execute(f"SELECT product_code, tags_json, registered FROM naver_product_tags WHERE product_code IN ({ph})", codes)
            for code, tj, reg in cur.fetchall():
                try:
                    tags = _j.loads(tj) if tj else []
                except Exception:
                    tags = []
                out[code]['tags'] = tags
                out[code]['tag_count'] = len(tags)
                out[code]['tag_registered'] = int(reg or 0)
        # 등록 속성 수
        with _c['myproduct'].cursor() as cur:
            q = f"SELECT seller_management_code, COUNT(*) FROM smartstore_product_missing_attrs WHERE status='registered' AND seller_management_code IN ({ph})"
            params = list(codes)
            if sid:
                q += " AND store_id=%s"; params.append(int(sid))
            q += " GROUP BY seller_management_code"
            cur.execute(q, params)
            for code, n in cur.fetchall():
                if code in out:
                    out[code]['attr_count'] = int(n)
        return Response({'meta': out})


class AttrAutoCheckPreviewView(APIView):
    """POST {store_id?, codes?, limit?, mode?} — AI 자동체크 미리보기(dry-run, 등록 안 함)."""
    def post(self, request):
        from . import auto_attr_service as aas
        d = request.data or {}
        codes = d.get('codes') or None
        return Response(aas.preview(
            store_id=d.get('store_id'), limit=min(int(d.get('limit', 60)), 300),
            codes=codes, mode=d.get('mode', 'auto')))


class AttrAutoCheckStartView(APIView):
    """POST {store_id?, limit?, mode?} — 백그라운드 자동등록 배치 실행."""
    def post(self, request):
        from . import auto_attr_service as aas
        d = request.data or {}
        return Response(aas.start(
            store_id=d.get('store_id'), limit=min(int(d.get('limit', 500)), 5000),
            mode=d.get('mode', 'auto')))


class AttrAutoCheckStatusView(APIView):
    """GET — 자동체크 워커 상태 + pending/registered 카운트."""
    def get(self, request):
        from . import auto_attr_service as aas
        return Response(aas.status())


class AttrVisionStartView(APIView):
    """POST {store_id?, limit?} — 썸네일 라이브 비전 배치(GPU 분산) 실행."""
    def post(self, request):
        from . import auto_attr_service as aas
        d = request.data or {}
        return Response(aas.start_vision(
            store_id=d.get('store_id'), limit=min(int(d.get('limit', 500)), 5000)))


class AttrVisionStatusView(APIView):
    """GET — 썸네일 비전 배치 상태 + 남은 대상 수."""
    def get(self, request):
        from . import auto_attr_service as aas
        return Response(aas.vision_status())


class AttrPipelineStatusView(APIView):
    """GET ?store_id= — 스토어별 3단계 파이프라인 카운트 + 수집 커버리지."""
    def get(self, request):
        from . import attr_sync_service as ass
        sid = request.query_params.get('store_id')
        return Response(ass.pipeline_status(store_id=int(sid) if sid else None))


class AttrSyncStartView(APIView):
    """POST {store_id?, limit?} — 스테이징(classified)된 추론값 네이버 동기화(PUT) 배치 실행."""
    def post(self, request):
        from . import attr_sync_service as ass
        d = request.data or {}
        return Response(ass.start_sync(
            store_id=d.get('store_id'), limit=min(int(d.get('limit', 5000)), 50000)))


class AttrSyncStatusView(APIView):
    """GET — 동기화 워커 상태 + classified/registered 카운트."""
    def get(self, request):
        from . import attr_sync_service as ass
        return Response(ass.sync_status())


class AttrProductView(APIView):
    """GET ?seller_code=(&store_id=) — 단일 상품 속성 추론/등록 현황(개별 표시용).
    store_id 미지정 시 W코드 보유 스토어로 자동 해석(My상품용)."""
    def get(self, request):
        from . import attr_sync_service as ass
        seller = request.query_params.get('seller_code')
        sid = request.query_params.get('store_id')
        if not seller:
            return Response({'ok': False, 'error': 'seller_code 필수'}, status=400)
        return Response(ass.product_attrs(seller, int(sid) if sid else None))


class AttrProductContextView(APIView):
    """GET ?seller_code=(&store_id=) — 단일 상품 실물 컨텍스트(상품명/카테고리/대표이미지/상세).
    커머스 API GET 1회. 모달 오픈/SKU 선택 시에만 호출(리스트 호출 금지)."""
    def get(self, request):
        from . import attr_sync_service as ass
        seller = request.query_params.get('seller_code')
        sid = request.query_params.get('store_id')
        if not seller:
            return Response({'ok': False, 'error': 'seller_code 필수'}, status=400)
        return Response(ass.product_context(seller, int(sid) if sid else None))


class AttrRecentChangesView(APIView):
    """GET ?store_id=&limit= — 속성 적용(PUT) 최근 변경 이력."""
    def get(self, request):
        from . import attr_sync_service as ass
        sid = request.query_params.get('store_id')
        limit = min(int(request.query_params.get('limit', 100)), 500)
        return Response(ass.recent_changes(store_id=int(sid) if sid else None, limit=limit))


class AttrListByStatusView(APIView):
    """GET ?status=&store_id=&limit=&search= — 특정 status SKU 목록(드릴다운)."""
    def get(self, request):
        from . import attr_sync_service as ass
        status = request.query_params.get('status', '')
        sid = request.query_params.get('store_id')
        limit = min(int(request.query_params.get('limit', 200)), 1000)
        search = request.query_params.get('search', '')
        return Response(ass.list_by_status(status, store_id=int(sid) if sid else None,
                                           limit=limit, search=search))


class AttrNeedsManualSkuView(APIView):
    """GET ?seller_code=&store_id= — 단일 SKU 의 needs_manual 속성+후보값."""
    def get(self, request):
        from . import attr_sync_service as ass
        seller = request.query_params.get('seller_code')
        sid = request.query_params.get('store_id')
        if not seller:
            return Response({'ok': False, 'error': 'seller_code 필수'}, status=400)
        return Response(ass.needs_manual_for_sku(seller, int(sid) if sid else None))


class AttrPipelineCrawlStartView(APIView):
    """POST {store_id?} — ① 수집(크롤)+탐지 실행."""
    def post(self, request):
        from . import attr_sync_service as ass
        return Response(ass.start_crawl(store_id=(request.data or {}).get('store_id')))


class AttrPipelineStageStartView(APIView):
    """POST {store_id?, limit?} — ② GPU 추론(스테이징) 실행."""
    def post(self, request):
        from . import attr_sync_service as ass
        d = request.data or {}
        return Response(ass.start_stage(store_id=d.get('store_id'), limit=min(int(d.get('limit', 5000)), 50000)))


class AttrPipelineWorkersView(APIView):
    """GET — 3단계 워커(수집/추론/동기화) 진행상태."""
    def get(self, request):
        from . import attr_sync_service as ass
        return Response(ass.workers())
