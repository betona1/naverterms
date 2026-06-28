"""네이버 상품 일괄등록 — 엑셀(일괄등록 양식) 생성.

확정상품(naver_name_confirmation 이력 + naver_product_name) → 세트로 판매가 계산
→ 양식 93컬럼 매핑 → 500개 단위 xlsx → ZIP.

업로드: https://sell.smartstore.naver.com/#/products/bulkadd
"""
from __future__ import annotations

import io
import os
import re
import zipfile

from django.conf import settings
from django.db import connections

from . import register_set_service as rss

NAVERDB = 'naverdb'
HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(HERE, '..', '..', 'docs', 'ExcelSaveTemplate_20260324.xlsx')

BATCH_SIZE = 500
HEADER_ROWS = 2       # 1행 섹션 + 2행 컬럼명 유지
GUIDE_ROWS = (3, 6)   # 3~6행(필수표기/예시/가이드/주석) 삭제
DATA_START = 3        # 데이터 시작 행

# 공인 미디어 베이스 (보정/업스케일 이미지를 쓰려면 .env 에 PUBLIC_MEDIA_BASE_URL 설정)
PUBLIC_MEDIA_BASE = getattr(settings, 'PUBLIC_MEDIA_BASE_URL', '') or os.getenv('PUBLIC_MEDIA_BASE_URL', '')

_NAME_BAN = re.compile(r'[\\\*\?"<>]')
_ws = re.compile(r'\s+')


def _norm(s) -> str:
    return _ws.sub('', str(s or ''))


def _sanitize_name(name: str) -> str:
    return _NAME_BAN.sub('', (name or '').strip())[:100]


def _pick_image(row: dict) -> str:
    """대표이미지 URL.

    업스케일 완료(upscaled_image_url 존재) 상품은 joacham 호스팅(W코드_1.jpg, 공인)을
    사용하고, 그 외에는 오너클랜 공인 CDN(image_large)을 사용한다.
    """
    if PUBLIC_MEDIA_BASE and row.get('upscaled_image_url'):
        code = (row.get('product_code') or '').strip()
        if code:
            return f"{PUBLIC_MEDIA_BASE.rstrip('/')}/{code}_1.jpg"
    return row.get('image_large') or row.get('image_medium') or row.get('image_small') or ''


def _map_origin(origin_text: str | None, set_origin_code: str) -> str:
    t = origin_text or ''
    if '국산' in t or '국내' in t:
        return '00'
    return set_origin_code or '03'


def _fix_detail_html(html: str) -> str:
    """상세설명 이미지 URL 보정 — 프로토콜 없는 //url → https:// (네이버 거부 회피)."""
    if not html:
        return html
    # src="//..." / src='//...' → https://
    return re.sub(r'(src\s*=\s*["\'])\s*//', r'\1https://', html, flags=re.I)


def _opt_values_str(v: str) -> str:
    """옵션값 콤마 구분, 각 값 25자 이내로 절단 (네이버 25자 제한)."""
    parts = [x.strip()[:25] for x in (v or '').split(',') if x.strip()]
    return ','.join(parts)


def _options(row: dict, default_stock: int) -> dict:
    """옵션 매핑 → {옵션형태, 옵션명, 옵션값, 옵션재고수량}."""
    n1, v1 = (row.get('option1_name') or '').strip(), _opt_values_str(row.get('option1_values'))
    n2, v2 = (row.get('option2_name') or '').strip(), _opt_values_str(row.get('option2_values'))
    names, vals = [], []
    if n1 and v1:
        names.append(n1[:25]); vals.append(v1)
    if n2 and v2:
        names.append(n2[:25]); vals.append(v2)
    if not names:
        return {}
    opt_type = '조합형' if len(names) >= 2 else '단독형'
    first_vals = [x for x in vals[0].split(',') if x.strip()]
    stock = ','.join([str(default_stock)] * len(first_vals)) if first_vals else str(default_stock)
    return {
        '옵션형태': opt_type,
        '옵션명': '\n'.join(names),
        '옵션값': '\n'.join(vals),
        '옵션재고수량': stock,
    }


PRODUCT_FIELDS = (
    'p.id, p.product_code, p.naver_product_name, p.category_code, '
    'p.ownerclan_price, p.shipping_fee, p.return_fee, '
    'p.brand, p.manufacturer, p.model_name, p.origin, '
    'p.image_large, p.image_medium, p.image_small, p.edited_image_url, p.upscaled_image_url, '
    'p.option1_name, p.option1_values, p.option2_name, p.option2_values, p.detail_html'
)


def fetch_confirmed_products(folder_id: int, limit: int | None = None) -> list[dict]:
    """상품명 확정(컨펌 이력 보유 + naver_product_name) 상품."""
    sql = f"""
        SELECT {PRODUCT_FIELDS}
        FROM naver_my_product p
        WHERE p.folder_id=%s
          AND p.naver_product_name IS NOT NULL AND p.naver_product_name<>''
          AND EXISTS (SELECT 1 FROM naver_name_confirmation c WHERE c.product_code=p.product_code)
        ORDER BY p.id
    """
    if limit:
        sql += f" LIMIT {int(limit)}"
    with connections[NAVERDB].cursor() as cur:
        cur.execute(sql, [int(folder_id)])
        cols = [c[0] for c in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


ADS_DB = 'ads'


def fetch_additional_images(wcodes: list[str]) -> dict:
    """이셀러스/플레이오토(ads.product_esproduct) 추가이미지 조회.

    ba_seller_code(=W코드) 기준 ba_image1~5 중 채워진 것만 엔터(\\n)로 연결.
    테이블이 비어있거나 매칭이 없으면 빈 dict.
    """
    if not wcodes:
        return {}
    out: dict[str, str] = {}
    ph = ','.join(['%s'] * len(wcodes))
    try:
        with connections[ADS_DB].cursor() as cur:
            cur.execute(
                f"""SELECT ba_seller_code, ba_image1, ba_image2, ba_image3, ba_image4, ba_image5
                    FROM product_esproduct WHERE ba_seller_code IN ({ph})""",
                wcodes)
            for r in cur.fetchall():
                imgs = [str(x).strip() for x in r[1:6] if x and str(x).strip()]
                if imgs:
                    out[r[0]] = '\n'.join(imgs[:9])   # 최대 9개
    except Exception:
        return {}
    return out


_MINOR_BLOCK_CACHE: set | None = None


def _minor_blocked_categories() -> set:
    """미성년자 구매(Y) 등록 불가 카테고리 (naver_register_minor_block). 1회 캐시."""
    global _MINOR_BLOCK_CACHE
    if _MINOR_BLOCK_CACHE is None:
        try:
            with connections[NAVERDB].cursor() as cur:
                cur.execute(
                    "CREATE TABLE IF NOT EXISTS naver_register_minor_block "
                    "(category_code VARCHAR(20) NOT NULL PRIMARY KEY, "
                    " note VARCHAR(200) NULL, created_at DATETIME DEFAULT CURRENT_TIMESTAMP) "
                    "ENGINE=InnoDB DEFAULT CHARSET=utf8mb4")
                cur.execute("SELECT category_code FROM naver_register_minor_block")
                _MINOR_BLOCK_CACHE = {r[0] for r in cur.fetchall()}
        except Exception:
            _MINOR_BLOCK_CACHE = set()
    return _MINOR_BLOCK_CACHE


def fetch_folder_products(folder_id: int) -> list[dict]:
    """폴더 내 전체 상품 (확정 여부 무관)."""
    sql = f"""
        SELECT {PRODUCT_FIELDS}
        FROM naver_my_product p
        WHERE p.folder_id=%s
          AND p.naver_product_name IS NOT NULL AND p.naver_product_name<>''
        ORDER BY p.id
    """
    with connections[NAVERDB].cursor() as cur:
        cur.execute(sql, [int(folder_id)])
        cols = [c[0] for c in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def fetch_stage_products(store_folder_id: int, stage: str,
                         exclude_registered: bool = True) -> list[dict]:
    """스토어 폴더 내 특정 단계(register_stage) 상품 — 작업대기 엑셀용.

    exclude_registered=True: 이미 등록완료(registered=1)건 제외 (중복등록 방지).
    """
    extra = 'AND p.registered=0' if exclude_registered else ''
    sql = f"""
        SELECT {PRODUCT_FIELDS}
        FROM naver_my_product p
        WHERE p.folder_id=%s AND p.register_stage=%s {extra}
          AND p.naver_product_name IS NOT NULL AND p.naver_product_name<>''
        ORDER BY p.id
    """
    with connections[NAVERDB].cursor() as cur:
        cur.execute(sql, [int(store_folder_id), stage])
        cols = [c[0] for c in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def build_row(product: dict, set_row: dict, add_images: str = '') -> dict:
    """상품 1건 → {컬럼명: 값} (정규화된 키 기반은 writer 가 처리).

    배송비는 오너클랜 배송비(product.shipping_fee)에 맞춘다:
      · 무료배송(free_shipping=1): 판매가가 오너클랜 배송비를 흡수
      · 유료배송(free_shipping=0): 기본배송비 = 오너클랜 배송비, 판매가엔 배송비 미포함(조정 0)
    반품/교환배송비도 오너클랜 반품비(return_fee) 기준.
    """
    B = int(product.get('shipping_fee') or 0)
    # 오너클랜 배송비 0원이면 무료배송으로 처리 (유료+0원은 네이버 거부)
    free = int(set_row.get('free_shipping') or 0) == 1 or B <= 0
    # 유료배송이면 오너클랜 배송비와 동일하게 부과 → 판매가 배송비조정 0
    ps = dict(set_row)
    if not free:
        ps['set_ship_fee'] = B
    price = rss.compute_price(ps, product.get('ownerclan_price'), B)

    detail = _fix_detail_html(product.get('detail_html') or '')
    if not detail.strip():
        img = _pick_image(product)
        detail = (f'<div><h3>{_sanitize_name(product.get("naver_product_name"))}</h3>'
                  f'<img src="{img}"></div>')

    ret_fee = int(product.get('return_fee') or set_row.get('return_fee') or 0)
    row = {
        '판매자 상품코드': (product.get('product_code') or '')[:30],
        '카테고리코드': product.get('category_code') or '',
        '상품명': _sanitize_name(product.get('naver_product_name')),
        '상품상태': set_row.get('product_state') or '신상품',
        '판매가': price['list_price'],
        '재고수량': int(set_row.get('default_stock') or 999),
        '대표이미지': _pick_image(product),
        '상세설명': detail,
        # 브랜드: 비워둠 (정책) — 제조사만 기입
        '제조사': (product.get('manufacturer') or '')[:100],
        '원산지코드': _map_origin(product.get('origin'), set_row.get('origin_code') or '03'),
        '복수원산지여부': 'N',
        '미성년자 구매': 'N' if (product.get('category_code') in _minor_blocked_categories()) else 'Y',
        '부가세': set_row.get('vat_type') or '과세상품',
        # 단위가격: 가격표시제 의무 카테고리 대비 — 표시대상 아님(N)으로 명시 (미선택 오류 방지)
        '단위가격 사용여부': 'N',
        # 배송 (템플릿 미사용 — 오너클랜 배송비와 일치)
        '배송방법': '택배, 소포, 등기',
        '택배사코드': set_row.get('delivery_company_code') or 'CJGLS',
        '배송비유형': '무료' if free else '유료',
        '반품배송비': ret_fee,
        '교환배송비': ret_fee * 2 if ret_fee else int(set_row.get('exchange_fee') or 0),
    }
    # 리뷰 포인트 (판매가 산정에 반영됨) — 0이면 셀 비움(네이버는 0 불가, 10원 이상만 허용)
    rp_text = int(set_row.get('review_point_text') or 0)
    rp_photo = int(set_row.get('review_point_photo') or 0)
    if rp_text >= 10:
        row['텍스트리뷰 작성시\n지급 포인트'] = rp_text
    if rp_photo >= 10:
        row['포토/동영상 리뷰 작성시\n지급 포인트'] = rp_photo
    # 추가이미지 (이셀러스/플레이오토 ba_image1~5, 엔터 구분)
    if add_images:
        row['추가이미지'] = add_images
    # 유료배송 — 오너클랜 배송비 부과
    if not free and B > 0:
        row['기본배송비'] = B
        row['배송비 결제방식'] = '선결제'
    # 즉시할인 (정액)
    if price['discount_amount'] > 0:
        row['즉시할인 값\n(기본할인)'] = price['discount_amount']
        row['즉시할인 단위\n(기본할인)'] = '원'
    # A/S (템플릿 미사용 → 전화/안내 직접)
    if set_row.get('as_phone'):
        row['A/S 전화번호'] = set_row['as_phone']
    if set_row.get('as_guide'):
        row['A/S 안내'] = set_row['as_guide']
    # 옵션
    row.update(_options(product, int(set_row.get('default_stock') or 999)))
    return row


def _write_batch_xlsx(rows: list[dict]) -> bytes:
    """양식 헤더(1~2행) 유지 + 가이드행 삭제 + 데이터 기입 → xlsx bytes."""
    import openpyxl
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    # 헤더 컬럼명(2행) → 정규화 인덱스
    name_to_col = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROWS, column=col).value
        if v is not None:
            name_to_col[_norm(v)] = col
    # 가이드행(3~6) 삭제
    ws.delete_rows(GUIDE_ROWS[0], GUIDE_ROWS[1] - GUIDE_ROWS[0] + 1)
    # 데이터 기입
    for i, rowdata in enumerate(rows):
        r = DATA_START + i
        for key, val in rowdata.items():
            col = name_to_col.get(_norm(key))
            if col:
                ws.cell(row=r, column=col, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def preview(folder_id: int, set_row: dict, n: int = 10) -> dict:
    """상위 n건 판매가 계산 미리보기 + 경고."""
    products = fetch_confirmed_products(folder_id, limit=n)
    addimg = fetch_additional_images([p['product_code'] for p in products if p.get('product_code')])
    free = int(set_row.get('free_shipping') or 0) == 1
    items = []
    for p in products:
        B = int(p.get('shipping_fee') or 0)
        ps = dict(set_row)
        if not free:
            ps['set_ship_fee'] = B   # 유료: 오너클랜 배송비와 일치 → 조정 0
        price = rss.compute_price(ps, p.get('ownerclan_price'), B)
        add = addimg.get(p.get('product_code'), '')
        items.append({
            'product_code': p['product_code'],
            'name': _sanitize_name(p.get('naver_product_name')),
            'category_code': p.get('category_code'),
            'cost': price['cost'],
            'orig_ship_fee': price['orig_ship_fee'],
            'ship_fee_charged': 0 if free else B,
            'target_price': price['target_price'],
            'list_price': price['list_price'],
            'discount_amount': price['discount_amount'],
            'net_margin': price['net_margin'],
            'image': _pick_image(p),
            'img_upscaled': bool(PUBLIC_MEDIA_BASE and p.get('upscaled_image_url')),
            'add_image_count': len(add.split('\n')) if add else 0,
            'has_detail': bool((p.get('detail_html') or '').strip()),
            'category_ok': bool(p.get('category_code')),
        })
    # 전체 카운트
    total = len(fetch_confirmed_products(folder_id))
    warnings = []
    if set_row.get('delivery_fee_type') != '무료' and not (set_row.get('as_phone') or set_row.get('as_guide')):
        warnings.append('A/S 전화번호/안내가 비어 있습니다 (템플릿 미사용 시 권장).')
    if not PUBLIC_MEDIA_BASE:
        warnings.append('대표이미지는 공인 CDN(image_large) 사용. 업스케일 이미지를 쓰려면 PUBLIC_MEDIA_BASE_URL 설정 필요.')
    else:
        cdn_fallback = sum(1 for it in items if not it['img_upscaled'])
        if cdn_fallback:
            warnings.append(f'업스케일 미완료 {cdn_fallback}건은 공인 CDN(image_large) 사용 (joacham 호스팅 W코드_1.jpg 없음).')
    missing_cat = [it['product_code'] for it in items if not it['category_ok']]
    if missing_cat:
        warnings.append(f'카테고리코드 없는 상품: {len(missing_cat)}건')
    return {
        'total_confirmed': total,
        'file_count': (total + BATCH_SIZE - 1) // BATCH_SIZE if total else 0,
        'batch_size': BATCH_SIZE,
        'items': items,
        'warnings': warnings,
    }


def _zip_from_products(products: list[dict], set_row: dict, store_name: str,
                       folder_id: int) -> tuple[bytes, dict]:
    """상품 리스트 → 500개 단위 xlsx → ZIP bytes."""
    if not products:
        return b'', {'total': 0, 'files': 0}
    addimg = fetch_additional_images([p['product_code'] for p in products if p.get('product_code')])
    rows = [build_row(p, set_row, addimg.get(p.get('product_code'), '')) for p in products]

    safe_store = re.sub(r'[^\w가-힣]+', '_', store_name or f'folder{folder_id}')[:40]
    zbuf = io.BytesIO()
    files = []
    with zipfile.ZipFile(zbuf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for bi in range(0, len(rows), BATCH_SIZE):
            chunk = rows[bi:bi + BATCH_SIZE]
            part = bi // BATCH_SIZE + 1
            fname = f'{safe_store}_일괄등록_{part:03d}_{len(chunk)}건.xlsx'
            zf.writestr(fname, _write_batch_xlsx(chunk))
            files.append(fname)
    return zbuf.getvalue(), {
        'total': len(rows), 'files': len(files),
        'file_names': files, 'batch_size': BATCH_SIZE,
    }


def generate_zip(folder_id: int, set_row: dict, store_name: str = '') -> tuple[bytes, dict]:
    """확정상품(컨펌이력) 전체 → 500개 단위 ZIP."""
    return _zip_from_products(fetch_confirmed_products(folder_id), set_row, store_name, folder_id)


def generate_zip_from_folder(folder_id: int, set_row: dict, store_name: str = '') -> tuple[bytes, dict]:
    """폴더 내 전체 상품 → 500개 단위 ZIP."""
    return _zip_from_products(fetch_folder_products(folder_id), set_row, store_name, folder_id)


def generate_zip_from_stage(store_folder_id: int, stage: str, set_row: dict,
                            store_name: str = '') -> tuple[bytes, dict]:
    """스토어 폴더 내 특정 단계(작업대기) 상품 → 500개 단위 ZIP."""
    return _zip_from_products(
        fetch_stage_products(store_folder_id, stage), set_row, store_name, store_folder_id)


def generate_all_queue_zip(stores: list[dict]) -> tuple[bytes, dict]:
    """여러 스토어 작업대기를 스토어별 폴더로 한 ZIP에 묶기.

    stores: [{'folder_id', 'name', 'set_row'}]  (미등록 작업대기 상품만, 500분할)
    ZIP 구조: {스토어}/{스토어}_일괄등록_001_NNN건.xlsx
    """
    zbuf = io.BytesIO()
    per_store = []
    total_rows = total_files = 0
    with zipfile.ZipFile(zbuf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for st in stores:
            fid = st['folder_id']
            products = fetch_stage_products(fid, 'queue')
            if not products:
                continue
            addimg = fetch_additional_images([p['product_code'] for p in products if p.get('product_code')])
            rows = [build_row(p, st['set_row'], addimg.get(p.get('product_code'), '')) for p in products]
            safe = re.sub(r'[^\w가-힣]+', '_', st.get('name') or f'folder{fid}')[:40]
            files = 0
            for bi in range(0, len(rows), BATCH_SIZE):
                chunk = rows[bi:bi + BATCH_SIZE]
                part = bi // BATCH_SIZE + 1
                fname = f'{safe}/{safe}_일괄등록_{part:03d}_{len(chunk)}건.xlsx'
                zf.writestr(fname, _write_batch_xlsx(chunk))
                files += 1
            per_store.append({'store': safe, 'total': len(rows), 'files': files})
            total_rows += len(rows)
            total_files += files
    return zbuf.getvalue(), {
        'stores': len(per_store), 'total': total_rows, 'files': total_files,
        'per_store': per_store, 'batch_size': BATCH_SIZE,
    }
