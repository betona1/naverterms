"""W코드 블랙리스트 서비스.

흐름:
  1. 사용자가 텍스트/엑셀로 W코드 등록 → naverdb.naver_wcode_blacklist 누적
  2. 매칭 실행 → 3 DB 에서 OR 검색
     - naverdb.naver_my_product (마이상품, 191K)
     - ads.preliminary_product (예비상품, 987K)
     - ads.ownerclan_product (오너클랜상품, 173K)
  3. 사용자가 모달에서 검토 후 체크 → 일괄 삭제
     - 안전상 기본은 naverdb.naver_my_product 행만 삭제
     - 마켓(ads.*) 행은 별도 옵션으로만 (현 단계 미구현)
"""
from __future__ import annotations

import logging
import re
from datetime import datetime

from django.db import connections

logger = logging.getLogger(__name__)

NAVERDB = 'naverdb'
ADSDB = 'ads'

WCODE_RE = re.compile(r'W[0-9A-Fa-f]{6}', re.IGNORECASE)


# ── 유틸 ─────────────────────────────────────────────────

def _normalize_code(raw: str) -> str | None:
    """입력에서 W코드 1개 추출. 'W' 빠진 hex 6자리도 보강."""
    if not raw:
        return None
    s = raw.strip().upper()
    if WCODE_RE.fullmatch(s):
        return s
    # 'W' 없이 들어온 경우
    if re.fullmatch(r'[0-9A-F]{6}', s):
        return f'W{s}'
    # 임의 문자열에서 첫 W코드만 뽑기
    m = WCODE_RE.search(s)
    return m.group(0).upper() if m else None


def _parse_codes(text: str | None) -> list[str]:
    """textarea/CSV/공백 등 어떤 구분자로 들어와도 W코드 list 로 정리 (중복 제거)."""
    if not text:
        return []
    parts = re.split(r'[\s,;|\n\r\t]+', text)
    out: list[str] = []
    seen: set = set()
    for p in parts:
        c = _normalize_code(p)
        if c and c not in seen:
            seen.add(c)
            out.append(c)
    return out


# ── 등록 / 해제 ──────────────────────────────────────────

def bulk_upsert(codes: list[str], reason: str | None = None,
                source: str = 'manual',
                reasons: dict[str, str | None] | None = None) -> dict:
    """다건 등록. 재업로드 안전:
      - 새 코드          → INSERT
      - 기존 & 사유 비어있음 → 사유 채움 (filled)
      - 기존 & 사유 있음    → 그대로 유지 (unchanged, 덮어쓰기 X)

    reasons: 코드별 사유 dict. 주어지면 우선 사용, 비어있으면 reason 파라미터 fallback.
    """
    if not codes:
        return {'ok': True, 'inserted': 0, 'filled': 0,
                'unchanged': 0, 'total': 0}
    reasons = reasons or {}

    # 1) 기존 사유 상태 조회 → 분류
    existing: dict[str, str | None] = {}
    CHUNK = 1000
    with connections[NAVERDB].cursor() as cur:
        for i in range(0, len(codes), CHUNK):
            chunk = codes[i:i + CHUNK]
            ph = ','.join(['%s'] * len(chunk))
            cur.execute(
                f"SELECT product_code, reason FROM naver_wcode_blacklist "
                f"WHERE product_code IN ({ph})",
                chunk,
            )
            for code_, rsn in cur.fetchall():
                existing[code_] = rsn

    inserted = filled = unchanged = 0
    with connections[NAVERDB].cursor() as cur:
        for code in codes:
            new_r = reasons.get(code) or reason
            if code not in existing:
                cur.execute(
                    """INSERT INTO naver_wcode_blacklist
                          (product_code, reason, source)
                       VALUES (%s, %s, %s)""",
                    [code, new_r, source],
                )
                inserted += 1
            else:
                cur_r = existing[code]
                if (cur_r is None or cur_r == '') and new_r:
                    cur.execute(
                        """UPDATE naver_wcode_blacklist
                              SET reason=%s, updated_at=NOW()
                            WHERE product_code=%s""",
                        [new_r, code],
                    )
                    filled += 1
                else:
                    unchanged += 1
    return {'ok': True, 'inserted': inserted, 'filled': filled,
            'unchanged': unchanged, 'total': len(codes)}


def delete(codes: list[str]) -> dict:
    """블랙리스트에서 지정 해제 (DB 삭제 아님 — 블랙리스트 등록 해제만)."""
    if not codes:
        return {'ok': True, 'deleted': 0}
    placeholders = ','.join(['%s'] * len(codes))
    with connections[NAVERDB].cursor() as cur:
        cur.execute(
            f"DELETE FROM naver_wcode_blacklist WHERE product_code IN ({placeholders})",
            codes,
        )
        return {'ok': True, 'deleted': cur.rowcount}


# ── 목록 ─────────────────────────────────────────────────

def list_all(search: str | None = None,
             only_unprocessed: bool = False,
             page: int = 1, per_page: int = 50) -> dict:
    where: list[str] = ['1=1']
    params: list = []
    if search:
        s = _normalize_code(search) or search.strip()
        where.append('product_code LIKE %s')
        params.append(f'%{s}%')
    if only_unprocessed:
        where.append('is_processed=0')
    where_sql = ' AND '.join(where)
    offset = max(0, (page - 1) * per_page)

    with connections[NAVERDB].cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM naver_wcode_blacklist WHERE {where_sql}",
                    params)
        total = cur.fetchone()[0]
        cur.execute(
            f"""SELECT product_code, reason, source, is_processed, processed_at,
                       matched_my, matched_pre, matched_oc, matched_at,
                       created_at, updated_at
                  FROM naver_wcode_blacklist
                 WHERE {where_sql}
                 ORDER BY is_processed ASC, created_at DESC
                 LIMIT %s OFFSET %s""",
            params + [per_page, offset],
        )
        items = []
        for r in cur.fetchall():
            items.append({
                'product_code': r[0],
                'reason': r[1],
                'source': r[2],
                'is_processed': bool(r[3]),
                'processed_at': r[4].isoformat() if r[4] else None,
                'matched_my': bool(r[5]),
                'matched_pre': bool(r[6]),
                'matched_oc': bool(r[7]),
                'matched_at': r[8].isoformat() if r[8] else None,
                'created_at': r[9].isoformat() if r[9] else None,
                'updated_at': r[10].isoformat() if r[10] else None,
            })
    return {'items': items, 'total': total, 'page': page, 'per_page': per_page}


# ── 매칭 (3-DB) ──────────────────────────────────────────

def _fetch_codes(processed: bool | None = None) -> list[str]:
    where = '1=1' if processed is None else (f'is_processed={1 if processed else 0}')
    with connections[NAVERDB].cursor() as cur:
        cur.execute(f"SELECT product_code FROM naver_wcode_blacklist WHERE {where}")
        return [r[0] for r in cur.fetchall()]


def _match_table(db: str, table: str, codes: list[str],
                 select_fields: str) -> list[dict]:
    if not codes:
        return []
    # IN (...) 가 너무 크면 청크 분할
    out: list[dict] = []
    CHUNK = 1000
    for i in range(0, len(codes), CHUNK):
        chunk = codes[i:i + CHUNK]
        ph = ','.join(['%s'] * len(chunk))
        with connections[db].cursor() as cur:
            cur.execute(
                f"SELECT {select_fields} FROM {table} WHERE product_code IN ({ph})",
                chunk,
            )
            cols = [c[0] for c in cur.description]
            for r in cur.fetchall():
                d = dict(zip(cols, r))
                # datetime → iso
                for k, v in list(d.items()):
                    if isinstance(v, datetime):
                        d[k] = v.isoformat()
                out.append(d)
    return out


def run_match(refresh_flags: bool = True) -> dict:
    """블랙리스트 W코드를 3 DB 에서 찾아 결과 반환 + matched_* 플래그 갱신."""
    codes = _fetch_codes(processed=False)
    my_rows = _match_table(
        NAVERDB, 'naver_my_product', codes,
        'id, product_code, folder_id, product_name, naver_product_name, '
        'category_name, brand, image_small, sale_status, updated_at',
    )
    pre_rows = _match_table(
        ADSDB, 'preliminary_product', codes,
        'id, product_code, supplier_code, seller_code1, '
        'product_name, market_product_name, category_name, '
        'ownerclan_price, market_price, updated_at',
    )
    oc_rows = _match_table(
        ADSDB, 'ownerclan_product', codes,
        'id, product_code, seller_code1, '
        'product_name, market_product_name, category_name, '
        'sale_status, is_synced, ownerclan_price, market_price, updated_at',
    )

    if refresh_flags and codes:
        my_set = {r['product_code'] for r in my_rows}
        pre_set = {r['product_code'] for r in pre_rows}
        oc_set = {r['product_code'] for r in oc_rows}
        with connections[NAVERDB].cursor() as cur:
            for code in codes:
                cur.execute(
                    """UPDATE naver_wcode_blacklist
                          SET matched_my=%s, matched_pre=%s, matched_oc=%s, matched_at=NOW()
                        WHERE product_code=%s""",
                    [int(code in my_set), int(code in pre_set), int(code in oc_set), code],
                )

    return {
        'ok': True,
        'total_codes': len(codes),
        'my': {'count': len(my_rows), 'items': my_rows},
        'preliminary': {'count': len(pre_rows), 'items': pre_rows},
        'ownerclan': {'count': len(oc_rows), 'items': oc_rows},
    }


# ── 일괄 삭제 (체크 후 액션) ──────────────────────────────

def process_delete(codes: list[str], targets: list[str]) -> dict:
    """체크된 W코드를 실제 삭제.

    targets:
      - 'my' → naverdb.naver_my_product 행 DELETE
      - 'pre' → ads.preliminary_product 행 DELETE
      - 'oc' → ads.ownerclan_product 행 DELETE

    안전 기본: targets=['my'] (마이상품만). 마켓 API 호출은 하지 않음.
    삭제 완료 W코드는 naver_wcode_blacklist.is_processed=1.
    """
    if not codes:
        return {'ok': True, 'deleted': {}}
    if not targets:
        return {'ok': False, 'error': 'targets required'}
    valid_targets = {'my', 'pre', 'oc'}
    targets = [t for t in targets if t in valid_targets]
    if not targets:
        return {'ok': False, 'error': 'invalid targets'}

    deleted: dict = {'my': 0, 'pre': 0, 'oc': 0}
    placeholders = ','.join(['%s'] * len(codes))

    if 'my' in targets:
        with connections[NAVERDB].cursor() as cur:
            cur.execute(
                f"DELETE FROM naver_my_product WHERE product_code IN ({placeholders})",
                codes,
            )
            deleted['my'] = cur.rowcount

    if 'pre' in targets:
        with connections[ADSDB].cursor() as cur:
            cur.execute(
                f"DELETE FROM preliminary_product WHERE product_code IN ({placeholders})",
                codes,
            )
            deleted['pre'] = cur.rowcount

    if 'oc' in targets:
        with connections[ADSDB].cursor() as cur:
            cur.execute(
                f"DELETE FROM ownerclan_product WHERE product_code IN ({placeholders})",
                codes,
            )
            deleted['oc'] = cur.rowcount

    # 처리 완료 마킹
    with connections[NAVERDB].cursor() as cur:
        cur.execute(
            f"""UPDATE naver_wcode_blacklist
                   SET is_processed=1, processed_at=NOW()
                 WHERE product_code IN ({placeholders})""",
            codes,
        )

    return {'ok': True, 'deleted': deleted, 'processed': len(codes), 'targets': targets}


# ── 가드 (재등록 차단) ────────────────────────────────────

def filter_blacklisted(codes: list[str] | set[str]) -> set[str]:
    """주어진 W코드 중 블랙리스트에 등록된 것만 set 으로 반환.

    sync/import 진입점에서 사전 필터로 사용.
    """
    if not codes:
        return set()
    codes = list(set(codes))
    out: set[str] = set()
    CHUNK = 1000
    with connections[NAVERDB].cursor() as cur:
        for i in range(0, len(codes), CHUNK):
            chunk = codes[i:i + CHUNK]
            ph = ','.join(['%s'] * len(chunk))
            cur.execute(
                f"SELECT product_code FROM naver_wcode_blacklist "
                f"WHERE product_code IN ({ph})",
                chunk,
            )
            for (c,) in cur.fetchall():
                out.add(c)
    return out


def enforce_blacklist() -> dict:
    """블랙리스트의 모든 W코드를 3개 타겟 테이블에서 일괄 DELETE.

    외부(11st/ai100) sync 가 다시 등록한 행을 일괄 정리할 때 사용.
    is_processed 와 무관하게 블랙리스트 전체에 적용.
    """
    codes = _fetch_codes(processed=None)
    deleted = {'my': 0, 'pre': 0, 'oc': 0}
    if not codes:
        return {'ok': True, 'deleted': deleted, 'total_codes': 0}
    CHUNK = 1000
    for i in range(0, len(codes), CHUNK):
        chunk = codes[i:i + CHUNK]
        ph = ','.join(['%s'] * len(chunk))
        with connections[NAVERDB].cursor() as cur:
            cur.execute(
                f"DELETE FROM naver_my_product WHERE product_code IN ({ph})",
                chunk,
            )
            deleted['my'] += cur.rowcount
        with connections[ADSDB].cursor() as cur:
            cur.execute(
                f"DELETE FROM preliminary_product WHERE product_code IN ({ph})",
                chunk,
            )
            deleted['pre'] += cur.rowcount
            cur.execute(
                f"DELETE FROM ownerclan_product WHERE product_code IN ({ph})",
                chunk,
            )
            deleted['oc'] += cur.rowcount
    return {'ok': True, 'deleted': deleted, 'total_codes': len(codes)}


# ── 엑셀 업로드 ─────────────────────────────────────────

def parse_excel(file_obj) -> dict:
    """업로드된 엑셀(.xlsx) 에서 W코드 + 사유 추출.

    1) 헤더에서 W코드 컬럼명 매칭 (상품번호/상품코드/W코드/product_code/code 등)
       + 사유 컬럼명 매칭 (위반사유/사유/reason/비고)
       → 같은 행의 W코드와 사유를 짝지음
    2) 헤더 매칭 실패 시 모든 셀에서 W-패턴 스캔 (사유 없음)

    return: { ok, codes: [...], count, reasons: {code: reason}, with_reason: N }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {'ok': False, 'error': 'openpyxl not installed'}
    try:
        wb = load_workbook(file_obj, read_only=True)
    except Exception as e:  # noqa: BLE001
        return {'ok': False, 'error': f'엑셀 로드 실패: {e}'}

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        return {'ok': False, 'error': '빈 파일'}

    header_norm = [str(h or '').strip().lower().replace(' ', '') for h in header]

    def _find_col(needles: tuple[str, ...]) -> int | None:
        for n in needles:
            if n in header_norm:
                return header_norm.index(n)
        return None

    target_idx = _find_col(('w코드', 'wcode', 'product_code', 'productcode',
                            'code', '상품코드', '상품번호', 'productno', 'product_no'))
    reason_idx = _find_col(('위반사유', '사유', 'reason', '비고', 'note', 'memo', '메모'))

    codes: list[str] = []
    reasons: dict[str, str] = {}
    seen: set = set()

    def _extract_wcodes(value) -> list[str]:
        if value is None:
            return []
        return [m.upper() for m in WCODE_RE.findall(str(value))]

    def _norm_reason(value) -> str | None:
        if value is None:
            return None
        s = str(value).strip()
        return s or None

    def _record(code: str, reason: str | None) -> None:
        if code not in seen:
            seen.add(code); codes.append(code)
        # 같은 W코드가 여러 행에 나오면 첫 번째 비어있지 않은 사유 채택
        if reason and code not in reasons:
            reasons[code] = reason

    def _walk(row_values) -> None:
        if target_idx is not None and target_idx < len(row_values):
            r = _norm_reason(row_values[reason_idx]) if (
                reason_idx is not None and reason_idx < len(row_values)
            ) else None
            for c in _extract_wcodes(row_values[target_idx]):
                _record(c, r)
        else:
            # 헤더 매칭 실패 → 모든 셀에서 W-패턴 추출
            for cell in row_values:
                for c in _extract_wcodes(cell):
                    _record(c, None)

    _walk(header)  # 헤더 행이 사실 데이터일 수 있음
    for row in rows:
        _walk(row)

    wb.close()
    return {
        'ok': True,
        'codes': codes,
        'count': len(codes),
        'reasons': reasons,
        'with_reason': len(reasons),
    }
