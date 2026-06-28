"""상품 일괄등록 단계 — 상태(register_stage) + 등록완료 관리.

물리 하위폴더 대신 naver_my_product.register_stage 상태값 사용:
  · NULL        — 단계 없음 (스토어 폴더에만 존재)
  · 'candidate' — 등록후보 (수동 선별)
  · 'queue'     — 작업대기 (엑셀 생성 대상)
등록완료 = naver_my_product.registered=1.
스토어 클릭 후 탭(전체/등록후보/작업대기)으로 필터링 — 폴더 안 늘어남.
"""
from __future__ import annotations

from django.db import connections

NAVERDB = 'naverdb'
MYPRODUCT_DB = 'myproduct'
ADS_DB = 'ads'
OWIMAGE_DB = 'owimage'

STAGE_CANDIDATE = 'candidate'   # 등록후보
STAGE_QUEUE = 'queue'           # 작업대기
VALID_STAGES = (STAGE_CANDIDATE, STAGE_QUEUE)


def list_store_stages() -> list[dict]:
    """스토어 폴더별 단계 카운트 (전체/등록후보/작업대기/등록완료)."""
    with connections[NAVERDB].cursor() as cur:
        cur.execute(
            """
            SELECT f.id, f.store_id, f.name, f.color,
                   COALESCE(c.total, 0)      AS total,
                   COALESCE(c.candidate, 0)  AS candidate_count,
                   COALESCE(c.queue, 0)      AS queue_count,
                   COALESCE(c.registered, 0) AS registered_count,
                   COALESCE(c.staged_pending, 0) AS staged_pending
              FROM naver_my_product_folder f
              LEFT JOIN (
                SELECT folder_id,
                       COUNT(*) total,
                       SUM(register_stage='candidate') candidate,
                       SUM(register_stage='queue') queue,
                       SUM(registered) registered,
                       SUM(register_stage IS NOT NULL AND registered=0) staged_pending
                  FROM naver_my_product GROUP BY folder_id
              ) c ON c.folder_id = f.id
             WHERE f.kind='store' OR f.kind IS NULL
             ORDER BY f.is_system DESC, f.sort_order, f.id
            """)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    for r in rows:
        for k in ('total', 'candidate_count', 'queue_count', 'registered_count', 'staged_pending'):
            r[k] = int(r.get(k) or 0)

    # 스토어별 현재 네이버 등록수 (smartstore_product) — 970 한도 계산용
    store_ids = [r['store_id'] for r in rows if r.get('store_id')]
    naver_cnt: dict = {}
    if store_ids:
        with connections[MYPRODUCT_DB].cursor() as cur:
            ph = ','.join(['%s'] * len(store_ids))
            cur.execute(
                f"SELECT store_id, COUNT(*) FROM smartstore_product "
                f"WHERE store_id IN ({ph}) GROUP BY store_id", store_ids)
            naver_cnt = {sid: int(c) for sid, c in cur.fetchall()}
    for r in rows:
        r['naver_count'] = naver_cnt.get(r.get('store_id'), 0)
    return rows


def _chunked_in(cur, sql_tmpl: str, codes: list, prefix: list = None, size: int = 500):
    """IN 절 배치 실행 → 전체 row 합산 반환."""
    out = []
    prefix = prefix or []
    for i in range(0, len(codes), size):
        chunk = codes[i:i + size]
        ph = ','.join(['%s'] * len(chunk))
        cur.execute(sql_tmpl.format(ph=ph), prefix + chunk)
        out.extend(cur.fetchall())
    return out


def score_candidates(folder_id: int, rows: list, weights: dict | None = None) -> dict:
    """후보(rows=[(id,code,category_code)]) 종합 점수.

    badimage(위험) 제외 + 내매출 우선 가중 종합:
      score = w_my·내매출 + w_cat·카테고리시장성 + w_11·11번가 + w_img·good이미지
    반환: {scored:[(id,score,detail)], excluded_bad:[codes]}
    """
    w = {'my': 0.5, 'cat': 0.25, 'eleven': 0.15, 'img': 0.10}
    if weights:
        w.update(weights)
    codes = [r[1] for r in rows]
    cats = sorted({r[2] for r in rows if r[2]})

    # 스토어 store_id
    with connections[NAVERDB].cursor() as cur:
        cur.execute("SELECT store_id FROM naver_my_product_folder WHERE id=%s", [int(folder_id)])
        srow = cur.fetchone()
    store_id = srow[0] if srow else None

    my_sales, cat_market, eleven, good, bad, live = {}, {}, {}, set(), set(), set()
    # 내매출 + 카테고리 시장성 + 라이브 등록 여부 (myproduct)
    with connections[MYPRODUCT_DB].cursor() as cur:
        if store_id and codes:
            for w_code, amt in _chunked_in(
                cur, "SELECT seller_management_code, all_order_amount FROM smartstore_product "
                     "WHERE store_id=%s AND seller_management_code IN ({ph})", codes, [int(store_id)]):
                my_sales[w_code] = float(amt or 0)
            # 이미 네이버에 등록된 W코드 (중복등록 방지 — 제외)
            for (w_code,) in _chunked_in(
                cur, "SELECT seller_management_code FROM smartstore_product "
                     "WHERE store_id=%s AND seller_management_code IN ({ph})", codes, [int(store_id)]):
                live.add(w_code)
        if cats:
            for leaf, amt in _chunked_in(
                cur, "SELECT leaf_category_id, SUM(all_order_amount) FROM smartstore_product "
                     "WHERE leaf_category_id IN ({ph}) GROUP BY leaf_category_id", cats):
                cat_market[str(leaf)] = float(amt or 0)
    # 11번가 매출 (ads)
    try:
        with connections[ADS_DB].cursor() as cur:
            for w_code, amt in _chunked_in(
                cur, "SELECT w_code, total_amount FROM eleven_sales_w_stats WHERE w_code IN ({ph})", codes):
                eleven[w_code] = float(amt or 0)
    except Exception:
        pass
    # good/bad 이미지 (owimage)
    try:
        with connections[OWIMAGE_DB].cursor() as cur:
            for (w_code,) in _chunked_in(cur, "SELECT DISTINCT w_code FROM goodimage WHERE w_code IN ({ph})", codes):
                good.add(w_code)
            for (w_code,) in _chunked_in(cur, "SELECT DISTINCT w_code FROM badimage WHERE w_code IN ({ph})", codes):
                bad.add(w_code)
    except Exception:
        pass

    def nz(d):  # 정규화 분모 (최댓값)
        m = max(d.values()) if d else 0
        return m or 1.0

    my_max, cat_max, el_max = nz(my_sales), nz(cat_market), nz(eleven)
    scored = []
    excluded_live = 0
    for pid, code, cat in rows:
        if code in bad:
            continue   # 위험(badimage) — 제외
        if code in live:
            excluded_live += 1
            continue   # 이미 라이브 등록 — 중복방지 제외
        s_my = my_sales.get(code, 0) / my_max
        s_cat = cat_market.get(str(cat), 0) / cat_max
        s_el = eleven.get(code, 0) / el_max
        s_img = 1.0 if code in good else 0.0
        score = w['my'] * s_my + w['cat'] * s_cat + w['eleven'] * s_el + w['img'] * s_img
        scored.append((pid, score, {
            'code': code, 'my_sales': my_sales.get(code, 0), 'cat_market': cat_market.get(str(cat), 0),
            'eleven': eleven.get(code, 0), 'good_img': code in good,
        }))
    scored.sort(key=lambda x: x[1], reverse=True)
    return {'scored': scored, 'excluded_bad': [c for c in codes if c in bad],
            'excluded_live': excluded_live, 'good_count': len(good)}


def select_ids(folder_id: int, stage: str | None = None, registered: int | None = None,
               category_code: str | None = None, search: str | None = None,
               recommend: bool = False, limit: int = 2000) -> dict:
    """필터에 맞는 상품 id 반환 (전체선택 / AI추천 체크).

    recommend=True: 미등록·업스케일·카테고리 보유 적격풀에서 badimage(위험) 제외 후,
      종합점수(내매출 우선 + 카테고리시장성 + 11번가 + good이미지) 상위 limit.
    """
    where = ['p.folder_id=%s', "p.naver_product_name<>''"]
    params: list = [int(folder_id)]
    if stage == 'none':
        where.append('p.register_stage IS NULL')
    elif stage in VALID_STAGES:
        where.append('p.register_stage=%s'); params.append(stage)
    if registered is not None:
        where.append('p.registered=%s'); params.append(int(registered))
    if category_code:
        where.append('p.category_code=%s'); params.append(str(category_code))
    if search:
        where.append('(p.product_code LIKE %s OR p.naver_product_name LIKE %s)')
        params += [f'%{search}%', f'%{search}%']
    if recommend:
        where += ['p.register_stage IS NULL', 'p.registered=0',
                  "p.upscaled_image_url<>''", "p.category_code<>''"]

    where_sql = ' AND '.join(where)
    sql_limit = 100000 if recommend else int(limit)
    sql = (f"SELECT p.id, p.product_code, p.category_code, p.naver_product_name FROM naver_my_product p "
           f"WHERE {where_sql} ORDER BY p.id LIMIT %s")
    with connections[NAVERDB].cursor() as cur:
        cur.execute(sql, params + [sql_limit])
        raw = cur.fetchall()
    # 불량 상품명(AI 안내문 누출 등) 제외
    if recommend:
        from . import bad_name_service as bn
        rows = [(i, c, cat) for i, c, cat, nm in raw if not bn.classify(nm)]
    else:
        rows = [(i, c, cat) for i, c, cat, nm in raw]

    if recommend and rows:
        res = score_candidates(folder_id, rows)
        scored = res['scored'][:int(limit)]
        return {'ids': [s[0] for s in scored], 'count': len(scored),
                'excluded_bad': len(res['excluded_bad']), 'eligible': len(rows)}

    return {'ids': [r[0] for r in rows], 'count': len(rows)}


def set_stage(ids: list[int], stage: str | None) -> dict:
    """상품들의 단계 설정 (None=단계해제/되돌리기, 'candidate', 'queue')."""
    if not ids:
        return {'ok': False, 'error': 'ids 없음'}
    if stage is not None and stage not in VALID_STAGES:
        return {'ok': False, 'error': f'잘못된 stage: {stage}'}
    with connections[NAVERDB].cursor() as cur:
        ph = ','.join(['%s'] * len(ids))
        cur.execute(
            f"UPDATE naver_my_product SET register_stage=%s WHERE id IN ({ph})",
            [stage] + [int(x) for x in ids])
        n = cur.rowcount
    return {'ok': True, 'updated': n, 'stage': stage}


def mark_registered(ids: list[int], value: int = 1) -> dict:
    """상품등록완료 표기/해제."""
    if not ids:
        return {'ok': False, 'error': 'ids 없음'}
    ts = 'NOW()' if value else 'NULL'
    with connections[NAVERDB].cursor() as cur:
        ph = ','.join(['%s'] * len(ids))
        cur.execute(
            f"UPDATE naver_my_product SET registered=%s, registered_at={ts} WHERE id IN ({ph})",
            [1 if value else 0] + [int(x) for x in ids])
        n = cur.rowcount
    return {'ok': True, 'updated': n, 'registered': bool(value)}


STATUS_LABEL = {
    'SALE': '판매중', 'OUTOFSTOCK': '품절', 'SUSPENSION': '판매중지',
    'PROHIBITION': '판매금지', 'CLOSE': '판매종료', 'WAIT': '판매대기',
    'DELETE': '삭제됨',
}


def inspect_registration(store_folder_id: int, run_sync: bool = True) -> dict:
    """전체 올린상품 동기화 + 특이사항 점검 + 로그.

    작업대기(queue) 상품을 네이버 동기화 결과와 대조:
      · 미발견(등록실패 추정)  · 판매중지/금지/종료/품절  · 가격 불일치
    found 는 register_verified=1, registered=1 갱신. 로그 파일도 저장.
    """
    import os
    from datetime import datetime
    from . import register_set_service as rss

    store_folder_id = int(store_folder_id)
    with connections[NAVERDB].cursor() as cur:
        cur.execute("SELECT store_id, name FROM naver_my_product_folder WHERE id=%s", [store_folder_id])
        row = cur.fetchone()
    if not row or not row[0]:
        return {'ok': False, 'error': '스토어 폴더/store_id 없음'}
    store_id, store_name = int(row[0]), row[1]
    set_row = rss.get_set(store_folder_id) or {}

    synced = False
    if run_sync:
        try:
            from . import smartstore_product_service as sps
            sps.sync_products(store_id)
            synced = True
        except Exception as e:
            return {'ok': False, 'error': f'동기화 실패: {str(e)[:200]}'}

    # 작업대기 상품 (원가/배송비 포함 — 가격 대조용)
    with connections[NAVERDB].cursor() as cur:
        cur.execute(
            "SELECT id, product_code, ownerclan_price, shipping_fee, naver_product_name "
            "FROM naver_my_product WHERE folder_id=%s AND register_stage='queue' AND product_code<>''",
            [store_folder_id])
        qrows = cur.fetchall()
    if not qrows:
        return {'ok': True, 'queue_count': 0, 'synced': synced, 'anomalies': [], 'log': []}

    codes = [r[1] for r in qrows]
    # 네이버 동기화 결과 조회
    found: dict = {}
    with connections[MYPRODUCT_DB].cursor() as cur:
        for i in range(0, len(codes), 500):
            chunk = codes[i:i + 500]
            ph = ','.join(['%s'] * len(chunk))
            cur.execute(
                f"SELECT seller_management_code, status_type, sale_price, whole_category_name "
                f"FROM smartstore_product WHERE store_id=%s AND seller_management_code IN ({ph})",
                [store_id] + chunk)
            for code, st, sp, cat in cur.fetchall():
                found[code] = {'status': st, 'sale_price': sp, 'category': cat}

    def expected_price(cost, ship):
        B = int(ship or 0)
        free = int(set_row.get('free_shipping') or 0) == 1 or B <= 0
        ps = dict(set_row)
        if not free:
            ps['set_ship_fee'] = B
        return rss.compute_price(ps, cost, B)['list_price'] if set_row else None

    anomalies = []
    verified_ids = []
    cnt = {'found': 0, 'missing': 0, 'soldout': 0, 'stopped': 0, 'price_diff': 0}
    for pid, code, cost, ship, pname in qrows:
        info = found.get(code)
        if not info:
            cnt['missing'] += 1
            anomalies.append({'code': code, 'type': '미발견', 'detail': '네이버에서 찾을 수 없음 (등록 실패 추정)', 'name': pname})
            continue
        cnt['found'] += 1
        verified_ids.append(pid)
        st = info['status']
        if st in ('SUSPENSION', 'PROHIBITION', 'CLOSE', 'DELETE'):
            cnt['stopped'] += 1
            anomalies.append({'code': code, 'type': STATUS_LABEL.get(st, st), 'detail': f'판매상태={st}', 'name': pname})
        elif st == 'OUTOFSTOCK':
            cnt['soldout'] += 1
            anomalies.append({'code': code, 'type': '품절', 'detail': '재고 0', 'name': pname})
        exp = expected_price(cost, ship)
        if exp and info['sale_price'] and int(info['sale_price']) != int(exp):
            cnt['price_diff'] += 1
            anomalies.append({'code': code, 'type': '가격불일치',
                              'detail': f"네이버 {int(info['sale_price']):,} ≠ 예상 {int(exp):,}", 'name': pname})

    # found 검증/등록완료 갱신
    with connections[NAVERDB].cursor() as cur:
        if verified_ids:
            ph = ','.join(['%s'] * len(verified_ids))
            cur.execute(
                f"UPDATE naver_my_product SET register_verified=1, registered=1, "
                f"register_checked_at=NOW(), registered_at=COALESCE(registered_at,NOW()) "
                f"WHERE id IN ({ph})", verified_ids)
        missing_ids = [r[0] for r in qrows if r[1] not in found]
        if missing_ids:
            ph = ','.join(['%s'] * len(missing_ids))
            cur.execute(
                f"UPDATE naver_my_product SET register_verified=0, register_checked_at=NOW() "
                f"WHERE id IN ({ph})", missing_ids)

    # 로그 작성
    ts = datetime.now()
    log = [
        f"[{ts:%Y-%m-%d %H:%M:%S}] 동기화+점검 — {store_name} (store_id={store_id})",
        f"작업대기 {len(qrows)}건 | 동기화 {'O' if synced else 'X'}",
        f"등록확인 {cnt['found']} / 미발견 {cnt['missing']} / 품절 {cnt['soldout']} / 판매중지등 {cnt['stopped']} / 가격불일치 {cnt['price_diff']}",
        "─" * 50,
    ]
    for a in anomalies:
        log.append(f"  [{a['type']}] {a['code']} {a['detail']} — {(a['name'] or '')[:30]}")
    if not anomalies:
        log.append("  특이사항 없음 ✓")

    # 로그 파일 저장
    log_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'exports')
    os.makedirs(log_dir, exist_ok=True)
    fname = f"register_inspect_{store_name}_{ts:%Y%m%d_%H%M%S}.log"
    fpath = os.path.join(log_dir, fname)
    try:
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(log))
    except Exception:
        fname = None

    return {
        'ok': True, 'synced': synced, 'queue_count': len(qrows),
        'counts': cnt, 'anomalies': anomalies, 'log': log,
        'log_file': fname,
    }


def parse_error_excel(file_bytes: bytes) -> dict:
    """네이버 실패 엑셀 파싱 → 실패 W코드 + 사유.

    실패 엑셀: 1행 섹션, 2행 헤더(1열='실패사유', 2열='판매자 상품코드'...),
    3행~ 데이터. 가이드행(예시 SSKR_2021 등)은 제외.
    """
    import io
    import re
    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(file_bytes)).active
    # 헤더(2행) → 컬럼 인덱스
    nmap = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(2, col).value
        if v is not None:
            nmap[re.sub(r'\s+', '', str(v))] = col
    code_col = nmap.get('판매자상품코드')
    reason_col = nmap.get('실패사유') or 1
    failed = []
    summary: dict = {}
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, code_col).value if code_col else None
        code = str(code).strip() if code else ''
        reason = str(ws.cell(r, reason_col).value or '').strip()
        # 가이드/예시행 제외
        if not code or code in ('비필수', '필수', 'SSKR_2021') or '최대' in code or len(code) > 20:
            continue
        lines = [l.strip() for l in reason.split('\n') if l.strip()]
        failed.append({'code': code, 'reasons': lines})
        for l in lines:
            key = l[:80]
            summary[key] = summary.get(key, 0) + 1
    return {
        'failed': failed,
        'failed_codes': [f['code'] for f in failed],
        'failed_count': len(failed),
        'reason_summary': [{'reason': k, 'count': v} for k, v in
                           sorted(summary.items(), key=lambda x: -x[1])],
    }


def apply_failure_result(store_folder_id: int, failed_codes: list[str],
                         mark_success: bool = True) -> dict:
    """업로드 결과 반영: 작업대기 미등록 중 실패코드 외 = 성공 → 등록완료 표기.
    실패코드는 작업대기에 미등록으로 유지(수정 후 재생성 대상).
    """
    store_folder_id = int(store_folder_id)
    fset = set(failed_codes or [])
    with connections[NAVERDB].cursor() as cur:
        cur.execute(
            "SELECT id, product_code FROM naver_my_product "
            "WHERE folder_id=%s AND register_stage='queue' AND registered=0",
            [store_folder_id])
        rows = cur.fetchall()
    success_ids = [i for i, code in rows if code not in fset]
    failed_in_queue = [code for i, code in rows if code in fset]
    if mark_success and success_ids:
        mark_registered(success_ids, 1)
    return {
        'ok': True,
        'queue_pending': len(rows),
        'success_marked': len(success_ids) if mark_success else 0,
        'failed_kept': len(failed_in_queue),
    }


def verify_registration(store_folder_id: int, run_sync: bool = False) -> dict:
    """작업대기(stage='queue') 상품이 실제 네이버에 등록됐는지 검증(동기화).

    smartstore_product(myproduct).seller_management_code 와 작업대기 W코드 매칭.
    run_sync=True 면 커머스 API 로 먼저 동기화 후 검증.
    """
    store_folder_id = int(store_folder_id)
    with connections[NAVERDB].cursor() as cur:
        cur.execute("SELECT store_id FROM naver_my_product_folder WHERE id=%s", [store_folder_id])
        row = cur.fetchone()
    if not row or not row[0]:
        return {'ok': False, 'error': '스토어 폴더/연결된 store_id 없음'}
    store_id = int(row[0])

    with connections[NAVERDB].cursor() as cur:
        cur.execute(
            "SELECT id, product_code FROM naver_my_product "
            "WHERE folder_id=%s AND register_stage='queue' AND product_code<>''",
            [store_folder_id])
        qrows = cur.fetchall()
    if not qrows:
        return {'ok': True, 'queue_count': 0, 'verified': 0, 'missing': 0, 'synced': False}
    wcodes = [r[1] for r in qrows]

    synced = False
    if run_sync:
        try:
            from . import smartstore_product_service as sps
            sps.sync_products(store_id)
            synced = True
        except Exception as e:
            return {'ok': False, 'error': f'동기화 실패: {str(e)[:200]}'}

    present = set()
    with connections[MYPRODUCT_DB].cursor() as cur:
        for i in range(0, len(wcodes), 500):
            chunk = wcodes[i:i + 500]
            ph = ','.join(['%s'] * len(chunk))
            cur.execute(
                f"SELECT seller_management_code FROM smartstore_product "
                f"WHERE store_id=%s AND seller_management_code IN ({ph})",
                [store_id] + chunk)
            present.update(r[0] for r in cur.fetchall())

    verified_ids = [r[0] for r in qrows if r[1] in present]
    unverified_ids = [r[0] for r in qrows if r[1] not in present]
    missing = [r[1] for r in qrows if r[1] not in present]

    with connections[NAVERDB].cursor() as cur:
        if verified_ids:
            ph = ','.join(['%s'] * len(verified_ids))
            cur.execute(
                f"UPDATE naver_my_product SET register_verified=1, registered=1, "
                f"register_checked_at=NOW(), registered_at=COALESCE(registered_at,NOW()) "
                f"WHERE id IN ({ph})", verified_ids)
        if unverified_ids:
            ph = ','.join(['%s'] * len(unverified_ids))
            cur.execute(
                f"UPDATE naver_my_product SET register_verified=0, register_checked_at=NOW() "
                f"WHERE id IN ({ph})", unverified_ids)

    return {
        'ok': True, 'queue_count': len(qrows), 'verified': len(verified_ids),
        'missing': len(missing), 'missing_codes': missing[:50], 'synced': synced,
    }
